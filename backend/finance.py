from __future__ import annotations

import io
import json
import mimetypes
import re
import secrets
import sqlite3
import time
from datetime import date
from http import HTTPStatus
from pathlib import Path
from sqlite_config import configure_connection

from auth import (
    user_can_manage_finances,
    user_can_pay_invoices,
    user_can_view_finances,
    user_has_any_role,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"
DOCUMENTS_DIR = DATA_DIR / "documents"
DB_PATH = DATA_DIR / "pmbi.sqlite3"
TODAY_ISO = date.today().isoformat()
MAX_UPLOAD_BYTES = 25 * 1024 * 1024
FINANCE_INVOICE_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".webp", ".xlsx", ".xls"}
FINANCE_EXCEL_PARSE_ERROR = "\u041e\u0448\u0438\u0431\u043a\u0430: \u0424\u043e\u0440\u043c\u0430\u0442 \u0444\u0430\u0439\u043b\u0430 \u043d\u0435 \u0441\u043e\u043e\u0442\u0432\u0435\u0442\u0441\u0442\u0432\u0443\u0435\u0442 \u0448\u0430\u0431\u043b\u043e\u043d\u0443 \u043e\u0442\u0447\u0435\u0442\u0430"
FINANCE_EXCEL_TEMPLATE_CELLS = {
    "category": "B2",
    "counterparty_name": "B3",
    "amount": "B4",
}


def now_ts() -> int:
    return int(time.time())


def db() -> sqlite3.Connection:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(DB_PATH)
    return configure_connection(connection)


def parse_path_int(path: str, index: int) -> int | None:
    parts = path.strip("/").split("/")
    try:
        return int(parts[index])
    except (IndexError, TypeError, ValueError):
        return None


def create_audit(
    con: sqlite3.Connection,
    user_id: int | None,
    action: str,
    entity: str | None = None,
    entity_id: int | None = None,
    payload: dict | None = None,
) -> None:
    con.execute(
        """
        INSERT INTO audit_log (user_id, action, entity, entity_id, payload, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (user_id, action, entity, entity_id, json.dumps(payload or {}, ensure_ascii=False), now_ts()),
    )


def sanitize_filename(name: str) -> str:
    cleaned = re.sub(r"[^\w.\- ]+", "_", name or "").strip().strip(".")
    return cleaned or "document"


def project_documents_dir(project_id: int) -> Path:
    path = DOCUMENTS_DIR / f"project_{project_id}"
    path.mkdir(parents=True, exist_ok=True)
    return path


def document_extension(name: str) -> str:
    return Path(name or "").suffix.lower()[:16]


class FinanceExcelParseError(ValueError):
    pass


def excel_cell_ref_to_indexes(cell_ref: str) -> tuple[int, int]:
    match = re.match(r"^([A-Za-z]+)([1-9][0-9]*)$", str(cell_ref or "").strip())
    if not match:
        raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR)
    col = 0
    for char in match.group(1).upper():
        col = col * 26 + (ord(char) - ord("A") + 1)
    return int(match.group(2)) - 1, col - 1


def clean_excel_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def parse_excel_amount(value: object) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_excel_value(value)
    text = re.sub(r"[^\d,.\-]+", "", text).replace(",", ".")
    if not text:
        raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR)
    try:
        return float(text)
    except ValueError as error:
        raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR) from error


def normalize_excel_label(value: object) -> str:
    """Normalize labels without dropping Cyrillic characters."""
    return "".join(char for char in clean_excel_value(value).lower() if char.isalnum())


def labeled_excel_value(rows: list[list[object]], labels: set[str]) -> object:
    normalized_labels = {normalize_excel_label(label) for label in labels}
    for row_index, row in enumerate(rows):
        for col_index, value in enumerate(row):
            label = normalize_excel_label(value)
            if not label or not any(expected in label for expected in normalized_labels):
                continue
            for next_col in range(col_index + 1, min(len(row), col_index + 5)):
                if clean_excel_value(row[next_col]):
                    return row[next_col]
            if row_index + 1 < len(rows) and col_index < len(rows[row_index + 1]):
                below = rows[row_index + 1][col_index]
                if clean_excel_value(below):
                    return below
    return None


def _parse_finance_rows_for_pdf(rows: list[list[object]]) -> dict[str, object]:
    """Apply the same label/value rules to rows returned by the PDF adapter."""
    category = clean_excel_value(
        labeled_excel_value(
            rows,
            {
                "\u043d\u0430\u0438\u043c\u0435\u043d\u043e\u0432\u0430\u043d\u0438\u0435",
                "\u043d\u0430\u0437\u043d\u0430\u0447\u0435\u043d\u0438\u0435",
                "\u0441\u0447\u0435\u0442",
                "\u0441\u0447\u0451\u0442",
            },
        )
    )
    counterparty = clean_excel_value(
        labeled_excel_value(
            rows,
            {
                "\u043a\u043e\u043d\u0442\u0440\u0430\u0433\u0435\u043d\u0442",
                "\u043f\u043e\u0441\u0442\u0430\u0432\u0449\u0438\u043a",
                "\u043f\u043e\u0434\u0440\u044f\u0434\u0447\u0438\u043a",
            },
        )
    )
    amount_value = labeled_excel_value(
        rows,
        {
            "\u0438\u0442\u043e\u0433\u043e\u0432\u0430\u044f \u0441\u0443\u043c\u043c\u0430",
            "\u0438\u0442\u043e\u0433\u043e",
            "\u0441\u0443\u043c\u043c\u0430 \u043a \u043e\u043f\u043b\u0430\u0442\u0435",
            "\u0432\u0441\u0435\u0433\u043e",
        },
    )
    amount = parse_excel_amount(amount_value)
    if not category or not counterparty or amount <= 0:
        raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR)
    return {"category": category, "counterparty_name": counterparty, "amount": amount}


def parse_finance_excel_invoice(raw: bytes, file_ext: str) -> dict[str, object]:
    if file_ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR) from error
        try:
            workbook = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
            sheet = workbook.active
            cell_values = {
                key: sheet[cell_ref].value
                for key, cell_ref in FINANCE_EXCEL_TEMPLATE_CELLS.items()
            }
            rows = [list(row) for row in sheet.iter_rows(min_row=1, max_row=40, max_col=12, values_only=True)]
            workbook.close()
        except Exception as error:
            raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR) from error
    elif file_ext == ".xls":
        try:
            import xlrd
        except ImportError as error:
            raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR) from error
        try:
            book = xlrd.open_workbook(file_contents=raw)
            sheet = book.sheet_by_index(0)
            cell_values = {}
            for key, cell_ref in FINANCE_EXCEL_TEMPLATE_CELLS.items():
                row_index, col_index = excel_cell_ref_to_indexes(cell_ref)
                cell_values[key] = sheet.cell_value(row_index, col_index) if row_index < sheet.nrows and col_index < sheet.ncols else None
            rows = [
                [sheet.cell_value(row_index, col_index) for col_index in range(min(sheet.ncols, 12))]
                for row_index in range(min(sheet.nrows, 40))
            ]
        except Exception as error:
            raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR) from error
    else:
        raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR)

    category = clean_excel_value(cell_values.get("category")) or clean_excel_value(
        labeled_excel_value(rows, {"наименование", "назначение", "счет", "счёт"})
    )
    counterparty = clean_excel_value(cell_values.get("counterparty_name")) or clean_excel_value(
        labeled_excel_value(rows, {"контрагент", "поставщик", "подрядчик"})
    )
    amount_value = cell_values.get("amount")
    if amount_value in (None, ""):
        amount_value = labeled_excel_value(rows, {"итоговая сумма", "итого", "сумма к оплате", "всего"})
    amount = parse_excel_amount(amount_value)
    if not category or not counterparty or amount <= 0:
        raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR)
    return {"category": category, "counterparty_name": counterparty, "amount": amount}


def parse_finance_pdf_invoice(raw: bytes) -> dict[str, object]:
    """Recognize a PDF invoice/table and return the same result as Excel parsing."""
    try:
        from pdf_excel_adapter import PdfExcelAdapter, PdfExcelAdapterError

        rows = PdfExcelAdapter().extract(raw).rows
        return _parse_finance_rows_for_pdf(rows)
    except (PdfExcelAdapterError, FinanceExcelParseError, ImportError) as error:
        raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR) from error


def recalc_project_finance_totals(handler, con: sqlite3.Connection, project_id: int) -> None:
    spent = con.execute(
        """
        SELECT COALESCE(SUM(amount), 0)
        FROM finance_entries
        WHERE project_id = ? AND direction = 'expense' AND status = 'paid'
        """,
        (project_id,),
    ).fetchone()[0]
    paid = con.execute(
        """
        SELECT COALESCE(SUM(amount), 0)
        FROM finance_entries
        WHERE project_id = ? AND direction = 'income' AND status = 'paid'
        """,
        (project_id,),
    ).fetchone()[0]
    con.execute(
        "UPDATE projects SET spent = ?, paid = ?, updated_at = ? WHERE id = ?",
        (float(spent or 0), float(paid or 0), now_ts(), project_id),
    )


def api_project_finances(handler, path: str) -> None:
    project_id = parse_path_int(path, 2)
    if not project_id:
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
        return
    user = handler.require_project_access(project_id)
    if not user:
        return
    if not user_can_view_finances(user):
        handler.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
        return
    if user_has_any_role(user, {"foreman"}) and not user_can_manage_finances(user):
        handler.send_json(
            HTTPStatus.OK,
            {
                "items": [],
                "summary": {},
                "canUploadInvoice": True,
                "limited": True,
            },
        )
        return
    with db() as con:
        rows = con.execute(
            """
            SELECT
                f.*,
                u.name AS created_by_name,
                d.original_name AS document_original_name,
                d.mime_type AS document_mime_type,
                d.file_ext AS document_file_ext,
                d.storage_path AS document_storage_path
            FROM finance_entries f
            LEFT JOIN users u ON u.id = f.created_by
            LEFT JOIN documents d ON d.id = f.document_id
            WHERE f.project_id = ?
            ORDER BY COALESCE(f.paid_date, f.planned_date, '') DESC, f.id DESC
            """,
            (project_id,),
        ).fetchall()
        estimate_total = con.execute(
            """
            SELECT COALESCE(SUM(planned_qty * planned_price), 0)
            FROM estimate_items
            WHERE project_id = ?
            """,
            (project_id,),
        ).fetchone()[0]
    items = []
    for row in rows:
        item = dict(row)
        if item.get("document_id"):
            item["document"] = {
                "id": item["document_id"],
                "original_name": item.get("document_original_name"),
                "mime_type": item.get("document_mime_type"),
                "file_ext": item.get("document_file_ext"),
                "download_url": f"/api/documents/{item['document_id']}/download",
                "view_url": f"/api/documents/{item['document_id']}/view",
                "can_preview": bool(item.get("document_storage_path"))
                and (
                    str(item.get("document_mime_type") or "").startswith("image/")
                    or str(item.get("document_file_ext") or "") == ".pdf"
                ),
            }
        items.append(item)
    summary = {
        "estimateTotal": float(estimate_total or 0),
        "plannedExpense": sum(float(item["amount"] or 0) for item in items if item["direction"] == "expense" and item["status"] != "cancelled"),
        "paidExpense": sum(float(item["amount"] or 0) for item in items if item["direction"] == "expense" and item["status"] == "paid"),
        "paidIncome": sum(float(item["amount"] or 0) for item in items if item["direction"] == "income" and item["status"] == "paid"),
    }
    summary["balance"] = summary["paidIncome"] - summary["paidExpense"]
    handler.send_json(HTTPStatus.OK, {"items": items, "summary": summary})


def api_create_finance_entry(handler, path: str) -> None:
    project_id = parse_path_int(path, 2)
    if not project_id:
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
        return
    user = handler.require_project_access(project_id)
    if not user:
        return
    payload = handler.read_json()
    direction = str(payload.get("direction", "expense")).strip() or "expense"
    if direction not in {"income", "expense"}:
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_direction"})
        return
    payment_kind = str(payload.get("payment_kind", payload.get("paymentKind", "cash"))).strip() or "cash"
    if payment_kind not in {"cash", "bank_no_vat", "bank_vat"}:
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_payment_kind"})
        return
    status = str(payload.get("status", "planned")).strip() or "planned"
    if status not in {"planned", "approved", "paid", "cancelled"}:
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_status"})
        return
    if not user_can_manage_finances(user):
        if not user_has_any_role(user, {"foreman"}) or direction != "expense" or status not in {"planned", "approved"}:
            handler.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
    try:
        amount = float(payload.get("amount", 0) or 0)
        vat_percent = float(payload.get("vat_percent", payload.get("vatPercent", 0)) or 0)
    except (TypeError, ValueError):
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_amount"})
        return
    if amount <= 0:
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "amount_required"})
        return
    paid_date = str(payload.get("paid_date", payload.get("paidDate", ""))).strip() or None
    if not user_can_manage_finances(user):
        paid_date = None
    with db() as con:
        cur = con.execute(
            """
            INSERT INTO finance_entries (
                project_id, direction, category, payment_kind, vat_percent, amount,
                planned_date, paid_date, counterparty_name, status, notes,
                created_by, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                project_id,
                direction,
                str(payload.get("category", "")).strip() or None,
                payment_kind,
                vat_percent,
                amount,
                str(payload.get("planned_date", payload.get("plannedDate", ""))).strip() or None,
                paid_date,
                str(payload.get("counterparty_name", payload.get("counterpartyName", ""))).strip() or None,
                status,
                str(payload.get("notes", "")).strip() or None,
                user["id"],
                now_ts(),
                now_ts(),
            ),
        )
        handler.recalc_project_finance_totals(con, project_id)
        create_audit(
            con,
            user["id"],
            "create_finance_entry",
            "finance_entry",
            cur.lastrowid,
            {"project_id": project_id, "direction": direction, "amount": amount, "status": status},
        )
        con.commit()
    handler.send_json(HTTPStatus.CREATED, {"id": cur.lastrowid})


def api_upload_finance_invoice(handler, path: str) -> None:
    project_id = parse_path_int(path, 2)
    if not project_id:
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
        return
    user = handler.require_project_access(project_id)
    if not user:
        return
    if not user_can_manage_finances(user) and not user_has_any_role(user, {"foreman"}):
        handler.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
        return

    form = handler.read_multipart()
    upload = form["file"] if "file" in form else None
    if upload is None or not getattr(upload, "file", None) or not getattr(upload, "filename", ""):
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "file_required"})
        return

    original_name = sanitize_filename(upload.filename)
    file_ext = document_extension(original_name)
    if file_ext not in FINANCE_INVOICE_EXTENSIONS:
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_invoice_format"})
        return

    raw = upload.file.read()
    if not raw:
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "empty_file"})
        return
    if len(raw) > MAX_UPLOAD_BYTES:
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "upload_too_large"})
        return

    parsed_invoice: dict[str, object] | None = None
    if file_ext in {".xlsx", ".xls", ".pdf"}:
        try:
            if file_ext == ".pdf":
                parsed_invoice = parse_finance_pdf_invoice(raw)
            else:
                parsed_invoice = parse_finance_excel_invoice(raw, file_ext)
        except FinanceExcelParseError:
            # PDF remains uploadable with manually supplied form fields.  If
            # no fields were supplied, report the same clear parse error as
            # for a malformed Excel file.
            if file_ext != ".pdf" or not any(
                str(form.getfirst(name, "")).strip()
                for name in ("category", "counterparty_name", "amount")
            ):
                handler.send_json(HTTPStatus.BAD_REQUEST, {"error": FINANCE_EXCEL_PARSE_ERROR})
                return

    category = str((parsed_invoice or {}).get("category") or form.getfirst("category", "")).strip()
    counterparty_name = str((parsed_invoice or {}).get("counterparty_name") or form.getfirst("counterparty_name", "")).strip()
    planned_date = str(form.getfirst("planned_date", "")).strip() or None
    payment_kind = str(form.getfirst("payment_kind", "bank_no_vat")).strip() or "bank_no_vat"
    if payment_kind not in {"cash", "bank_no_vat", "bank_vat"}:
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_payment_kind"})
        return
    status = str(form.getfirst("status", "approved")).strip() or "approved"
    if status not in {"planned", "approved"}:
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_status"})
        return
    try:
        amount = float((parsed_invoice or {}).get("amount") or form.getfirst("amount", "0") or 0)
    except (TypeError, ValueError):
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_amount"})
        return
    if amount <= 0:
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "amount_required"})
        return
    if not category:
        category = Path(original_name).stem

    notes = str(form.getfirst("notes", "")).strip() or None
    vat_percent = 20 if payment_kind == "bank_vat" else 0
    mime_type = upload.type or mimetypes.guess_type(original_name)[0] or "application/octet-stream"
    storage_name = f"{now_ts()}_{secrets.token_hex(8)}{file_ext}"
    file_path = project_documents_dir(project_id) / storage_name
    file_path.write_bytes(raw)

    with db() as con:
        doc_cur = con.execute(
            """
            INSERT INTO documents (
                project_id, title, doc_type, status, original_name, storage_name, storage_path,
                mime_type, file_ext, size_bytes, notes, uploaded_by, is_client_visible, created_at, updated_at
            )
            VALUES (?, ?, 'invoice', 'submitted', ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
            """,
            (
                project_id,
                category,
                original_name,
                storage_name,
                str(file_path.relative_to(PROJECT_ROOT)),
                mime_type,
                file_ext,
                len(raw),
                notes,
                user["id"],
                now_ts(),
                now_ts(),
            ),
        )
        finance_cur = con.execute(
            """
            INSERT INTO finance_entries (
                project_id, direction, category, payment_kind, vat_percent, amount,
                planned_date, paid_date, counterparty_name, document_id, status, notes,
                created_by, created_at, updated_at
            )
            VALUES (?, 'expense', ?, ?, ?, ?, ?, NULL, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                project_id,
                category,
                payment_kind,
                vat_percent,
                amount,
                planned_date,
                counterparty_name or None,
                doc_cur.lastrowid,
                status,
                notes,
                user["id"],
                now_ts(),
                now_ts(),
            ),
        )
        handler.recalc_project_finance_totals(con, project_id)
        create_audit(
            con,
            user["id"],
            "upload_finance_invoice",
            "finance_entry",
            finance_cur.lastrowid,
            {
                "project_id": project_id,
                "document_id": doc_cur.lastrowid,
                "amount": amount,
                "status": status,
                "parsed_excel": bool(parsed_invoice),
            },
        )
        con.commit()
    handler.send_json(
        HTTPStatus.CREATED,
        {
            "id": finance_cur.lastrowid,
            "document_id": doc_cur.lastrowid,
            "parsedInvoice": parsed_invoice,
            "status": status,
        },
    )


def api_update_finance_entry(handler, path: str) -> None:
    finance_id = parse_path_int(path, 2)
    if not finance_id:
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_finance_id"})
        return
    payload = handler.read_json()
    with db() as con:
        row = con.execute("SELECT * FROM finance_entries WHERE id = ?", (finance_id,)).fetchone()
        if not row:
            handler.send_json(HTTPStatus.NOT_FOUND, {"error": "finance_not_found"})
            return
        user = handler.require_project_access(int(row["project_id"]))
        if not user:
            return
        if not user_can_manage_finances(user):
            handler.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        status = str(payload.get("status", row["status"])).strip() or row["status"]
        if status not in {"planned", "approved", "paid", "cancelled"}:
            handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_status"})
            return
        if status == "paid" and str(row["status"] or "") != "paid" and not user_can_pay_invoices(user):
            handler.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        con.execute(
            """
            UPDATE finance_entries
            SET planned_date = ?, paid_date = ?, status = ?, notes = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                str(payload.get("planned_date", payload.get("plannedDate", row["planned_date"] or ""))).strip() or None,
                str(payload.get("paid_date", payload.get("paidDate", row["paid_date"] or ""))).strip() or None,
                status,
                str(payload.get("notes", row["notes"] or "")).strip() or None,
                now_ts(),
                finance_id,
            ),
        )
        handler.recalc_project_finance_totals(con, int(row["project_id"]))
        create_audit(
            con,
            user["id"],
            "update_finance_entry",
            "finance_entry",
            finance_id,
            {"project_id": row["project_id"], "status": status},
        )
        con.commit()
    handler.send_json(HTTPStatus.OK, {"ok": True})


def api_pay_invoice(handler) -> None:
    payer = handler.require_role({"admin", "director"})
    if not payer:
        return
    payload = handler.read_json()
    try:
        finance_id = int(payload.get("finance_id", payload.get("financeId", payload.get("invoice_id", payload.get("invoiceId", payload.get("id"))))))
    except (TypeError, ValueError):
        handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_finance_id"})
        return
    paid_date = str(payload.get("paid_date", payload.get("paidDate", TODAY_ISO))).strip() or TODAY_ISO
    with db() as con:
        row = con.execute("SELECT * FROM finance_entries WHERE id = ?", (finance_id,)).fetchone()
        if not row:
            handler.send_json(HTTPStatus.NOT_FOUND, {"error": "finance_not_found"})
            return
        if row["direction"] != "expense":
            handler.send_json(HTTPStatus.BAD_REQUEST, {"error": "not_expense_invoice"})
            return
        user = handler.require_project_access(int(row["project_id"]))
        if not user:
            return
        if not user_can_pay_invoices(user):
            handler.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        con.execute(
            """
            UPDATE finance_entries
            SET status = 'paid', paid_date = ?, updated_at = ?
            WHERE id = ?
            """,
            (paid_date, now_ts(), finance_id),
        )
        handler.recalc_project_finance_totals(con, int(row["project_id"]))
        create_audit(
            con,
            payer["id"],
            "pay_invoice",
            "finance_entry",
            finance_id,
            {"project_id": row["project_id"], "paid_date": paid_date},
        )
        con.commit()
    handler.send_json(HTTPStatus.OK, {"ok": True, "id": finance_id, "status": "paid", "paidDate": paid_date})
