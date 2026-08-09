from __future__ import annotations

import base64
import cgi
import hashlib
import html
import hmac
import io
import json
import mimetypes
import os
import re
import secrets
import sqlite3
import sys
import time
import unicodedata
import urllib.parse
import urllib.request
from datetime import date, timedelta
from http import HTTPStatus
from http.cookies import SimpleCookie
from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from pathlib import Path

import jwt


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEPLOY_ROOT = PROJECT_ROOT / "deploy"
FRONTEND_ROOT = PROJECT_ROOT / "frontend"
FRONTEND_TEMPLATES = FRONTEND_ROOT / "templates"
FRONTEND_PAGES = FRONTEND_ROOT / "pages"
FRONTEND_ASSETS = FRONTEND_ROOT / "assets"
DATA_DIR = PROJECT_ROOT / "data"
DOCUMENTS_DIR = DATA_DIR / "documents"
DB_PATH = DATA_DIR / "pmbi.sqlite3"
BOOTSTRAP_PATH = DATA_DIR / "INITIAL_ADMIN.txt"

HOST = os.environ.get("PMBI_HOST", "127.0.0.1")
PORT = int(os.environ.get("PMBI_PORT", os.environ.get("PORT", "8080")))
SESSION_COOKIE = "pmbi_session"
SESSION_TTL_SECONDS = 60 * 60 * 12
PASSWORD_ITERATIONS = 220_000
MAX_UPLOAD_BYTES = 50 * 1024 * 1024
FINANCE_INVOICE_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".xlsx", ".xls"}
FINANCE_EXCEL_TEMPLATE_CELLS = {
    "category": os.environ.get("PMBI_FINANCE_TEMPLATE_NAME_CELL", "B2"),
    "counterparty_name": os.environ.get("PMBI_FINANCE_TEMPLATE_COUNTERPARTY_CELL", "B3"),
    "amount": os.environ.get("PMBI_FINANCE_TEMPLATE_TOTAL_CELL", "B4"),
}
FINANCE_EXCEL_PARSE_ERROR = "Ошибка: Формат файла не соответствует шаблону отчета"

LOGIN_PATH = "/login"
DEFAULT_AUTH_PATH = "/app/dashboard"
PMBI_PUBLIC_BASE_URL = (os.environ.get("PMBI_PUBLIC_BASE_URL", "") or "").strip().rstrip("/")
PMBI_AUTOBOT_BASE_URL = (os.environ.get("PMBI_AUTOBOT_BASE_URL", "http://127.0.0.1:8765") or "").strip().rstrip("/")

CLERK_PUBLISHABLE_KEY = (os.environ.get("CLERK_PUBLISHABLE_KEY", "") or "").strip()
CLERK_SECRET_KEY = (os.environ.get("CLERK_SECRET_KEY", "") or "").strip()
CLERK_JWT_KEY = (os.environ.get("CLERK_JWT_KEY", "") or "").replace("\\n", "\n").strip()
CLERK_SIGN_IN_FALLBACK_REDIRECT_URL = (
    os.environ.get("CLERK_SIGN_IN_FALLBACK_REDIRECT_URL", "") or DEFAULT_AUTH_PATH
).strip() or DEFAULT_AUTH_PATH
CLERK_SIGN_UP_FALLBACK_REDIRECT_URL = (
    os.environ.get("CLERK_SIGN_UP_FALLBACK_REDIRECT_URL", "") or DEFAULT_AUTH_PATH
).strip() or DEFAULT_AUTH_PATH
CLERK_AUTHORIZED_PARTIES = {
    item.strip().rstrip("/")
    for item in (os.environ.get("CLERK_AUTHORIZED_PARTIES", "") or "").split(",")
    if item.strip()
}
if PMBI_PUBLIC_BASE_URL:
    CLERK_AUTHORIZED_PARTIES.add(PMBI_PUBLIC_BASE_URL)
CLERK_ADMIN_EMAILS = {
    item.strip().lower()
    for item in (os.environ.get("CLERK_ADMIN_EMAILS", "") or "").split(",")
    if item.strip()
}
CLERK_API_BASE = "https://api.clerk.com/v1"


ROLE_ALLOWED_PREFIXES = {
    "admin": ["*"],
    "director": ["*"],
    "foreman": [
        "/app",
        "/app/dashboard",
        "/app/projects",
        "/app/autobot",
        "/app/schedule",
        "/app/logs",
        "/app/warehouse",
        "/app/suppliers",
        "/app/chats",
    ],
    "buyer": [
        "/app",
        "/app/dashboard",
        "/app/projects",
        "/app/autobot",
        "/app/logs",
        "/app/warehouse",
        "/app/suppliers",
        "/app/chats",
    ],
    "purchaser": [
        "/app",
        "/app/dashboard",
        "/app/projects",
        "/app/autobot",
        "/app/logs",
        "/app/warehouse",
        "/app/suppliers",
        "/app/chats",
    ],
    "financier": [
        "/app",
        "/app/dashboard",
        "/app/projects",
        "/app/autobot",
        "/app/reports",
    ],
    "accountant": [
        "/app",
        "/app/dashboard",
        "/app/projects",
        "/app/autobot",
        "/app/reports",
    ],
    "client": [
        "/app",
        "/app/dashboard",
        "/app/projects",
        "/app/schedule",
        "/app/logs",
        "/app/chats",
    ],
    "customer": [
        "/app",
        "/app/dashboard",
        "/app/projects",
        "/app/schedule",
        "/app/logs",
        "/app/chats",
    ],
}

ROLE_LABELS = {
    "admin": "Администратор",
    "director": "Директор",
    "foreman": "Прораб",
    "purchaser": "Закупщик",
    "buyer": "Закупщик",
    "financier": "Ответственный за финплан",
    "accountant": "Бухгалтер",
    "customer": "Заказчик",
    "client": "Заказчик",
}

ROLE_DESCRIPTIONS = {
    "admin": "Технический администратор с полным доступом.",
    "director": "Руководитель: видит все объекты, назначает роли и контролирует систему.",
    "foreman": "Прораб/куратор: ведёт назначенные объекты и ежедневные отчёты.",
    "purchaser": "Закупщик/снабженец: отвечает за поставщиков, материалы и поставки.",
    "buyer": "Закупщик/снабженец: устаревший код роли для совместимости.",
    "financier": "Ответственный за финпланы и платежи.",
    "accountant": "Бухгалтер: работает с финансовыми и подтверждающими документами.",
    "customer": "Заказчик: ограниченный доступ к своим объектам.",
    "client": "Заказчик: устаревший код роли для совместимости.",
}

LEGACY_ROLE_ALIASES = {
    "buyer": "purchaser",
    "client": "customer",
}

ROLE_CODES = tuple(ROLE_LABELS.keys())

PUBLIC_STATIC_PATHS = {
    "/",
    "/index.html",
    LOGIN_PATH,
    "/robots.txt",
}

APP_PAGES = {
    "/app": ("dashboard", "Панель"),
    "/app/dashboard": ("dashboard", "Панель"),
    "/app/projects": ("projects", "Объекты"),
    "/app/companies": ("companies", "Компании"),
    "/app/warehouse": ("warehouse", "Склад"),
    "/app/schedule": ("schedule", "График работ"),
    "/app/logs": ("logs", "Журнал работ"),
    "/app/chats": ("chats", "Чаты"),
    "/app/users": ("users", "Пользователи"),
    "/app/reports": ("reports", "Отчётность"),
}

APP_PAGES["/app/suppliers"] = ("suppliers", "Контрагенты")

APP_PAGES["/app/autobot"] = ("autobot", "AutoBot")

DIRECTOR_ACTION_BLOCK_RE = re.compile(
    r"<(?P<tag>section|button)\b(?=[^>]*\sdata-director-action\b)[\s\S]*?</(?P=tag)>",
    re.IGNORECASE,
)

STAGE_CATEGORY_KEYWORDS = {
    "prep": ["подготов", "демонтаж", "временн", "мобилиз", "размет", "геодез"],
    "concrete": ["котлован", "землян", "фундамент", "бетон", "монолит", "арматур", "стяжк"],
    "masonry": ["кладк", "кирпич", "блок", "перегород", "каркас", "сварк", "металлоконструк"],
    "roof": ["кровл", "крыша", "гидроизоляц", "утеплен"],
    "facade": ["фасад", "витраж", "остеклен", "окн", "двер"],
    "electrical": ["электр", "кабель", "щит", "слаботоч", "освещен"],
    "plumbing": ["сантех", "водоснаб", "канализ", "отоплен", "труб", "радиатор"],
    "ventilation": ["вентиляц", "дымоудал", "кондицион", "чиллер"],
    "finishing": ["отдел", "штукатур", "шпаклев", "плитк", "окраск", "ламинат", "потол", "дверн"],
    "landscape": ["благоустрой", "асфальт", "бордюр", "озеленен", "наруж"],
    "handover": ["пуск", "исполн", "сдач", "акт", "комис", "документ"],
}


EXECUTIVE_TEMPLATE_RULES = {
    "hidden_work_act": {
        "title": "Акт скрытых работ",
        "doc_type": "hidden_work_act",
        "categories": {"concrete", "masonry", "roof", "facade", "electrical", "plumbing", "ventilation", "landscape"},
        "optional": False,
        "default_notes": "Подготовить акт скрытых работ по этапу, проверить объемы и подписи стройконтроля.",
    },
    "inspection_act": {
        "title": "Акт осмотра",
        "doc_type": "inspection_act",
        "categories": {"prep", "concrete", "masonry", "roof", "facade", "landscape", "handover"},
        "optional": False,
        "default_notes": "Нужен акт осмотра/освидетельствования перед закрытием этапа.",
    },
    "executive": {
        "title": "Исполнительная документация",
        "doc_type": "executive",
        "categories": {"electrical", "plumbing", "ventilation", "handover"},
        "optional": False,
        "default_notes": "Собрать исполнительную документацию и приложить схемы, фото, подтверждающие листы.",
    },
    "technical_solution": {
        "title": "Техрешение",
        "doc_type": "technical_solution",
        "categories": {"general"},
        "optional": True,
        "default_notes": "Использовать, если на этапе выявлено расхождение проекта, сметы или факта работ.",
    },
}


def now_ts() -> int:
    return int(time.time())


TODAY_ISO = date.today().isoformat()


def normalize_estimate_item_kind(value: object) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return "material"
    work_markers = ("work", "works", "service", "services", "labor", "labour", "job", "работ", "услуг", "труд")
    material_markers = ("material", "materials", "supply", "goods", "equipment", "матер", "товар", "оборуд")
    work_markers = work_markers + (
        "\u0440\u0430\u0431\u043e\u0442",
        "\u0443\u0441\u043b\u0443\u0433",
        "\u0442\u0440\u0443\u0434",
    )
    material_markers = material_markers + (
        "\u043c\u0430\u0442\u0435\u0440",
        "\u0442\u043e\u0432\u0430\u0440",
        "\u043e\u0431\u043e\u0440\u0443\u0434",
    )
    if any(marker in text for marker in work_markers):
        return "work"
    if any(marker in text for marker in material_markers):
        return "material"
    return "material"


def estimate_unit_multiplier(unit: object) -> float:
    match = re.match(r"^\s*(\d+(?:[\.,]\d+)?)\s+\S+", str(unit or "").strip())
    if not match:
        return 1.0
    try:
        multiplier = float(match.group(1).replace(",", "."))
    except ValueError:
        return 1.0
    return multiplier if multiplier > 1 else 1.0


def normalize_estimate_planned_qty(unit: object, qty: object) -> float:
    try:
        value = float(qty or 0)
    except (TypeError, ValueError):
        return 0.0
    multiplier = estimate_unit_multiplier(unit)
    if multiplier >= 100 and value >= multiplier:
        return value / multiplier
    return value


def normalize_estimate_planned_values(unit: object, qty: object, price: object) -> tuple[float, float]:
    try:
        raw_qty = float(qty or 0)
    except (TypeError, ValueError):
        raw_qty = 0.0
    try:
        raw_price = float(price or 0)
    except (TypeError, ValueError):
        raw_price = 0.0
    planned_qty = normalize_estimate_planned_qty(unit, raw_qty)
    if raw_qty > 0 and planned_qty > 0 and planned_qty != raw_qty:
        raw_price *= raw_qty / planned_qty
    return planned_qty, raw_price


FUZZY_MATCH_THRESHOLD = 0.70


def normalize_fuzzy_text(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).lower().replace("ё", "е")
    text = re.sub(r"\bпровод\b", "кабель", text)
    text = re.sub(r"\bпровода\b", "кабель", text)
    text = re.sub(r"(?<=\d)[xх×](?=\d)", "x", text)
    text = re.sub(r"(?<=\d)[,.](?=\d)", ".", text)
    text = re.sub(r"[^0-9a-zа-я.]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def levenshtein_distance(left: str, right: str) -> int:
    if left == right:
        return 0
    if not left:
        return len(right)
    if not right:
        return len(left)
    if len(left) < len(right):
        left, right = right, left
    previous = list(range(len(right) + 1))
    for i, left_char in enumerate(left, 1):
        current = [i]
        for j, right_char in enumerate(right, 1):
            current.append(
                min(
                    previous[j] + 1,
                    current[j - 1] + 1,
                    previous[j - 1] + (0 if left_char == right_char else 1),
                )
            )
        previous = current
    return previous[-1]


def dice_coefficient(left: str, right: str) -> float:
    def bigrams(text: str) -> list[str]:
        compact = re.sub(r"\s+", "", text)
        if len(compact) < 2:
            return [compact] if compact else []
        return [compact[index : index + 2] for index in range(len(compact) - 1)]

    left_bigrams = bigrams(left)
    right_bigrams = bigrams(right)
    if not left_bigrams or not right_bigrams:
        return 0.0
    right_counts: dict[str, int] = {}
    for gram in right_bigrams:
        right_counts[gram] = right_counts.get(gram, 0) + 1
    overlap = 0
    for gram in left_bigrams:
        count = right_counts.get(gram, 0)
        if count:
            overlap += 1
            right_counts[gram] = count - 1
    return (2.0 * overlap) / (len(left_bigrams) + len(right_bigrams))


def fuzzy_similarity(left_value: object, right_value: object) -> float:
    left = normalize_fuzzy_text(left_value)
    right = normalize_fuzzy_text(right_value)
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0
    if left in right or right in left:
        short = min(len(left), len(right))
        long = max(len(left), len(right))
        return max(0.82, short / long)
    lev = 1.0 - (levenshtein_distance(left, right) / max(len(left), len(right), 1))
    dice = dice_coefficient(left, right)
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    token_score = len(left_tokens & right_tokens) / max(len(left_tokens | right_tokens), 1)
    return max(lev, dice, token_score)


def extract_labeled_note_value(notes: object, labels: tuple[str, ...]) -> str | None:
    text = str(notes or "").strip()
    if not text:
        return None
    normalized_labels = tuple(str(label).strip().lower() for label in labels if str(label).strip())
    for chunk in re.split(r"[;\n]+", text):
        part = chunk.strip()
        if not part:
            continue
        lowered = part.lower()
        for label in normalized_labels:
            colon_prefix = f"{label}:"
            if lowered.startswith(colon_prefix):
                value = part[len(colon_prefix):].strip()
                return value or None
            dash_prefix = f"{label} -"
            if lowered.startswith(dash_prefix):
                value = part[len(dash_prefix):].strip()
                return value or None
    return None


def estimate_item_kind_from_payload(item: dict | sqlite3.Row) -> str:
    raw_value = (
        item.get("item_kind")
        or item.get("itemKind")
        or item.get("type")
        or item.get("type_label")
        or item.get("typeLabel")
        or item.get("position_type")
        or item.get("positionType")
        or ""
    )
    raw_text = str(raw_value or "").strip()
    if raw_text:
        return normalize_estimate_item_kind(raw_text)
    note_value = extract_labeled_note_value(item.get("notes"), ("Тип", "Type"))
    note_value = note_value or extract_labeled_note_value(item.get("notes"), ("\u0422\u0438\u043f", "Type"))
    return normalize_estimate_item_kind(note_value)


def estimate_section_title_from_payload(item: dict | sqlite3.Row) -> str | None:
    direct_value = (
        item.get("section_title")
        or item.get("sectionTitle")
        or item.get("section_name")
        or item.get("sectionName")
        or item.get("section")
        or item.get("chapter")
        or ""
    )
    direct_text = str(direct_value or "").strip()
    if direct_text:
        return direct_text
    return extract_labeled_note_value(item.get("notes"), ("Раздел", "Section", "Chapter"))


def b64url(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).decode("ascii").rstrip("=")


def hash_password(password: str, salt: bytes | None = None) -> str:
    salt = salt or secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt,
        PASSWORD_ITERATIONS,
    )
    return f"pbkdf2_sha256${PASSWORD_ITERATIONS}${b64url(salt)}${b64url(digest)}"


def verify_password(password: str, stored: str) -> bool:
    try:
        algo, iterations, salt_b64, digest_b64 = stored.split("$", 3)
        if algo != "pbkdf2_sha256":
            return False
        salt = base64.urlsafe_b64decode(salt_b64 + "=" * (-len(salt_b64) % 4))
        expected = base64.urlsafe_b64decode(digest_b64 + "=" * (-len(digest_b64) % 4))
        actual = hashlib.pbkdf2_hmac(
            "sha256",
            password.encode("utf-8"),
            salt,
            int(iterations),
        )
        return hmac.compare_digest(actual, expected)
    except Exception:
        return False


def token_hash(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def db() -> sqlite3.Connection:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    return connection


def ensure_columns(con: sqlite3.Connection, table: str, columns: dict[str, str]) -> None:
    existing = {row["name"] for row in con.execute(f"PRAGMA table_info({table})").fetchall()}
    for name, definition in columns.items():
        if name not in existing:
            con.execute(f"ALTER TABLE {table} ADD COLUMN {name} {definition}")


def normalize_role(role: str | None) -> str:
    role_code = str(role or "").strip()
    return LEGACY_ROLE_ALIASES.get(role_code, role_code)


def create_audit(
    con: sqlite3.Connection,
    user_id: int | None,
    action: str,
    entity: str | None = None,
    entity_id: int | None = None,
    payload: dict | None = None,
) -> None:
    con.execute(
        """
        INSERT INTO audit_log (user_id, action, entity, entity_id, payload, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (user_id, action, entity, entity_id, json.dumps(payload or {}, ensure_ascii=False), now_ts()),
    )


def migrate_users_table(con: sqlite3.Connection) -> None:
    create_sql_row = con.execute(
        "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'users'"
    ).fetchone()
    create_sql = create_sql_row["sql"] if create_sql_row else ""
    if "CHECK(role IN ('director','foreman','buyer','client'))" not in create_sql:
        return

    con.execute("ALTER TABLE users RENAME TO users_legacy")
    con.execute(
        """
        CREATE TABLE users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            login TEXT NOT NULL UNIQUE,
            email TEXT,
            phone TEXT,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            name TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'active',
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at INTEGER NOT NULL,
            updated_at INTEGER
        )
        """
    )
    con.execute(
        """
        INSERT INTO users (id, login, password_hash, role, name, status, is_active, created_at, updated_at)
        SELECT id, login, password_hash, role, name,
               CASE WHEN is_active = 1 THEN 'active' ELSE 'blocked' END,
               is_active, created_at, created_at
        FROM users_legacy
        """
    )
    # Keep the legacy table until foreign keys in already-created tables are
    # rebuilt to point back to users. SQLite rewrites FK targets on rename.


def table_exists(con: sqlite3.Connection, table: str) -> bool:
    return con.execute(
        "SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table,),
    ).fetchone() is not None


def table_sql(con: sqlite3.Connection, table: str) -> str:
    row = con.execute(
        "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = ?",
        (table,),
    ).fetchone()
    return row["sql"] if row and row["sql"] else ""


def rebuild_table(con: sqlite3.Connection, table: str, create_sql: str) -> None:
    old_table = f"{table}__old"
    con.execute(f"DROP TABLE IF EXISTS {old_table}")
    con.execute("PRAGMA legacy_alter_table = ON")
    con.execute(f"ALTER TABLE {table} RENAME TO {old_table}")
    con.execute(create_sql)
    old_columns = [row["name"] for row in con.execute(f"PRAGMA table_info({old_table})").fetchall()]
    new_columns = [row["name"] for row in con.execute(f"PRAGMA table_info({table})").fetchall()]
    common_columns = [column for column in new_columns if column in old_columns]
    if common_columns:
        column_sql = ", ".join(common_columns)
        con.execute(f"INSERT INTO {table} ({column_sql}) SELECT {column_sql} FROM {old_table}")
    con.execute(f"DROP TABLE {old_table}")
    con.execute("PRAGMA legacy_alter_table = OFF")


def repair_legacy_user_references(con: sqlite3.Connection) -> None:
    affected = [
        table
        for table in [
            "sessions",
            "projects",
            "user_project_access",
            "stock_moves",
            "tasks",
            "chat_messages",
            "daily_logs",
            "audit_log",
            "object_assignments",
            "user_roles",
        ]
        if table_exists(con, table) and "users_legacy" in table_sql(con, table)
    ]
    if not affected:
        return

    con.execute("PRAGMA foreign_keys = OFF")
    con.execute("PRAGMA legacy_alter_table = ON")
    schemas = {
        "sessions": """
            CREATE TABLE sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                token_hash TEXT NOT NULL UNIQUE,
                created_at INTEGER NOT NULL,
                expires_at INTEGER NOT NULL,
                user_agent TEXT,
                ip TEXT
            )
        """,
        "projects": """
            CREATE TABLE projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                client_id INTEGER REFERENCES clients(id) ON DELETE SET NULL,
                customer_company_id INTEGER REFERENCES companies(id) ON DELETE SET NULL,
                own_legal_entity_id INTEGER REFERENCES companies(id) ON DELETE SET NULL,
                address TEXT NOT NULL,
                city TEXT,
                region TEXT,
                client_name TEXT NOT NULL,
                contract_no TEXT,
                contract_date TEXT,
                director_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                foreman_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                buyer_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                client_user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                status TEXT NOT NULL,
                progress INTEGER NOT NULL DEFAULT 0,
                budget REAL NOT NULL DEFAULT 0,
                paid REAL NOT NULL DEFAULT 0,
                spent REAL NOT NULL DEFAULT 0,
                started_at TEXT,
                deadline_at TEXT,
                internal_schedule_status TEXT NOT NULL DEFAULT 'draft',
                internal_schedule_version INTEGER NOT NULL DEFAULT 1,
                internal_schedule_approved_at TEXT,
                customer_schedule_status TEXT NOT NULL DEFAULT 'draft',
                customer_schedule_version INTEGER NOT NULL DEFAULT 1,
                customer_schedule_approved_at TEXT,
                schedule_generated_at TEXT,
                description TEXT,
                created_at INTEGER NOT NULL,
                updated_at INTEGER
            )
        """,
        "user_project_access": """
            CREATE TABLE user_project_access (
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                PRIMARY KEY (user_id, project_id)
            )
        """,
        "stock_moves": """
            CREATE TABLE stock_moves (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                estimate_item_id INTEGER REFERENCES estimate_items(id) ON DELETE SET NULL,
                move_type TEXT NOT NULL CHECK(move_type IN ('purchase','receipt','use','writeoff')),
                qty REAL NOT NULL,
                price REAL NOT NULL DEFAULT 0,
                comment TEXT,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at INTEGER NOT NULL
            )
        """,
        "tasks": """
            CREATE TABLE tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                title TEXT NOT NULL,
                description TEXT,
                status TEXT NOT NULL DEFAULT 'open',
                priority TEXT NOT NULL DEFAULT 'normal',
                assignee_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                due_at TEXT,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at INTEGER NOT NULL
            )
        """,
        "chat_messages": """
            CREATE TABLE chat_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                chat_id INTEGER NOT NULL REFERENCES chats(id) ON DELETE CASCADE,
                user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                body TEXT NOT NULL,
                created_at INTEGER NOT NULL
            )
        """,
        "daily_logs": """
            CREATE TABLE daily_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                report_date TEXT NOT NULL,
                title TEXT NOT NULL,
                work_done TEXT NOT NULL DEFAULT '',
                workers_count INTEGER NOT NULL DEFAULT 0,
                equipment TEXT NOT NULL DEFAULT '',
                blockers TEXT NOT NULL DEFAULT '',
                next_steps TEXT NOT NULL DEFAULT '',
                is_client_visible INTEGER NOT NULL DEFAULT 1,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at INTEGER NOT NULL
            )
        """,
        "audit_log": """
            CREATE TABLE audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                action TEXT NOT NULL,
                entity TEXT,
                entity_id INTEGER,
                payload TEXT,
                created_at INTEGER NOT NULL
            )
        """,
        "object_assignments": """
            CREATE TABLE object_assignments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                object_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                role_code TEXT NOT NULL,
                responsibility TEXT,
                is_primary INTEGER NOT NULL DEFAULT 0,
                assigned_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                assigned_at INTEGER NOT NULL,
                UNIQUE(object_id, user_id, role_code)
            )
        """,
        "user_roles": """
            CREATE TABLE user_roles (
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                role_id INTEGER NOT NULL REFERENCES roles(id) ON DELETE CASCADE,
                created_at INTEGER NOT NULL,
                PRIMARY KEY (user_id, role_id)
            )
        """,
    }
    for table in affected:
        rebuild_table(con, table, schemas[table])
    if table_exists(con, "users_legacy"):
        con.execute("DROP TABLE users_legacy")
    con.execute("PRAGMA legacy_alter_table = OFF")
    con.execute("PRAGMA foreign_keys = ON")


def init_db() -> None:
    DOCUMENTS_DIR.mkdir(parents=True, exist_ok=True)
    with db() as con:
        con.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                login TEXT NOT NULL UNIQUE,
                email TEXT,
                phone TEXT,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL,
                name TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'active',
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at INTEGER NOT NULL,
                updated_at INTEGER
            );

            CREATE TABLE IF NOT EXISTS roles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                description TEXT
            );

            CREATE TABLE IF NOT EXISTS user_roles (
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                role_id INTEGER NOT NULL REFERENCES roles(id) ON DELETE CASCADE,
                created_at INTEGER NOT NULL,
                PRIMARY KEY (user_id, role_id)
            );

            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                type TEXT NOT NULL CHECK(type IN ('own_legal_entity','client','supplier','contractor','other')),
                name TEXT NOT NULL,
                inn TEXT,
                kpp TEXT,
                ogrn TEXT,
                phone TEXT,
                email TEXT,
                address TEXT,
                notes TEXT,
                created_at INTEGER NOT NULL
            );

            CREATE TABLE IF NOT EXISTS clients (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER REFERENCES companies(id) ON DELETE SET NULL,
                name TEXT NOT NULL,
                contact_person TEXT,
                phone TEXT,
                email TEXT,
                notes TEXT
            );

            CREATE TABLE IF NOT EXISTS sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                token_hash TEXT NOT NULL UNIQUE,
                created_at INTEGER NOT NULL,
                expires_at INTEGER NOT NULL,
                user_agent TEXT,
                ip TEXT
            );

            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                client_id INTEGER REFERENCES clients(id) ON DELETE SET NULL,
                customer_company_id INTEGER REFERENCES companies(id) ON DELETE SET NULL,
                own_legal_entity_id INTEGER REFERENCES companies(id) ON DELETE SET NULL,
                address TEXT NOT NULL,
                city TEXT,
                region TEXT,
                client_name TEXT NOT NULL,
                contract_no TEXT,
                contract_date TEXT,
                director_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                foreman_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                buyer_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                client_user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                status TEXT NOT NULL,
                progress INTEGER NOT NULL DEFAULT 0,
                budget REAL NOT NULL DEFAULT 0,
                paid REAL NOT NULL DEFAULT 0,
                spent REAL NOT NULL DEFAULT 0,
                started_at TEXT,
                deadline_at TEXT,
                internal_schedule_status TEXT NOT NULL DEFAULT 'draft',
                internal_schedule_version INTEGER NOT NULL DEFAULT 1,
                internal_schedule_approved_at TEXT,
                customer_schedule_status TEXT NOT NULL DEFAULT 'draft',
                customer_schedule_version INTEGER NOT NULL DEFAULT 1,
                customer_schedule_approved_at TEXT,
                schedule_generated_at TEXT,
                description TEXT,
                created_at INTEGER NOT NULL,
                updated_at INTEGER
            );

            CREATE TABLE IF NOT EXISTS object_assignments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                object_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                role_code TEXT NOT NULL,
                responsibility TEXT,
                is_primary INTEGER NOT NULL DEFAULT 0,
                assigned_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                assigned_at INTEGER NOT NULL,
                UNIQUE(object_id, user_id, role_code)
            );

            CREATE TABLE IF NOT EXISTS user_project_access (
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                PRIMARY KEY (user_id, project_id)
            );

            CREATE TABLE IF NOT EXISTS estimate_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                title TEXT NOT NULL,
                unit TEXT NOT NULL,
                planned_qty REAL NOT NULL,
                planned_price REAL NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS stock_moves (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                estimate_item_id INTEGER REFERENCES estimate_items(id) ON DELETE SET NULL,
                move_type TEXT NOT NULL CHECK(move_type IN ('purchase','receipt','use','writeoff')),
                qty REAL NOT NULL,
                price REAL NOT NULL DEFAULT 0,
                comment TEXT,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at INTEGER NOT NULL
            );

            CREATE TABLE IF NOT EXISTS warehouse_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                item_type TEXT NOT NULL DEFAULT 'material' CHECK(item_type IN ('material','tool')),
                category TEXT,
                name TEXT NOT NULL,
                sku TEXT,
                unit TEXT NOT NULL DEFAULT 'шт',
                qty REAL NOT NULL DEFAULT 0,
                condition_status TEXT,
                created_at INTEGER NOT NULL,
                updated_at INTEGER
            );

            CREATE TABLE IF NOT EXISTS warehouse_transfers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                warehouse_item_id INTEGER NOT NULL REFERENCES warehouse_items(id) ON DELETE CASCADE,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                estimate_item_id INTEGER REFERENCES estimate_items(id) ON DELETE SET NULL,
                qty REAL NOT NULL,
                unit TEXT NOT NULL,
                comment TEXT,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at INTEGER NOT NULL
            );

            CREATE TABLE IF NOT EXISTS supplier_offers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                estimate_item_id INTEGER REFERENCES estimate_items(id) ON DELETE SET NULL,
                company_id INTEGER REFERENCES companies(id) ON DELETE SET NULL,
                candidate_type TEXT NOT NULL DEFAULT 'supplier',
                candidate_name TEXT NOT NULL,
                source_type TEXT NOT NULL DEFAULT 'manual',
                source_url TEXT,
                contact_name TEXT,
                phone TEXT,
                price REAL NOT NULL DEFAULT 0,
                qty REAL NOT NULL DEFAULT 0,
                unit TEXT,
                status TEXT NOT NULL DEFAULT 'new',
                notes TEXT,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at INTEGER NOT NULL,
                updated_at INTEGER
            );

            CREATE TABLE IF NOT EXISTS material_schedule_snapshots (
                project_id INTEGER PRIMARY KEY REFERENCES projects(id) ON DELETE CASCADE,
                payload TEXT NOT NULL,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at INTEGER NOT NULL,
                updated_at INTEGER NOT NULL
            );

            CREATE TABLE IF NOT EXISTS work_stages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                title TEXT NOT NULL,
                position INTEGER NOT NULL DEFAULT 0,
                parent_id INTEGER,
                stage_kind TEXT NOT NULL DEFAULT 'section',
                status_code TEXT NOT NULL DEFAULT 'not_started',
                planned_start TEXT,
                planned_end TEXT,
                customer_start TEXT,
                customer_end TEXT,
                fact_start TEXT,
                fact_end TEXT,
                progress INTEGER NOT NULL DEFAULT 0,
                responsible TEXT,
                notes TEXT,
                is_client_visible INTEGER NOT NULL DEFAULT 1,
                depends_on_materials INTEGER NOT NULL DEFAULT 0,
                created_at INTEGER NOT NULL,
                updated_at INTEGER
            );

            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                title TEXT NOT NULL,
                description TEXT,
                status TEXT NOT NULL DEFAULT 'open',
                priority TEXT NOT NULL DEFAULT 'normal',
                assignee_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                due_at TEXT,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at INTEGER NOT NULL
            );

            CREATE TABLE IF NOT EXISTS chats (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                chat_type TEXT NOT NULL CHECK(chat_type IN ('team','client')),
                title TEXT NOT NULL,
                created_at INTEGER NOT NULL
            );

            CREATE TABLE IF NOT EXISTS chat_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                chat_id INTEGER NOT NULL REFERENCES chats(id) ON DELETE CASCADE,
                user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                body TEXT NOT NULL,
                created_at INTEGER NOT NULL
            );

            CREATE TABLE IF NOT EXISTS documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                title TEXT NOT NULL,
                doc_type TEXT NOT NULL DEFAULT 'file',
                status TEXT NOT NULL DEFAULT 'draft',
                original_name TEXT,
                storage_name TEXT,
                storage_path TEXT,
                mime_type TEXT,
                file_ext TEXT,
                size_bytes INTEGER,
                notes TEXT,
                uploaded_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                is_client_visible INTEGER NOT NULL DEFAULT 0,
                created_at INTEGER NOT NULL,
                updated_at INTEGER
            );

            CREATE TABLE IF NOT EXISTS finance_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                direction TEXT NOT NULL CHECK(direction IN ('income','expense')),
                category TEXT,
                payment_kind TEXT NOT NULL DEFAULT 'cash' CHECK(payment_kind IN ('cash','bank_no_vat','bank_vat')),
                vat_percent REAL NOT NULL DEFAULT 0,
                amount REAL NOT NULL DEFAULT 0,
                planned_date TEXT,
                paid_date TEXT,
                counterparty_name TEXT,
                company_id INTEGER REFERENCES companies(id) ON DELETE SET NULL,
                document_id INTEGER REFERENCES documents(id) ON DELETE SET NULL,
                status TEXT NOT NULL DEFAULT 'planned' CHECK(status IN ('planned','approved','paid','cancelled')),
                notes TEXT,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at INTEGER NOT NULL,
                updated_at INTEGER
            );

            CREATE TABLE IF NOT EXISTS daily_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                report_date TEXT NOT NULL,
                title TEXT NOT NULL,
                work_done TEXT NOT NULL DEFAULT '',
                workers_count INTEGER NOT NULL DEFAULT 0,
                equipment TEXT NOT NULL DEFAULT '',
                blockers TEXT NOT NULL DEFAULT '',
                next_steps TEXT NOT NULL DEFAULT '',
                progress_percent REAL,
                raw_input TEXT,
                is_client_visible INTEGER NOT NULL DEFAULT 1,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at INTEGER NOT NULL,
                updated_at INTEGER
            );

            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                action TEXT NOT NULL,
                entity TEXT,
                entity_id INTEGER,
                payload TEXT,
                created_at INTEGER NOT NULL
            );
            """
        )

        migrate_users_table(con)
        repair_legacy_user_references(con)

        ensure_columns(
            con,
            "users",
            {
                "email": "TEXT",
                "phone": "TEXT",
                "clerk_user_id": "TEXT",
                "status": "TEXT NOT NULL DEFAULT 'active'",
                "updated_at": "INTEGER",
            },
        )

        ensure_columns(
            con,
            "projects",
            {
                "client_id": "INTEGER REFERENCES clients(id) ON DELETE SET NULL",
                "customer_company_id": "INTEGER REFERENCES companies(id) ON DELETE SET NULL",
                "own_legal_entity_id": "INTEGER REFERENCES companies(id) ON DELETE SET NULL",
                "city": "TEXT",
                "region": "TEXT",
                "contract_no": "TEXT",
                "contract_date": "TEXT",
                "director_id": "INTEGER REFERENCES users(id) ON DELETE SET NULL",
                "foreman_id": "INTEGER REFERENCES users(id) ON DELETE SET NULL",
                "buyer_id": "INTEGER REFERENCES users(id) ON DELETE SET NULL",
                "client_user_id": "INTEGER REFERENCES users(id) ON DELETE SET NULL",
                "budget": "REAL NOT NULL DEFAULT 0",
                "paid": "REAL NOT NULL DEFAULT 0",
                "spent": "REAL NOT NULL DEFAULT 0",
                "internal_schedule_status": "TEXT NOT NULL DEFAULT 'draft'",
                "internal_schedule_version": "INTEGER NOT NULL DEFAULT 1",
                "internal_schedule_approved_at": "TEXT",
                "customer_schedule_status": "TEXT NOT NULL DEFAULT 'draft'",
                "customer_schedule_version": "INTEGER NOT NULL DEFAULT 1",
                "customer_schedule_approved_at": "TEXT",
                "schedule_generated_at": "TEXT",
                "description": "TEXT",
                "updated_at": "INTEGER",
            },
        )

        ensure_columns(
            con,
            "finance_entries",
            {
                "category": "TEXT",
                "payment_kind": "TEXT NOT NULL DEFAULT 'cash'",
                "vat_percent": "REAL NOT NULL DEFAULT 0",
                "planned_date": "TEXT",
                "paid_date": "TEXT",
                "counterparty_name": "TEXT",
                "company_id": "INTEGER REFERENCES companies(id) ON DELETE SET NULL",
                "document_id": "INTEGER REFERENCES documents(id) ON DELETE SET NULL",
                "status": "TEXT NOT NULL DEFAULT 'planned'",
                "notes": "TEXT",
                "created_by": "INTEGER REFERENCES users(id) ON DELETE SET NULL",
                "updated_at": "INTEGER",
            },
        )

        ensure_columns(
            con,
            "documents",
            {
                "original_name": "TEXT",
                "storage_name": "TEXT",
                "storage_path": "TEXT",
                "mime_type": "TEXT",
                "file_ext": "TEXT",
                "size_bytes": "INTEGER",
                "notes": "TEXT",
                "uploaded_by": "INTEGER REFERENCES users(id) ON DELETE SET NULL",
                "stage_id": "INTEGER REFERENCES work_stages(id) ON DELETE SET NULL",
                "template_code": "TEXT",
                "generated_by_system": "INTEGER NOT NULL DEFAULT 0",
                "updated_at": "INTEGER",
            },
        )

        ensure_columns(
            con,
            "estimate_items",
            {
                "stage_id": "INTEGER REFERENCES work_stages(id) ON DELETE SET NULL",
                "item_kind": "TEXT NOT NULL DEFAULT 'material'",
                "section_title": "TEXT",
                "article": "TEXT",
                "procurement_status": "TEXT NOT NULL DEFAULT 'Купить'",
                "warehouse_source": "TEXT",
                "warehouse_item_id": "INTEGER REFERENCES warehouse_items(id) ON DELETE SET NULL",
                "delivery_days": "INTEGER",
                "need_by_date": "TEXT",
                "notes": "TEXT",
                "updated_at": "INTEGER",
            },
        )

        ensure_columns(
            con,
            "warehouse_items",
            {
                "item_type": "TEXT NOT NULL DEFAULT 'material'",
                "category": "TEXT",
                "name": "TEXT NOT NULL DEFAULT ''",
                "sku": "TEXT",
                "unit": "TEXT NOT NULL DEFAULT 'шт'",
                "qty": "REAL NOT NULL DEFAULT 0",
                "condition_status": "TEXT",
                "updated_at": "INTEGER",
            },
        )

        ensure_columns(
            con,
            "work_stages",
            {
                "parent_id": "INTEGER",
                "stage_kind": "TEXT NOT NULL DEFAULT 'section'",
                "status_code": "TEXT NOT NULL DEFAULT 'not_started'",
                "customer_start": "TEXT",
                "customer_end": "TEXT",
                "notes": "TEXT",
                "is_client_visible": "INTEGER NOT NULL DEFAULT 1",
                "updated_at": "INTEGER",
            },
        )

        ensure_columns(
            con,
            "tasks",
            {
                "updated_at": "INTEGER",
            },
        )

        ensure_columns(
            con,
            "daily_logs",
            {
                "progress_percent": "REAL",
                "raw_input": "TEXT",
                "updated_at": "INTEGER",
            },
        )

        for code in ROLE_CODES:
            con.execute(
                """
                INSERT INTO roles (code, name, description)
                VALUES (?, ?, ?)
                ON CONFLICT(code) DO UPDATE SET
                    name = excluded.name,
                    description = excluded.description
                """,
                (code, ROLE_LABELS[code], ROLE_DESCRIPTIONS.get(code, "")),
            )

        user_rows = con.execute("SELECT id, role FROM users").fetchall()
        for row in user_rows:
            role_code = row["role"]
            role = con.execute("SELECT id FROM roles WHERE code = ?", (role_code,)).fetchone()
            if role:
                con.execute(
                    """
                    INSERT OR IGNORE INTO user_roles (user_id, role_id, created_at)
                    VALUES (?, ?, ?)
                    """,
                    (row["id"], role["id"], now_ts()),
                )

        if table_exists(con, "projects") and table_exists(con, "object_assignments"):
            con.execute(
                """
                INSERT OR IGNORE INTO object_assignments (object_id, user_id, role_code, responsibility, is_primary, assigned_by, assigned_at)
                SELECT id, foreman_id, 'foreman', 'РџСЂРѕСЂР°Р± РѕР±СЉРµРєС‚Р°', 1, director_id, COALESCE(created_at, ?)
                FROM projects
                WHERE foreman_id IS NOT NULL
                """,
                (now_ts(),),
            )
            con.execute(
                """
                INSERT OR IGNORE INTO user_project_access (user_id, project_id)
                SELECT foreman_id, id
                FROM projects
                WHERE foreman_id IS NOT NULL
                """
            )

        seed_warehouse_items(con)

        user_count = con.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        if user_count == 0:
            bootstrap_password = os.environ.get("PMBI_ADMIN_PASSWORD") or secrets.token_urlsafe(18)
            admin_cur = con.execute(
                """
                INSERT INTO users (login, password_hash, role, name, status, created_at, updated_at)
                VALUES (?, ?, 'director', 'Главный администратор', 'active', ?, ?)
                """,
                ("admin", hash_password(bootstrap_password), now_ts(), now_ts()),
            )
            director_role = con.execute("SELECT id FROM roles WHERE code = 'director'").fetchone()
            if director_role:
                con.execute(
                    """
                    INSERT OR IGNORE INTO user_roles (user_id, role_id, created_at)
                    VALUES (?, ?, ?)
                    """,
                    (admin_cur.lastrowid, director_role["id"], now_ts()),
                )
            con.commit()
            BOOTSTRAP_PATH.write_text(
                "PM.bi initial admin\n"
                "login: admin\n"
                f"password: {bootstrap_password}\n\n"
                "Delete this file after creating your real users.\n",
                encoding="utf-8",
            )

        con.commit()



def user_payload(row: sqlite3.Row) -> dict:
    role = normalize_role(row["role"])
    roles = [role]
    try:
        with db() as con:
            role_rows = con.execute(
                """
                SELECT r.code
                FROM user_roles ur
                JOIN roles r ON r.id = ur.role_id
                WHERE ur.user_id = ?
                ORDER BY r.id
                """,
                (row["id"],),
            ).fetchall()
            roles = [normalize_role(role_row["code"]) for role_row in role_rows] or roles
    except sqlite3.Error:
        roles = [role]
    return {
        "id": row["id"],
        "login": row["login"],
        "email": row["email"] if "email" in row.keys() else None,
        "phone": row["phone"] if "phone" in row.keys() else None,
        "clerkUserId": row["clerk_user_id"] if "clerk_user_id" in row.keys() else None,
        "role": role,
        "roles": sorted(set(roles)),
        "roleLabel": ROLE_LABELS.get(role, role),
        "name": row["name"],
    }


def clerk_enabled() -> bool:
    return bool(CLERK_PUBLISHABLE_KEY and CLERK_SECRET_KEY and CLERK_JWT_KEY)


def split_person_name(full_name: str) -> tuple[str, str]:
    cleaned = re.sub(r"\s+", " ", str(full_name or "").strip())
    if not cleaned:
        return "", ""
    parts = cleaned.split(" ", 1)
    return parts[0], parts[1] if len(parts) > 1 else ""


def decode_unverified_jwt(token: str) -> dict:
    try:
        return jwt.decode(token, options={"verify_signature": False, "verify_exp": False})
    except Exception:
        return {}


def verify_clerk_session_token(token: str) -> dict | None:
    if not clerk_enabled() or not token:
        return None
    try:
        claims = jwt.decode(
            token,
            CLERK_JWT_KEY,
            algorithms=["RS256"],
            options={"require": ["sub", "exp", "iat", "nbf"], "verify_aud": False},
        )
    except Exception:
        return None
    azp = str(claims.get("azp") or "").rstrip("/")
    if CLERK_AUTHORIZED_PARTIES and azp and azp not in CLERK_AUTHORIZED_PARTIES:
        return None
    return claims


def clerk_api_request(path: str, *, method: str = "GET", payload: dict | None = None) -> dict:
    if not CLERK_SECRET_KEY:
        raise ValueError("clerk_secret_key_missing")
    body = json.dumps(payload).encode("utf-8") if payload is not None else None
    request = urllib.request.Request(
        CLERK_API_BASE + path,
        data=body,
        method=method,
        headers={
            "Authorization": f"Bearer {CLERK_SECRET_KEY}",
            "Accept": "application/json",
            **({"Content-Type": "application/json"} if body is not None else {}),
        },
    )
    with urllib.request.urlopen(request, timeout=15) as response:
        raw = response.read().decode("utf-8") or "{}"
    return json.loads(raw)


def clerk_primary_email(user_data: dict) -> str | None:
    primary_id = user_data.get("primary_email_address_id")
    email_rows = user_data.get("email_addresses") or []
    for item in email_rows:
        if item.get("id") == primary_id:
            value = str(item.get("email_address") or "").strip().lower()
            if value:
                return value
    for item in email_rows:
        value = str(item.get("email_address") or "").strip().lower()
        if value:
            return value
    return None


def clerk_primary_phone(user_data: dict) -> str | None:
    primary_id = user_data.get("primary_phone_number_id")
    phone_rows = user_data.get("phone_numbers") or []
    for item in phone_rows:
        if item.get("id") == primary_id:
            value = str(item.get("phone_number") or "").strip()
            if value:
                return value
    for item in phone_rows:
        value = str(item.get("phone_number") or "").strip()
        if value:
            return value
    return None


def role_can_open(role: str, path: str) -> bool:
    if path in PUBLIC_STATIC_PATHS:
        return True
    allowed = ROLE_ALLOWED_PREFIXES.get(role, [])
    if "*" in allowed:
        return True
    return any(path == prefix or (prefix.endswith("/") and path.startswith(prefix)) for prefix in allowed)


def user_has_any_role(user: dict, roles: set[str]) -> bool:
    user_roles = {normalize_role(user.get("role"))}
    user_roles.update(normalize_role(role) for role in user.get("roles", []))
    return bool(user_roles & roles)


def user_can_open(user: dict, path: str) -> bool:
    user_roles = {normalize_role(user.get("role"))}
    user_roles.update(normalize_role(role) for role in user.get("roles", []))
    return any(role_can_open(role, path) for role in user_roles)


def user_can_manage_documents(user: dict) -> bool:
    return user_has_any_role(user, {"admin", "director", "foreman", "purchaser", "financier", "accountant"})


def user_can_manage_suppliers(user: dict) -> bool:
    return user_has_any_role(user, {"admin", "director", "foreman", "purchaser"})


def user_can_manage_finances(user: dict) -> bool:
    return user_has_any_role(user, {"admin", "director", "financier", "accountant"})


def user_can_view_finances(user: dict) -> bool:
    return user_can_manage_finances(user) or user_has_any_role(user, {"foreman"})


def user_can_pay_invoices(user: dict) -> bool:
    return user_has_any_role(user, {"director"})


def user_can_manage_schedule(user: dict) -> bool:
    return user_has_any_role(user, {"admin", "director", "foreman"})


def sanitize_filename(name: str) -> str:
    cleaned = re.sub(r"[^\w.\- ]+", "_", name or "").strip().strip(".")
    return cleaned or "document"


def project_documents_dir(project_id: int) -> Path:
    path = DOCUMENTS_DIR / f"project_{project_id}"
    path.mkdir(parents=True, exist_ok=True)
    return path


def document_extension(name: str) -> str:
    return Path(name or "").suffix.lower()[:16]


class FinanceExcelParseError(ValueError):
    pass


def excel_cell_ref_to_indexes(cell_ref: str) -> tuple[int, int]:
    match = re.match(r"^([A-Za-z]+)([1-9][0-9]*)$", str(cell_ref or "").strip())
    if not match:
        raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR)
    col = 0
    for char in match.group(1).upper():
        col = col * 26 + (ord(char) - ord("A") + 1)
    return int(match.group(2)) - 1, col - 1


def clean_excel_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def parse_excel_amount(value: object) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_excel_value(value)
    text = re.sub(r"[^\d,.\-]+", "", text).replace(",", ".")
    if not text:
        raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR)
    try:
        return float(text)
    except ValueError as error:
        raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR) from error


def normalize_excel_label(value: object) -> str:
    return re.sub(r"[^a-zа-яё0-9]+", "", clean_excel_value(value).lower())


def labeled_excel_value(rows: list[list[object]], labels: set[str]) -> object:
    normalized_labels = {normalize_excel_label(label) for label in labels}
    for row_index, row in enumerate(rows):
        for col_index, value in enumerate(row):
            label = normalize_excel_label(value)
            if not label or not any(expected in label for expected in normalized_labels):
                continue
            for next_col in range(col_index + 1, min(len(row), col_index + 5)):
                if clean_excel_value(row[next_col]):
                    return row[next_col]
            if row_index + 1 < len(rows) and col_index < len(rows[row_index + 1]):
                below = rows[row_index + 1][col_index]
                if clean_excel_value(below):
                    return below
    return None


def parse_finance_excel_invoice(raw: bytes, file_ext: str) -> dict[str, object]:
    if file_ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR) from error
        try:
            workbook = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
            sheet = workbook.active
            cell_values = {
                key: sheet[cell_ref].value
                for key, cell_ref in FINANCE_EXCEL_TEMPLATE_CELLS.items()
            }
            rows = [list(row) for row in sheet.iter_rows(min_row=1, max_row=40, max_col=12, values_only=True)]
            workbook.close()
        except Exception as error:
            raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR) from error
    elif file_ext == ".xls":
        try:
            import xlrd
        except ImportError as error:
            raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR) from error
        try:
            book = xlrd.open_workbook(file_contents=raw)
            sheet = book.sheet_by_index(0)
            cell_values = {}
            for key, cell_ref in FINANCE_EXCEL_TEMPLATE_CELLS.items():
                row_index, col_index = excel_cell_ref_to_indexes(cell_ref)
                cell_values[key] = sheet.cell_value(row_index, col_index) if row_index < sheet.nrows and col_index < sheet.ncols else None
            rows = [
                [sheet.cell_value(row_index, col_index) for col_index in range(min(sheet.ncols, 12))]
                for row_index in range(min(sheet.nrows, 40))
            ]
        except Exception as error:
            raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR) from error
    else:
        raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR)

    category = clean_excel_value(cell_values.get("category")) or clean_excel_value(
        labeled_excel_value(rows, {"наименование", "назначение", "счет", "счёт"})
    )
    counterparty = clean_excel_value(cell_values.get("counterparty_name")) or clean_excel_value(
        labeled_excel_value(rows, {"контрагент", "поставщик", "подрядчик"})
    )
    amount_value = cell_values.get("amount")
    if amount_value in (None, ""):
        amount_value = labeled_excel_value(rows, {"итоговая сумма", "итого", "сумма к оплате", "всего"})
    amount = parse_excel_amount(amount_value)
    if not category or not counterparty or amount <= 0:
        raise FinanceExcelParseError(FINANCE_EXCEL_PARSE_ERROR)
    return {"category": category, "counterparty_name": counterparty, "amount": amount}


def serialize_project(row: sqlite3.Row, user: dict) -> dict:
    data = dict(row)
    role = user["role"]
    if role == "customer":
        for key in ["budget", "paid", "spent", "director_id", "foreman_id", "buyer_id", "client_user_id"]:
            data.pop(key, None)
    elif role == "purchaser":
        for key in ["paid", "spent"]:
            data.pop(key, None)
    elif role == "foreman":
        data.pop("paid", None)
    try:
        with db() as con:
            assigned_rows = con.execute(
                """
                SELECT user_id
                FROM object_assignments
                WHERE object_id = ? AND role_code = 'foreman'
                ORDER BY is_primary DESC, assigned_at DESC, id DESC
                """,
                (row["id"],),
            ).fetchall()
            data["assigned_foremen"] = [assigned_row["user_id"] for assigned_row in assigned_rows]
    except sqlite3.Error:
        data["assigned_foremen"] = []
    data["scheduleControl"] = project_schedule_payload(row)
    return data


def project_schedule_payload(row: sqlite3.Row | dict) -> dict:
    getter = row.__getitem__ if isinstance(row, sqlite3.Row) else row.get
    return {
        "internal": {
            "status": str(getter("internal_schedule_status") or "draft"),
            "version": int(getter("internal_schedule_version") or 1),
            "approvedAt": getter("internal_schedule_approved_at"),
        },
        "customer": {
            "status": str(getter("customer_schedule_status") or "draft"),
            "version": int(getter("customer_schedule_version") or 1),
            "approvedAt": getter("customer_schedule_approved_at"),
        },
        "generatedAt": getter("schedule_generated_at"),
    }


def mark_project_schedule_draft(
    con: sqlite3.Connection,
    project_id: int,
    *,
    touch_internal: bool = True,
    touch_customer: bool = True,
    generated_at: str | None = None,
) -> None:
    project = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
    if not project:
        return

    updates: list[str] = []
    params: list[object] = []

    if touch_internal:
        internal_status = str(project["internal_schedule_status"] or "draft")
        internal_version = int(project["internal_schedule_version"] or 1)
        if internal_status == "approved":
            internal_version += 1
        updates.extend(
            [
                "internal_schedule_status = ?",
                "internal_schedule_version = ?",
                "internal_schedule_approved_at = NULL",
            ]
        )
        params.extend(["draft", internal_version])

    if touch_customer:
        customer_status = str(project["customer_schedule_status"] or "draft")
        customer_version = int(project["customer_schedule_version"] or 1)
        if customer_status == "approved":
            customer_version += 1
        updates.extend(
            [
                "customer_schedule_status = ?",
                "customer_schedule_version = ?",
                "customer_schedule_approved_at = NULL",
            ]
        )
        params.extend(["draft", customer_version])

    if generated_at is not None:
        updates.append("schedule_generated_at = ?")
        params.append(generated_at)

    if not updates:
        return

    updates.append("updated_at = ?")
    params.extend([now_ts(), project_id])
    con.execute(f"UPDATE projects SET {', '.join(updates)} WHERE id = ?", params)


def update_project_schedule_status(
    con: sqlite3.Connection,
    project_id: int,
    schedule_type: str,
    action: str,
) -> sqlite3.Row | None:
    project = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
    if not project:
        return None

    if action == "reset_to_draft":
        mark_project_schedule_draft(
            con,
            project_id,
            touch_internal=schedule_type in {"internal", "both"},
            touch_customer=schedule_type in {"customer", "both"},
        )
        return con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()

    updates: list[str] = []
    params: list[object] = []
    approved_at = TODAY_ISO
    if schedule_type in {"internal", "both"}:
        updates.extend(["internal_schedule_status = ?", "internal_schedule_approved_at = ?"])
        params.extend(["approved", approved_at])
    if schedule_type in {"customer", "both"}:
        updates.extend(["customer_schedule_status = ?", "customer_schedule_approved_at = ?"])
        params.extend(["approved", approved_at])
    updates.append("updated_at = ?")
    params.extend([now_ts(), project_id])
    con.execute(f"UPDATE projects SET {', '.join(updates)} WHERE id = ?", params)
    return con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()


def material_summary_rows(con: sqlite3.Connection, project_id: int) -> list[dict]:
    stage_rows = con.execute(
        """
        SELECT id, title, parent_id, stage_kind, position
        FROM work_stages
        WHERE project_id = ?
        ORDER BY position, id
        """,
        (project_id,),
    ).fetchall()
    stage_map = {int(row["id"]): row for row in stage_rows}

    def resolve_stage_root_title(stage_id: int | None) -> str | None:
        if not stage_id:
            return None
        current = stage_map.get(int(stage_id))
        root = None
        guard = 0
        while current and guard < 32:
            root = current
            parent_id = current["parent_id"]
            if not parent_id:
                break
            current = stage_map.get(int(parent_id))
            guard += 1
        return str(root["title"]).strip() if root and root["title"] else None

    rows = con.execute(
        """
        SELECT
            e.id,
            e.title,
            e.unit,
            e.planned_qty,
            e.planned_price,
            e.item_kind,
            e.section_title,
            e.article,
            e.procurement_status,
            e.warehouse_source,
            e.warehouse_item_id,
            e.delivery_days,
            e.need_by_date,
            e.notes,
            e.stage_id,
            ws.title AS stage_title,
            ws.planned_start AS stage_planned_start,
            ws.planned_end AS stage_planned_end,
            COALESCE(SUM(CASE WHEN s.move_type = 'purchase' THEN s.qty ELSE 0 END), 0) AS purchased_qty,
            COALESCE(SUM(CASE WHEN s.move_type = 'receipt' THEN s.qty ELSE 0 END), 0) AS received_qty,
            COALESCE(SUM(CASE WHEN s.move_type = 'use' THEN s.qty ELSE 0 END), 0) AS used_qty
        FROM estimate_items e
        LEFT JOIN work_stages ws ON ws.id = e.stage_id
        LEFT JOIN stock_moves s ON s.estimate_item_id = e.id
        WHERE e.project_id = ?
        GROUP BY e.id
        ORDER BY e.id
        """,
        (project_id,),
    ).fetchall()
    items = []
    for row in rows:
        planned = float(row["planned_qty"])
        purchased = float(row["purchased_qty"])
        received = float(row["received_qty"])
        used = float(row["used_qty"])
        covered = max(purchased, received)
        stock_base = received if received else purchased
        stock = max(stock_base - used, 0)
        missing = max(planned - covered, 0)
        usage_progress = round(min(100, used / planned * 100), 1) if planned else 0
        purchase_progress = round(min(100, covered / planned * 100), 1) if planned else 0
        estimated_delivery_days = estimate_material_lead_days(
            {
                "title": row["title"],
                "notes": row["notes"],
                "unit": row["unit"],
                "planned_qty": planned,
                "planned_price": float(row["planned_price"] or 0),
            }
        )
        delivery_days = int(row["delivery_days"]) if row["delivery_days"] is not None else int(estimated_delivery_days)
        need_by_date = str(row["need_by_date"] or row["stage_planned_start"] or row["stage_planned_end"] or "")
        soon_threshold = (parse_iso_date(TODAY_ISO) + timedelta(days=13)).isoformat()
        if missing <= 0:
            supply_status = "in_stock"
            supply_label = "Есть в наличии"
        elif need_by_date and need_by_date < TODAY_ISO:
            supply_status = "required"
            supply_label = "Требуется"
        elif need_by_date and need_by_date <= soon_threshold:
            supply_status = "soon"
            supply_label = "Скоро потребуется"
        else:
            supply_status = "planned"
            supply_label = "Нужно запланировать"
        items.append(
            {
                "id": row["id"],
                "title": row["title"],
                "itemKind": resolved_estimate_item_kind(row),
                "unit": row["unit"],
                "article": row["article"] or "",
                "plannedQty": planned,
                "plannedPrice": float(row["planned_price"] or 0),
                "purchasedQty": purchased,
                "receivedQty": received,
                "usedQty": used,
                "stockQty": stock,
                "missingQty": missing,
                "usageProgress": usage_progress,
                "purchaseProgress": purchase_progress,
                "needByDate": row["need_by_date"] or row["stage_planned_start"] or row["stage_planned_end"],
                "stageStartDate": row["stage_planned_start"],
                "stageEndDate": row["stage_planned_end"],
                "notes": row["notes"] or "",
                "procurementStatus": row["procurement_status"] or "",
                "warehouseSource": row["warehouse_source"] or "",
                "warehouseItemId": row["warehouse_item_id"],
                "deliveryDays": delivery_days,
                "estimatedDeliveryDays": int(estimated_delivery_days),
                "stageId": row["stage_id"],
                "stageTitle": row["stage_title"],
                "sectionTitle": str(resolved_estimate_section_title(row) or resolve_stage_root_title(row["stage_id"]) or "").strip(),
                "supplyStatus": supply_status,
                "supplyLabel": (row["procurement_status"] or supply_label) if row["warehouse_source"] else supply_label,
            }
        )
    return items


def serialize_warehouse_item(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "type": row["item_type"],
        "itemType": row["item_type"],
        "category": row["category"] or "",
        "name": row["name"],
        "sku": row["sku"] or "",
        "unit": row["unit"],
        "qty": float(row["qty"] or 0),
        "condition": row["condition_status"] or "",
        "conditionStatus": row["condition_status"] or "",
        "createdAt": row["created_at"],
        "updatedAt": row["updated_at"],
    }


def warehouse_item_rows(con: sqlite3.Connection, *, only_available: bool = False) -> list[sqlite3.Row]:
    where = "WHERE qty > 0" if only_available else ""
    return con.execute(
        f"""
        SELECT id, item_type, category, name, sku, unit, qty, condition_status, created_at, updated_at
        FROM warehouse_items
        {where}
        ORDER BY item_type, category, name, id
        """
    ).fetchall()


def warehouse_search_score(material: dict | sqlite3.Row, item: sqlite3.Row) -> float:
    def value(source: dict | sqlite3.Row, *keys: str) -> object:
        if isinstance(source, sqlite3.Row):
            for key in keys:
                if key in source.keys():
                    return source[key]
            return None
        for key in keys:
            if key in source:
                return source.get(key)
        return None

    material_article = str(value(material, "article", "sku") or "").strip()
    item_sku = str(item["sku"] or "").strip()
    if material_article and item_sku and normalize_fuzzy_text(material_article) == normalize_fuzzy_text(item_sku):
        return 1.0
    title = value(material, "title", "name") or ""
    return max(
        fuzzy_similarity(title, item["name"]),
        fuzzy_similarity(material_article, item_sku) if material_article and item_sku else 0.0,
    )


def best_warehouse_match(
    material: dict | sqlite3.Row,
    warehouse_items: list[sqlite3.Row],
    *,
    threshold: float = FUZZY_MATCH_THRESHOLD,
) -> dict | None:
    best_row = None
    best_score = 0.0
    for item in warehouse_items:
        if item["item_type"] != "material" or float(item["qty"] or 0) <= 0:
            continue
        score = warehouse_search_score(material, item)
        if score > best_score:
            best_score = score
            best_row = item
    if not best_row or best_score < threshold:
        return None
    payload = serialize_warehouse_item(best_row)
    payload["score"] = round(best_score, 3)
    return payload


def warehouse_matches_for_project(con: sqlite3.Connection, project_id: int) -> dict[int, dict]:
    warehouse_items = warehouse_item_rows(con, only_available=True)
    matches: dict[int, dict] = {}
    for material in material_summary_rows(con, project_id):
        if normalize_estimate_item_kind(material.get("itemKind")) == "work":
            continue
        if float(material.get("missingQty") or 0) <= 0 and not str(material.get("procurementStatus") or "").lower().startswith("куп"):
            continue
        match = best_warehouse_match(material, warehouse_items)
        if match:
            matches[int(material["id"])] = match
    return matches


def find_warehouse_receipt_target(
    con: sqlite3.Connection,
    *,
    item_type: str,
    name: str,
    unit: str,
    sku: str | None = None,
    warehouse_item_id: int | None = None,
) -> sqlite3.Row | None:
    if warehouse_item_id:
        row = con.execute(
            """
            SELECT id, item_type, category, name, sku, unit, qty, condition_status, created_at, updated_at
            FROM warehouse_items
            WHERE id = ?
            """,
            (warehouse_item_id,),
        ).fetchone()
        if row:
            return row

    normalized_sku = normalize_fuzzy_text(sku or "")
    normalized_name = normalize_fuzzy_text(name)
    rows = warehouse_item_rows(con)
    if normalized_sku:
        for row in rows:
            if normalize_fuzzy_text(row["sku"] or "") == normalized_sku:
                return row
    for row in rows:
        if row["item_type"] == item_type and normalize_fuzzy_text(row["name"]) == normalized_name and str(row["unit"] or "") == unit:
            return row

    best_row = None
    best_score = 0.0
    for row in rows:
        if row["item_type"] != item_type:
            continue
        score = fuzzy_similarity(name, row["name"])
        if score > best_score:
            best_score = score
            best_row = row
    return best_row if best_row and best_score >= FUZZY_MATCH_THRESHOLD else None


def add_qty_to_warehouse(
    con: sqlite3.Connection,
    *,
    item_type: str,
    name: str,
    qty: float,
    unit: str,
    sku: str | None = None,
    condition_status: str | None = None,
    category: str | None = None,
    warehouse_item_id: int | None = None,
) -> sqlite3.Row:
    target = find_warehouse_receipt_target(
        con,
        item_type=item_type,
        name=name,
        unit=unit,
        sku=sku,
        warehouse_item_id=warehouse_item_id,
    )
    if target:
        con.execute(
            """
            UPDATE warehouse_items
            SET qty = qty + ?,
                unit = COALESCE(NULLIF(?, ''), unit),
                condition_status = CASE WHEN ? != '' THEN ? ELSE condition_status END,
                updated_at = ?
            WHERE id = ?
            """,
            (qty, unit, condition_status or "", condition_status or "", now_ts(), target["id"]),
        )
        return con.execute(
            """
            SELECT id, item_type, category, name, sku, unit, qty, condition_status, created_at, updated_at
            FROM warehouse_items
            WHERE id = ?
            """,
            (target["id"],),
        ).fetchone()

    cur = con.execute(
        """
        INSERT INTO warehouse_items (item_type, category, name, sku, unit, qty, condition_status, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            item_type,
            category,
            name,
            sku or None,
            unit,
            qty,
            condition_status or ("" if item_type == "material" else "Б/У"),
            now_ts(),
            now_ts(),
        ),
    )
    return con.execute(
        """
        SELECT id, item_type, category, name, sku, unit, qty, condition_status, created_at, updated_at
        FROM warehouse_items
        WHERE id = ?
        """,
        (cur.lastrowid,),
    ).fetchone()


def seed_warehouse_items(con: sqlite3.Connection) -> None:
    if con.execute("SELECT COUNT(*) FROM warehouse_items").fetchone()[0] > 0:
        return
    rows = [
        ("material", "Кабель и электрика", "Кабель ВВГнг 3х2.5", "EL-VVGNG-3X2.5", "м", 240.0, ""),
        ("material", "Металлопрокат", "Швеллер 10П", "MET-SHV-10P", "м", 36.0, ""),
        ("material", "Изоляция", "Лента ФУМ", "SAN-FUM-12", "рулон", 18.0, ""),
        ("material", "Крепеж", "Саморез по металлу 4.2х19", "KRP-4219", "шт", 1200.0, ""),
        ("tool", "Инструмент", "Перфоратор Bosch GBH 2-26", "TOOL-GBH-226", "шт", 3.0, "Б/У"),
        ("tool", "Инструмент", "Лазерный уровень DeWalt", "TOOL-DW-LASER", "шт", 1.0, "Новый"),
        ("tool", "Инструмент", "Шуруповерт Makita DDF", "TOOL-MAK-DDF", "шт", 2.0, "Требует ремонта"),
    ]
    con.executemany(
        """
        INSERT INTO warehouse_items (item_type, category, name, sku, unit, qty, condition_status, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [(item_type, category, name, sku, unit, qty, condition, now_ts(), now_ts()) for item_type, category, name, sku, unit, qty, condition in rows],
    )


def build_material_schedule_payload(
    con: sqlite3.Connection,
    project_id: int,
    *,
    warning_days: int = 5,
    neutral_days: int = 7,
) -> dict:
    project = con.execute(
        "SELECT id, title, started_at, deadline_at FROM projects WHERE id = ?",
        (project_id,),
    ).fetchone()
    materials = [
        item
        for item in material_summary_rows(con, project_id)
        if normalize_estimate_item_kind(item.get("itemKind")) != "work"
    ]
    selected_offers = con.execute(
        """
        SELECT estimate_item_id, candidate_name, company_id, source_url
        FROM supplier_offers
        WHERE project_id = ? AND candidate_type = 'supplier' AND status = 'selected'
        ORDER BY updated_at DESC, created_at DESC, id DESC
        """,
        (project_id,),
    ).fetchall()
    supplier_by_item: dict[int, dict] = {}
    for row in selected_offers:
        item_id = row["estimate_item_id"]
        if item_id is None or int(item_id) in supplier_by_item:
            continue
        supplier_by_item[int(item_id)] = {
            "name": row["candidate_name"] or "",
            "companyId": row["company_id"],
            "sourceUrl": row["source_url"] or "",
        }

    today = parse_iso_date(TODAY_ISO) or date.today()
    items: list[dict] = []
    range_dates: list[date] = [today]
    summary = {
        "total": 0,
        "purchased": 0,
        "overdue": 0,
        "warning": 0,
        "neutral": 0,
        "unscheduled": 0,
    }

    for material in materials:
        material_id = int(material.get("id") or 0)
        planned_qty = float(material.get("plannedQty") or 0)
        planned_price = float(material.get("plannedPrice") or 0)
        missing_qty = float(material.get("missingQty") or 0)
        estimated_lead_days = estimate_material_lead_days(
            {
                "title": material.get("title"),
                "notes": material.get("notes"),
                "unit": material.get("unit"),
                "planned_qty": planned_qty,
                "planned_price": planned_price,
            }
        )
        try:
            lead_days = int(material.get("deliveryDays") if material.get("deliveryDays") is not None else estimated_lead_days)
        except (TypeError, ValueError):
            lead_days = int(estimated_lead_days)
        lead_days = max(0, min(lead_days, 90))

        deadline_date = parse_iso_date(str(material.get("needByDate") or material.get("stageStartDate") or material.get("stageEndDate") or ""))
        purchase_start = deadline_date - timedelta(days=lead_days) if deadline_date else None
        alert_start = purchase_start - timedelta(days=warning_days) if purchase_start else None

        days_until_purchase = (purchase_start - today).days if purchase_start else None
        days_until_deadline = (deadline_date - today).days if deadline_date else None
        if missing_qty <= 0:
            if float(material.get("receivedQty") or 0) >= planned_qty:
                status = "purchased"
                status_label = "Закуплено"
            else:
                status = "in_transit"
                status_label = "В пути"
            color = "done"
            summary["purchased"] += 1
        elif not deadline_date:
            status = "unscheduled"
            status_label = "Нет дедлайна"
            color = "muted"
            summary["unscheduled"] += 1
        elif days_until_purchase is not None and days_until_purchase < 0:
            status = "overdue"
            status_label = "Просрочено"
            color = "red"
            summary["overdue"] += 1
        elif days_until_purchase is not None and days_until_purchase <= warning_days:
            status = "warning"
            status_label = "Пора платить"
            color = "yellow"
            summary["warning"] += 1
        else:
            status = "neutral"
            status_label = "В плане"
            color = "green"
            summary["neutral"] += 1

        for value in (purchase_start, deadline_date, alert_start):
            if value:
                range_dates.append(value)

        summary["total"] += 1
        items.append(
            {
                "id": material_id,
                "projectId": project_id,
                "title": material.get("title") or "",
                "unit": material.get("unit") or "",
                "plannedQty": planned_qty,
                "purchasedQty": float(material.get("purchasedQty") or 0),
                "receivedQty": float(material.get("receivedQty") or 0),
                "missingQty": missing_qty,
                "purchaseProgress": float(material.get("purchaseProgress") or 0),
                "status": status,
                "statusLabel": status_label,
                "color": color,
                "purchaseStartDate": purchase_start.isoformat() if purchase_start else None,
                "purchaseByDate": purchase_start.isoformat() if purchase_start else None,
                "alertStartDate": alert_start.isoformat() if alert_start else None,
                "deadlineDate": deadline_date.isoformat() if deadline_date else None,
                "deliveryTargetDate": deadline_date.isoformat() if deadline_date else None,
                "deliveryLeadDays": int(lead_days),
                "estimatedDeliveryDays": int(estimated_lead_days),
                "warningDays": int(warning_days),
                "daysUntilPurchase": int(days_until_purchase) if days_until_purchase is not None else None,
                "daysUntilDeadline": int(days_until_deadline) if days_until_deadline is not None else None,
                "sectionTitle": material.get("sectionTitle") or "",
                "relatedWork": {
                    "stageId": material.get("stageId"),
                    "title": material.get("stageTitle") or material.get("sectionTitle") or "",
                    "startDate": material.get("stageStartDate"),
                    "endDate": material.get("stageEndDate"),
                },
                "supplier": supplier_by_item.get(material_id),
                "materialUrl": f"/app/projects?openProject={project_id}&tab=materials&materialId={material_id}",
            }
        )

    items.sort(key=lambda item: (item["deadlineDate"] or "9999-12-31", item["purchaseStartDate"] or "9999-12-31", item["title"]))
    start_date = min(range_dates)
    end_date = max(range_dates)
    if start_date == end_date:
        end_date = start_date + timedelta(days=max(neutral_days, warning_days, 1))
    return {
        "projectId": project_id,
        "projectTitle": project["title"] if project else "",
        "today": today.isoformat(),
        "settings": {
            "warningDays": int(warning_days),
            "neutralDays": int(neutral_days),
        },
        "range": {
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
        },
        "summary": summary,
        "items": items,
    }


def normalize_market_title_key(value: str | None) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"[^0-9a-zа-я]+", " ", text, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", text).strip()


def extract_estimate_id_from_project(project: sqlite3.Row | dict) -> str | None:
    def value(key: str) -> str:
        if isinstance(project, sqlite3.Row):
            return str(project[key] or "")
        return str(project.get(key, "") or "")

    contract_no = value("contract_no")
    match = re.search(r"ESTIMATE-([A-Za-z0-9]+)", contract_no)
    if match:
        return match.group(1)
    description = value("description")
    match = re.search(r"/estimates/([A-Za-z0-9]+)", description)
    if match:
        return match.group(1)
    return None


def estimate_position_from_notes(notes: str | None, estimate_id: str | None = None) -> int | None:
    raw = str(notes or "")
    if not raw:
        return None
    patterns = []
    if estimate_id:
        patterns.append(re.escape(estimate_id) + r";\s*[^;]*:\s*(\d+)")
    patterns.append(r";\s*[^;]*:\s*(\d+);\s*[^;]*Excel")
    for pattern in patterns:
        match = re.search(pattern, raw, re.IGNORECASE)
        if match:
            try:
                return int(match.group(1))
            except (TypeError, ValueError):
                pass
    return None


def strip_html_fragment(value: str | None) -> str:
    text = re.sub(r"<[^>]+>", " ", str(value or ""))
    text = html.unescape(text)
    return re.sub(r"\s+", " ", text).strip()


def parse_money_value(value: str | None) -> float | None:
    raw = str(value or "").strip().lower()
    if not raw or raw in {"—", "-", "nan", "none"}:
        return None
    raw = raw.replace("\xa0", " ").replace("₽", "").replace("руб.", "").replace("руб", "")
    raw = re.sub(r"[^0-9,.; -]", "", raw)
    numbers = []
    for chunk in re.split(r"[;|/]", raw):
        cleaned = chunk.replace(" ", "").replace(",", ".")
        if not cleaned:
            continue
        try:
            numbers.append(float(cleaned))
        except ValueError:
            continue
    if not numbers:
        return None
    return min(numbers)


def offer_domain_label(url: str | None) -> str:
    try:
        host = urllib.parse.urlparse(str(url or "")).netloc.lower()
    except ValueError:
        return ""
    return host.removeprefix("www.")


def parse_market_view_html(html_text: str, market_type: str) -> list[dict]:
    articles = re.findall(r"<article class=\"item\">(.*?)</article>", html_text, flags=re.S | re.I)
    rows: list[dict] = []
    for article in articles:
        title_match = re.search(r"<div class=\"item-title\">(.*?)</div>", article, flags=re.S | re.I)
        index_match = re.search(r"<div class=\"item-index\">(.*?)</div>", article, flags=re.S | re.I)
        meta_match = re.search(r"<div class=\"meta\">(.*?)</div>\s*(?:<div class=\"offers\">|<div class=\"status-note\">)", article, flags=re.S | re.I)
        offers_match = re.search(r"<div class=\"offers\">(.*?)</div>\s*(?:<div class=\"status-note\">|</article>)", article, flags=re.S | re.I)
        status_match = re.search(r"<div class=\"status-note\">(.*?)</div>", article, flags=re.S | re.I)
        kind_match = re.search(r"<span class=\"tag\">(.*?)</span>\s*</div>\s*<div class=\"meta\">", article, flags=re.S | re.I)
        title = strip_html_fragment(title_match.group(1) if title_match else "")
        if not title:
            continue
        try:
            position_index = int(strip_html_fragment(index_match.group(1) if index_match else "0") or 0)
        except ValueError:
            position_index = 0
        meta_tags = [
            strip_html_fragment(chunk)
            for chunk in re.findall(r"<span class=\"tag\">(.*?)</span>", meta_match.group(1) if meta_match else "", flags=re.S | re.I)
        ]
        meta_map: dict[str, str] = {}
        for tag in meta_tags:
            if ":" not in tag:
                continue
            key, value = tag.split(":", 1)
            meta_map[key.strip().lower()] = value.strip()

        offers: list[dict] = []
        offers_html = offers_match.group(1) if offers_match else ""
        for block in re.split(r"<div class=\"offer\">", offers_html, flags=re.I)[1:]:
            link_match = re.search(r"<a href=\"([^\"]+)\"[^>]*>(.*?)</a>", block, flags=re.S | re.I)
            price_match = re.search(r"<div class=\"num\">(.*?)</div>", block, flags=re.S | re.I)
            snippet_match = re.search(r"<div class=\"offer-snippet\">(.*?)</div>", block, flags=re.S | re.I)
            url = html.unescape(link_match.group(1)) if link_match else ""
            title_text = strip_html_fragment(link_match.group(2) if link_match else "")
            price_value = parse_money_value(strip_html_fragment(price_match.group(1) if price_match else ""))
            offers.append(
                {
                    "title": title_text or offer_domain_label(url) or "Источник",
                    "url": url,
                    "domain": offer_domain_label(url),
                    "price": price_value,
                    "snippet": strip_html_fragment(snippet_match.group(1) if snippet_match else ""),
                }
            )

        estimate_unit_price = parse_money_value(meta_map.get("смета за ед.") or meta_map.get("смета за ед"))
        estimate_total = parse_money_value(meta_map.get("смета всего"))
        market_price = parse_money_value(meta_map.get("рынок"))
        if market_price is None:
            offer_prices = [offer["price"] for offer in offers if offer.get("price") is not None]
            market_price = min(offer_prices) if offer_prices else None

        rows.append(
            {
                "positionIndex": position_index or None,
                "title": title,
                "titleKey": normalize_market_title_key(title),
                "marketType": market_type,
                "kindLabel": strip_html_fragment(kind_match.group(1) if kind_match else ""),
                "qtyText": meta_map.get("кол-во", ""),
                "unitText": meta_map.get("ед.", ""),
                "estimateUnitPrice": estimate_unit_price,
                "estimateTotal": estimate_total,
                "marketPrice": market_price,
                "marketPriceText": meta_map.get("рынок", ""),
                "offers": offers,
                "statusNote": strip_html_fragment(status_match.group(1) if status_match else ""),
            }
        )
    return rows


def fetch_autobot_market_rows(estimate_id: str, market_type: str) -> list[dict]:
    url = f"{PMBI_AUTOBOT_BASE_URL}/estimates/{urllib.parse.quote(estimate_id)}/market-view?market_type={urllib.parse.quote(market_type)}"
    request = urllib.request.Request(url, headers={"User-Agent": "PM.bi/1.0", "Accept": "text/html"})
    with urllib.request.urlopen(request, timeout=20) as response:
        charset = response.headers.get_content_charset() or "utf-8"
        html_text = response.read().decode(charset, errors="replace")
    return parse_market_view_html(html_text, market_type)


def build_project_market_analysis(con: sqlite3.Connection, project_id: int, kind: str) -> dict:
    project = con.execute("SELECT id, contract_no, description FROM projects WHERE id = ?", (project_id,)).fetchone()
    if not project:
        raise ValueError("project_not_found")
    estimate_id = extract_estimate_id_from_project(project)
    if not estimate_id:
        raise LookupError("estimate_not_linked")

    market_types = ["work"] if kind == "work" else ["material", "product", "service", "other"]
    market_rows: list[dict] = []
    for market_type in market_types:
        market_rows.extend(fetch_autobot_market_rows(estimate_id, market_type))

    items = material_summary_rows(con, project_id)
    if kind == "work":
        items = [item for item in items if normalize_estimate_item_kind(item.get("itemKind")) == "work"]
    else:
        items = [item for item in items if normalize_estimate_item_kind(item.get("itemKind")) != "work"]

    market_by_index: dict[int, dict] = {}
    market_by_title: dict[str, dict] = {}
    for row in market_rows:
        if row.get("positionIndex"):
            market_by_index[int(row["positionIndex"])] = row
        market_by_title[row["titleKey"]] = row

    merged_rows: list[dict] = []
    for item in items:
        position_index = estimate_position_from_notes(str(item.get("notes") or ""), estimate_id)
        title_key = normalize_market_title_key(str(item.get("title") or ""))
        market_row = None
        if kind == "work" and position_index and position_index in market_by_index:
            market_row = market_by_index[position_index]
        if not market_row and title_key and title_key in market_by_title:
            market_row = market_by_title[title_key]
        offers = list((market_row or {}).get("offers") or [])
        market_price = (market_row or {}).get("marketPrice")
        estimate_unit_price = float(item.get("plannedPrice") or 0)
        merged_rows.append(
            {
                "estimateItemId": int(item.get("id") or 0),
                "positionIndex": position_index,
                "title": str(item.get("title") or ""),
                "titleKey": title_key,
                "sectionTitle": str(item.get("sectionTitle") or "").strip(),
                "stageTitle": str(item.get("stageTitle") or "").strip(),
                "unit": str(item.get("unit") or "").strip(),
                "plannedQty": float(item.get("plannedQty") or 0),
                "estimateUnitPrice": estimate_unit_price,
                "estimateTotal": round(float(item.get("plannedQty") or 0) * estimate_unit_price, 2),
                "marketPrice": market_price,
                "marketPriceText": (market_row or {}).get("marketPriceText") or "",
                "marketType": (market_row or {}).get("marketType") or "",
                "statusNote": (market_row or {}).get("statusNote") or "",
                "sources": offers[:3],
                "sourceCount": len(offers),
                "deltaPerUnit": round(market_price - estimate_unit_price, 2) if market_price is not None and estimate_unit_price else None,
                "hasMarketData": market_row is not None,
            }
        )

    merged_rows.sort(
        key=lambda row: (
            str(row.get("sectionTitle") or ""),
            int(row.get("positionIndex") or 10**9),
            str(row.get("title") or ""),
        )
    )
    coverage = sum(1 for row in merged_rows if row["hasMarketData"])
    return {
        "projectId": project_id,
        "estimateId": estimate_id,
        "kind": kind,
        "rows": merged_rows,
        "summary": {
            "total": len(merged_rows),
            "withMarketData": coverage,
            "withoutMarketData": max(0, len(merged_rows) - coverage),
        },
    }


def parse_iso_date(value: str | None) -> date | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return date.fromisoformat(raw[:10])
    except ValueError:
        return None


def count_threshold_hits(value: float, thresholds: list[float]) -> int:
    return sum(1 for threshold in thresholds if value >= threshold)


def classify_scope(text: str | None) -> str:
    haystack = str(text or "").strip().lower()
    for category, keywords in STAGE_CATEGORY_KEYWORDS.items():
        if any(keyword in haystack for keyword in keywords):
            return category
    return "general"


def executive_templates_for_stage(stage: sqlite3.Row | dict) -> list[dict]:
    title = str(stage["title"] if isinstance(stage, sqlite3.Row) else stage.get("title", ""))
    stage_kind = str(stage["stage_kind"] if isinstance(stage, sqlite3.Row) else stage.get("stage_kind", "work") or "work")
    progress = int(stage["progress"] if isinstance(stage, sqlite3.Row) else stage.get("progress", 0) or 0)
    status_code = str(stage["status_code"] if isinstance(stage, sqlite3.Row) else stage.get("status_code", "not_started") or "not_started")
    category = classify_scope(title)
    if stage_kind == "section":
        return []

    templates: list[dict] = []
    for code, config in EXECUTIVE_TEMPLATE_RULES.items():
        categories = set(config.get("categories", set()))
        is_match = category in categories or "general" in categories
        if not is_match:
            continue
        if code == "technical_solution":
            if status_code not in {"problem", "paused", "blocked"} and progress < 25:
                continue
        elif status_code == "not_started" and progress <= 0:
            continue
        templates.append(
            {
                "code": code,
                "title": config["title"],
                "doc_type": config["doc_type"],
                "optional": bool(config.get("optional")),
                "default_notes": config.get("default_notes", ""),
            }
        )
    return templates


def executive_ready_status(status: str | None) -> bool:
    return str(status or "").strip() in {"reviewed", "approved", "signed", "ready"}


def estimate_stage_duration(stage: sqlite3.Row | dict, materials: list[dict]) -> int:
    title = str(stage["title"] if isinstance(stage, sqlite3.Row) else stage.get("title", "")).strip().lower()
    stage_kind = str(stage["stage_kind"] if isinstance(stage, sqlite3.Row) else stage.get("stage_kind", "work")).strip() or "work"
    category = classify_scope(title)
    base_map = {
        "prep": 3,
        "concrete": 8,
        "masonry": 7,
        "roof": 6,
        "facade": 7,
        "electrical": 8,
        "plumbing": 8,
        "ventilation": 9,
        "finishing": 10,
        "landscape": 5,
        "handover": 3,
        "general": 5,
    }
    base = base_map.get(category, 5)
    if stage_kind == "section":
        base = max(2, base - 2)
    elif stage_kind == "subsection":
        base = max(3, base - 1)
    total_cost = sum(float(item.get("planned_qty", 0) or 0) * float(item.get("planned_price", 0) or 0) for item in materials)
    total_qty = sum(float(item.get("planned_qty", 0) or 0) for item in materials)
    material_count = len(materials)
    duration = (
        base
        + count_threshold_hits(total_cost, [150_000, 400_000, 900_000, 1_800_000, 3_500_000])
        + count_threshold_hits(total_qty, [20, 60, 140, 260, 500])
        + count_threshold_hits(material_count, [2, 5, 9, 15])
    )
    if category == "handover":
        duration = max(2, duration - 1)
    return max(2 if stage_kind != "work" else 3, min(duration, 32 if stage_kind == "work" else 40))


def estimate_material_lead_days(material: dict) -> int:
    category = classify_scope(" ".join([
        str(material.get("title", "")),
        str(material.get("notes", "")),
        str(material.get("unit", "")),
    ]))
    base_map = {
        "facade": 16,
        "electrical": 12,
        "plumbing": 12,
        "ventilation": 16,
        "finishing": 8,
        "roof": 10,
        "concrete": 6,
        "masonry": 7,
        "prep": 4,
        "landscape": 5,
        "handover": 3,
        "general": 7,
    }
    planned_qty = float(material.get("planned_qty", material.get("plannedQty", 0)) or 0)
    planned_price = float(material.get("planned_price", material.get("plannedPrice", 0)) or 0)
    amount = planned_qty * planned_price
    extra = count_threshold_hits(amount, [250_000, 700_000, 1_500_000])
    return min(24, base_map.get(category, 7) + extra)


def build_procurement_alerts(
    materials: list[dict],
    stages: list[sqlite3.Row],
    today_date: date,
    section_start_dates: dict[str, str] | None = None,
) -> dict:
    stage_map = {int(row["id"]): dict(row) for row in stages}
    section_start_dates = section_start_dates or {}
    alerts: list[dict] = []
    summary = {"critical": 0, "soon": 0, "watch": 0}
    for material in materials:
        if normalize_estimate_item_kind(material.get("itemKind", material.get("item_kind"))) == "work":
            continue
        missing_qty = float(material.get("missingQty", material.get("missing_qty", 0)) or 0)
        if missing_qty <= 0:
            continue
        raw_stage_id = material.get("stageId", material.get("stage_id"))
        try:
            stage_id = int(raw_stage_id) if raw_stage_id not in (None, "", 0, "0") else 0
        except (TypeError, ValueError):
            stage_id = 0
        stage = stage_map.get(stage_id)
        section_title = str(material.get("sectionTitle") or "").strip()
        stage_start = parse_iso_date(str((stage or {}).get("planned_start") or material.get("stageStartDate") or ""))
        if not stage_start and section_title:
            stage_start = parse_iso_date(section_start_dates.get(section_title))
        if not stage_start:
            continue
        lead_days = estimate_material_lead_days(material)
        order_by = parse_iso_date(str(material.get("needByDate", material.get("need_by_date", "")) or ""))
        if not order_by:
            order_by = stage_start - timedelta(days=lead_days)
        days_until_start = (stage_start - today_date).days
        days_until_order = (order_by - today_date).days
        if days_until_start > 30 and days_until_order > 14:
            continue
        if days_until_order < 0:
            status = "critical"
            summary["critical"] += 1
        elif days_until_order <= 3:
            status = "critical"
            summary["critical"] += 1
        elif days_until_order <= 10:
            status = "soon"
            summary["soon"] += 1
        else:
            status = "watch"
            summary["watch"] += 1
        action_window_days = max(0, days_until_order)
        alerts.append(
            {
                "materialId": int(material.get("id") or 0),
                "title": str(material.get("title") or ""),
                "unit": str(material.get("unit") or ""),
                "missingQty": missing_qty,
                "sectionTitle": section_title or str((stage or {}).get("title") or "").strip(),
                "stageTitle": str(material.get("stageTitle") or (stage or {}).get("title") or section_title or "").strip(),
                "startDate": stage_start.isoformat(),
                "orderByDate": order_by.isoformat(),
                "leadDays": int(lead_days),
                "daysUntilStart": int(days_until_start),
                "daysUntilOrder": int(days_until_order),
                "actionWindowDays": int(action_window_days),
                "status": status,
            }
        )
    alerts.sort(key=lambda item: (item["daysUntilOrder"], item["daysUntilStart"], item["title"]))
    return {"items": alerts[:10], "summary": summary}


SECTION_SCHEDULE_RULES = [
    {
        "keywords": ("демонтаж оконных коробок", "каменных стенах"),
        "hours_per_qty": 128.73,
        "crew_size": 4,
        "source_label": "ФЕРр 56-9-1: демонтаж оконных коробок в каменных стенах",
        "source_url": "https://www.defsmeta.com/rfer/rfer56/rfer-56-09-001-01.php",
        "assumption": False,
    },
    {
        "keywords": ("установка в жилых и общественных зданиях оконных блоков из пвх профилей", "трехстворчатых"),
        "hours_per_qty": 149.16,
        "crew_size": 4,
        "source_label": "ФЕР 10-01-034-08: установка трехстворчатых ПВХ-окон",
        "source_url": "https://www.defsmeta.com/rfer/rfer10/rfer-10-01-034-08.php",
        "assumption": False,
    },
    {
        "keywords": ("установка подоконных досок из пвх",),
        "hours_per_qty": 21.19,
        "crew_size": 3,
        "source_label": "ФЕР 10-01-035-01: установка подоконных досок из ПВХ",
        "source_url": "https://files.stroyinf.ru/Data2/1/4293814/4293814969.htm",
        "assumption": False,
    },
    {
        "keywords": ("установка уголков пвх на клее",),
        "hours_per_qty": 6.7,
        "crew_size": 2,
        "source_label": "ФЕР 15-01-070-01: установка уголков ПВХ на клее",
        "source_url": "https://www.defsmeta.com/rfer15_2/rfer-15-01-070-01.php",
        "assumption": False,
    },
    {
        "keywords": ("устройство вентилируемых фасадов", "без теплоизоляционного слоя"),
        "hours_per_qty": 207.98,
        "crew_size": 4,
        "source_label": "ФЕР 15-01-090-01: вентфасад без теплоизоляции",
        "source_url": "https://www.defsmeta.com/rfer15_2/rfer-15-01-090-01.php",
        "assumption": False,
    },
    {
        "keywords": ("устройство вентилируемых фасадов", "с устройством теплоизоляционного слоя"),
        "hours_per_qty": 334.66,
        "crew_size": 4,
        "source_label": "ФЕР 15-01-090-02: вентфасад с теплоизоляцией",
        "source_url": "https://www.defsmeta.com/rfer15_2/rfer-15-01-090-02.php",
        "assumption": False,
    },
    {
        "keywords": ("устройство стяжек: цементных толщиной 20 мм",),
        "hours_per_qty": 39.51,
        "crew_size": 4,
        "source_label": "ГЭСН 11-01-011-01: цементная стяжка 20 мм",
        "source_url": "https://cs.smetnoedelo.ru/gesn2/gesn11-01-011-01.html",
        "assumption": False,
    },
    {
        "keywords": ("на каждые 5 мм изменения толщины стяжки",),
        "hours_per_qty": 0.5,
        "crew_size": 4,
        "source_label": "ГЭСН 11-01-011-09: добавка на каждые 5 мм стяжки",
        "source_url": "https://fgisrf.ru/frsn/fer/element/4f0007cc-f11b-4d16-98ed-cdf8494fd866",
        "assumption": False,
    },
    {
        "keywords": ("самовыравнивающейся смеси", "толщиной 3 мм"),
        "hours_per_qty": 26.14,
        "crew_size": 3,
        "source_label": "ГЭСН 11-01-011-10: самовыравнивающаяся смесь 3 мм",
        "source_url": "https://cs.smetnoedelo.ru/gesn2/gesn11-01-011-10.html",
        "assumption": False,
    },
    {
        "keywords": ("на каждый последующий слой толщиной 1 мм",),
        "hours_per_qty": 2.33,
        "crew_size": 3,
        "source_label": "ГЭСН 11-01-011-11: добавка на каждый 1 мм слоя",
        "source_url": "https://www.defsmeta.com/rgsn13/gsn_11/giesn-11-01-011-11.php",
        "assumption": False,
    },
    {
        "keywords": ("устройство покрытий: из линолеума", "свариванием полотнищ"),
        "hours_per_qty": 51.82,
        "crew_size": 3,
        "source_label": "ГЭСН 11-01-036-01: укладка линолеума на клее",
        "source_url": "https://www.defsmeta.com/rgsn14/gsn_11/giesn-11-01-036-01.php",
        "assumption": False,
    },
    {
        "keywords": ("разборка элементов облицовки потолков", "плит растровых"),
        "hours_per_qty": 34.51,
        "crew_size": 3,
        "source_label": "ГЭСНр 63-7-2: разборка растрового потолка",
        "source_url": "https://www.defsmeta.com/rgsnr3/gsnr_63/giesnr-63-07-002.php",
        "assumption": False,
    },
    {
        "keywords": ("устройство потолков: плитно-ячеистых", "оцинкованного профиля"),
        "hours_per_qty": 102.46,
        "crew_size": 3,
        "source_label": "ГЭСН 15-01-047-15: потолок Армстронг по каркасу",
        "source_url": "https://www.defsmeta.com/rgsn14/gsn_15/giesn-15-01-047-15.php",
        "assumption": False,
    },
    {
        "keywords": ("облицовка стен по одинарному металлическому каркасу", "гипсоволокнистыми листами", "одним слоем"),
        "hours_per_qty": 84.0,
        "crew_size": 3,
        "source_label": "ФЕР 10-06-037-03: облицовка стен ГВЛ по каркасу",
        "source_url": "https://www.defsmeta.com/rfer10/rfer-10-06-037-03.php",
        "assumption": False,
    },
    {
        "keywords": ("окраска водно-дисперсионными акриловыми составами улучшенная", "по сборным конструкциям стен"),
        "hours_per_qty": 32.73,
        "crew_size": 2,
        "source_label": "ФЕР 15-04-005-03: улучшенная окраска стен водно-дисперсионными составами",
        "source_url": "https://www.defsmeta.com/rfer15_2/rfer-15-04-005-03.php",
        "assumption": False,
    },
    {
        "keywords": ("розетка скрытого монтажа",),
        "hours_per_qty": 0.3048,
        "crew_size": 2,
        "source_label": "ГЭСНм 08-03-591-09: розетка скрытой проводки",
        "source_url": "https://cs.smetnoedelo.ru/gesnm2/gesnm08-03-591-09.html",
        "assumption": False,
    },
    {
        "keywords": ("разработка грунта вручную в траншеях",),
        "hours_per_qty": 123.6,
        "crew_size": 3,
        "source_label": "ГЭСН 01-02-060-06: ручная разработка траншей, группа 2",
        "source_url": "https://base.garant.ru/57505366/53925f69af584b25346d0c0b3ee74ea1/",
        "assumption": False,
    },
    {
        "keywords": ("прокладка трубопроводов канализации", "диаметром: 110 мм"),
        "hours_per_qty": 56.0,
        "crew_size": 3,
        "source_label": "ФЕР 16-04-001-01: прокладка канализации ПЭ 110 мм",
        "source_url": "https://cs.smetnoedelo.ru/fer/fer16-04-001-01.html",
        "assumption": False,
    },
    {
        "keywords": ("засыпка вручную траншей",),
        "hours_per_qty": 88.5,
        "crew_size": 3,
        "source_label": "ГЭСН 01-02-061-01: ручная засыпка траншей",
        "source_url": "https://cs.smetnoedelo.ru/fer/fer01-02-061-02.html",
        "assumption": False,
    },
    {
        "keywords": ("погрузка в автотранспортное средство", "мусор"),
        "hours_per_qty": 0.35,
        "crew_size": 2,
        "source_label": "Укрупнённое допущение по ручной погрузке строительного мусора",
        "source_url": "https://rags.ru/stroyka/text/50012/",
        "assumption": True,
    },
    {
        "keywords": ("перевозка грузов i класса автомобилями-самосвалами",),
        "hours_per_qty": 0.18,
        "crew_size": 1,
        "source_label": "Укрупнённое допущение по рейсу самосвала на 25 км",
        "source_url": "https://fgisrf.ru/frsn/fsscpg",
        "assumption": True,
    },
]


SECTION_SCHEDULE_FALLBACKS = {
    "facade": {"crew_size": 4, "area100": 88.0, "len100": 18.0, "pcs100": 52.0, "area1": 1.0, "len1": 0.18, "ton": 24.0, "generic": 8.0},
    "concrete": {"crew_size": 4, "area100": 46.0, "len100": 16.0, "vol100": 120.0, "area1": 0.46, "len1": 0.16, "generic": 7.0},
    "finishing": {"crew_size": 3, "area100": 62.0, "len100": 14.0, "pcs100": 28.0, "area1": 0.62, "len1": 0.14, "generic": 6.0},
    "electrical": {"crew_size": 2, "pcs1": 0.35, "pcs100": 35.0, "len100": 22.0, "len1": 0.22, "ton": 42.0, "generic": 4.0},
    "plumbing": {"crew_size": 3, "len100": 56.0, "len1": 0.56, "vol100": 104.0, "area100": 72.0, "generic": 6.0},
    "prep": {"crew_size": 3, "area100": 34.0, "len100": 10.0, "pcs100": 44.0, "ton": 8.0, "generic": 5.0},
    "general": {"crew_size": 3, "area100": 60.0, "len100": 16.0, "pcs100": 36.0, "area1": 0.6, "len1": 0.16, "generic": 6.0},
}


def normalize_schedule_text(value: object) -> str:
    return str(value or "").strip().lower()


SCHEDULE_SCOPE_KEYWORDS = {
    "facade": ("окн", "фасад", "откос", "подокон", "жалюз", "витраж", "водоотлив", "уголок пвх"),
    "concrete": ("стяжк", "бетон", "гидроизоляц", "пол", "наливн"),
    "finishing": ("стен", "потол", "окраск", "облицовк", "линолеум", "плинтус", "гипс"),
    "electrical": ("электр", "розет", "светиль", "кабел", "экран", "щит"),
    "plumbing": ("канализац", "труб", "транше", "грунт", "радиатор"),
    "prep": ("демонтаж", "разборка", "снятие", "вывоз мусора", "погрузка", "перевозка"),
}


def classify_schedule_scope(text: str | None) -> str:
    normalized = normalize_schedule_text(text)
    for scope, keywords in SCHEDULE_SCOPE_KEYWORDS.items():
        if any(keyword in normalized for keyword in keywords):
            return scope
    return "general"


def infer_schedule_section_title(raw_section: str | None, title: str) -> str:
    section = str(raw_section or "").strip()
    if section:
        return section
    normalized_title = normalize_schedule_text(title)
    if any(marker in normalized_title for marker in ("окн", "подокон", "жалюз", "откос", "фасад", "водоотлив", "отлив", "желоб", "уголк", "пвх")):
        return "Раздел 1. Окна и фасад"
    scope = classify_schedule_scope(title)
    mapping = {
        "facade": "Раздел 1. Окна и фасад",
        "concrete": "Раздел 1. Общестроительные работы",
        "electrical": "Раздел 1. Электромонтаж",
        "plumbing": "Раздел 1. Инженерные сети",
        "finishing": "Раздел 1. Отделка",
        "prep": "Раздел 1. Подготовка",
        "general": "Раздел 1. Прочие работы",
    }
    return mapping.get(scope, "Раздел 1. Прочие работы")


def schedule_unit_family(unit: str) -> str:
    text = normalize_schedule_text(unit)
    if "100 м3" in text:
        return "vol100"
    if text == "м3":
        return "vol1"
    if "100 м2" in text:
        return "area100"
    if text == "м2":
        return "area1"
    if "100 м" in text:
        return "len100"
    if text == "м":
        return "len1"
    if "100 шт" in text:
        return "pcs100"
    if text.endswith("шт") or text == "шт":
        return "pcs1"
    if "1т" in text or text == "т":
        return "ton"
    return "generic"


def normalized_schedule_qty(qty: float, unit_family: str) -> float:
    if unit_family == "pcs100" and qty >= 20:
        return qty / 100.0
    return qty


def estimate_schedule_work_item(item: sqlite3.Row | dict) -> dict:
    title = str(item["title"] if isinstance(item, sqlite3.Row) else item.get("title", "")).strip()
    unit = str(item["unit"] if isinstance(item, sqlite3.Row) else item.get("unit", "")).strip()
    qty = float(item["planned_qty"] if isinstance(item, sqlite3.Row) else item.get("planned_qty", 0) or 0)
    family = schedule_unit_family(unit)
    qty = normalized_schedule_qty(qty, family)
    text = normalize_schedule_text(title)
    for rule in SECTION_SCHEDULE_RULES:
        if all(keyword in text for keyword in rule["keywords"]):
            return {
                "hours": max(0.0, qty * float(rule["hours_per_qty"])),
                "crew_size": int(rule["crew_size"]),
                "source_label": rule["source_label"],
                "source_url": rule["source_url"],
                "assumption": bool(rule["assumption"]),
                "method": "exact_norm",
            }
    scope = classify_schedule_scope(" ".join([title, unit]))
    fallback = SECTION_SCHEDULE_FALLBACKS.get(scope, SECTION_SCHEDULE_FALLBACKS["general"])
    rate = float(fallback.get(family, fallback.get("generic", 6.0)))
    return {
        "hours": max(0.0, qty * rate),
        "crew_size": int(fallback.get("crew_size", 3)),
        "source_label": "Укрупнённая оценка по типу работ и единице измерения",
        "source_url": "",
        "assumption": True,
        "method": "heuristic",
    }


def build_section_schedule_forecast(project: sqlite3.Row, work_items: list[sqlite3.Row], start_at: date) -> dict:
    sections: list[dict] = []
    by_title: dict[str, dict] = {}
    for row in work_items:
        section_title = infer_schedule_section_title(row["section_title"], str(row["title"] or ""))
        bucket = by_title.get(section_title)
        if not bucket:
            bucket = {
                "title": section_title,
                "scope": classify_schedule_scope(section_title + " " + str(row["title"] or "")),
                "items": [],
                "estimated_hours": 0.0,
                "crew_size": 0,
                "source_map": {},
                "assumptions": False,
                "first_item_id": int(row["id"]),
            }
            by_title[section_title] = bucket
            sections.append(bucket)
        estimate = estimate_schedule_work_item(row)
        bucket["items"].append({
            "id": int(row["id"]),
            "title": str(row["title"] or ""),
            "unit": str(row["unit"] or ""),
            "planned_qty": float(row["planned_qty"] or 0),
            "estimated_hours": round(float(estimate["hours"]), 2),
            "method": estimate["method"],
            "assumption": bool(estimate["assumption"]),
            "source_label": estimate["source_label"],
            "source_url": estimate["source_url"],
        })
        bucket["estimated_hours"] += float(estimate["hours"])
        bucket["crew_size"] = max(int(bucket["crew_size"] or 0), int(estimate["crew_size"]))
        bucket["assumptions"] = bucket["assumptions"] or bool(estimate["assumption"])
        if estimate["source_label"] not in bucket["source_map"]:
            bucket["source_map"][estimate["source_label"]] = {
                "label": estimate["source_label"],
                "url": estimate["source_url"],
            }
    sections.sort(key=lambda item: (item["first_item_id"], item["title"]))

    cursor = start_at
    total_hours = 0.0
    for section in sections:
        work_count = len(section["items"])
        unmatched_count = sum(1 for item in section["items"] if item["assumption"])
        buffer_factor = 1.0 + min(0.18, unmatched_count * 0.03 + max(0, work_count - 3) * 0.01)
        buffered_hours = section["estimated_hours"] * buffer_factor
        crew_size = max(1, int(section["crew_size"] or SECTION_SCHEDULE_FALLBACKS["general"]["crew_size"]))
        duration_days = max(1, int((buffered_hours / max(1, crew_size * 8)) + 0.9999))
        section["estimated_hours"] = round(section["estimated_hours"], 1)
        section["buffered_hours"] = round(buffered_hours, 1)
        section["estimated_days"] = duration_days
        section["crew_size"] = crew_size
        section["start_date"] = cursor.isoformat()
        section["end_date"] = (cursor + timedelta(days=duration_days - 1)).isoformat()
        section["sources"] = list(section["source_map"].values())
        section["work_items"] = work_count
        section.pop("source_map", None)
        section.pop("first_item_id", None)
        total_hours += buffered_hours
        cursor = cursor + timedelta(days=duration_days)

    finish_at = cursor - timedelta(days=1) if sections else start_at
    total_days = max(0, (finish_at - start_at).days + 1) if sections else 0
    return {
        "projectId": int(project["id"]),
        "startDate": start_at.isoformat(),
        "finishDate": finish_at.isoformat() if sections else start_at.isoformat(),
        "totalDays": total_days,
        "totalHours": round(total_hours, 1),
        "sections": [
            {
                "title": section["title"],
                "scope": section["scope"],
                "crewSize": section["crew_size"],
                "estimatedHours": section["estimated_hours"],
                "bufferedHours": section["buffered_hours"],
                "estimatedDays": section["estimated_days"],
                "startDate": section["start_date"],
                "endDate": section["end_date"],
                "workItems": section["work_items"],
                "hasAssumptions": section["assumptions"],
                "sources": section["sources"],
                "items": section["items"],
            }
            for section in sections
        ],
    }


def stage_sort_key(stage: sqlite3.Row | dict) -> tuple[int, int]:
    if isinstance(stage, sqlite3.Row):
        return (int(stage["position"] or 0), int(stage["id"] or 0))
    return (int(stage.get("position", 0) or 0), int(stage.get("id", 0) or 0))


def build_auto_schedule_plan(project: sqlite3.Row, stages: list[sqlite3.Row], materials: list[sqlite3.Row], start_at: date) -> dict:
    stage_by_id = {int(stage["id"]): stage for stage in stages}
    children_map: dict[int | None, list[sqlite3.Row]] = {}
    for stage in stages:
        parent_id = int(stage["parent_id"]) if stage["parent_id"] and int(stage["parent_id"]) in stage_by_id else None
        children_map.setdefault(parent_id, []).append(stage)
    for children in children_map.values():
        children.sort(key=stage_sort_key)

    leaf_ids = {stage_id for stage_id in stage_by_id if stage_id not in children_map}
    candidate_stage_ids = [stage_id for stage_id in sorted(leaf_ids, key=lambda item: stage_sort_key(stage_by_id[item]))]
    materials_by_stage: dict[int, list[dict]] = {}
    auto_linked_materials: list[dict] = []
    for raw in materials:
        item = dict(raw)
        stage_id = int(item["stage_id"]) if item["stage_id"] and int(item["stage_id"]) in stage_by_id else None
        if not stage_id and candidate_stage_ids:
            material_category = classify_scope(" ".join([str(item.get("title", "")), str(item.get("notes", ""))]))
            for candidate_id in candidate_stage_ids:
                stage_category = classify_scope(str(stage_by_id[candidate_id]["title"]))
                if material_category != "general" and material_category == stage_category:
                    stage_id = candidate_id
                    break
            if not stage_id:
                stage_id = candidate_stage_ids[0]
            auto_linked_materials.append({"id": int(item["id"]), "stage_id": stage_id})
        if stage_id:
            materials_by_stage.setdefault(stage_id, []).append(item)

    stage_updates: list[dict] = []
    material_updates: list[dict] = []

    def walk(parent_id: int | None, cursor: date) -> tuple[date | None, date | None, date]:
        first_start = None
        last_end = None
        for stage in children_map.get(parent_id, []):
            stage_id = int(stage["id"])
            linked_materials = materials_by_stage.get(stage_id, [])
            child_nodes = children_map.get(stage_id, [])
            if child_nodes:
                child_start, child_end, next_cursor = walk(stage_id, cursor)
                start = child_start or cursor
                end = child_end or cursor
                cursor = next_cursor
            else:
                duration = estimate_stage_duration(stage, linked_materials)
                start = cursor
                end = start + timedelta(days=duration - 1)
                cursor = end + timedelta(days=1)
            duration_days = max((end - start).days + 1, 1)
            customer_shift = 0 if stage["stage_kind"] != "work" else min(2, max(0, duration_days // 10))
            customer_tail = min(2, max(0, duration_days // 12))
            depends_on_materials = bool(stage["depends_on_materials"]) or bool(linked_materials)
            stage_updates.append(
                {
                    "id": stage_id,
                    "planned_start": start.isoformat(),
                    "planned_end": end.isoformat(),
                    "customer_start": (start + timedelta(days=customer_shift)).isoformat(),
                    "customer_end": (end + timedelta(days=customer_tail)).isoformat(),
                    "depends_on_materials": 1 if depends_on_materials else 0,
                    "duration_days": duration_days,
                    "title": stage["title"],
                }
            )
            if linked_materials:
                for material in linked_materials:
                    lead_days = estimate_material_lead_days(material)
                    material_updates.append(
                        {
                            "id": int(material["id"]),
                            "stage_id": stage_id,
                            "need_by_date": start.isoformat(),
                            "lead_days": lead_days,
                            "title": material["title"],
                        }
                    )
            if first_start is None or start < first_start:
                first_start = start
            if last_end is None or end > last_end:
                last_end = end
        return first_start, last_end, cursor

    plan_start, plan_end, _ = walk(None, start_at)
    sorted_stage_updates = sorted(stage_updates, key=lambda item: (item["planned_start"], item["id"]))
    sorted_material_updates = sorted(material_updates, key=lambda item: (item["need_by_date"], item["title"]))
    deadline_at = parse_iso_date(project["deadline_at"])
    deadline_overrun_days = 0
    if deadline_at and plan_end and plan_end > deadline_at:
        deadline_overrun_days = (plan_end - deadline_at).days
    return {
        "project_start": (plan_start or start_at).isoformat(),
        "project_end": (plan_end or start_at).isoformat(),
        "stage_updates": sorted_stage_updates,
        "material_updates": sorted_material_updates,
        "auto_linked_materials": auto_linked_materials,
        "deadline_overrun_days": deadline_overrun_days,
        "longest_stages": sorted(sorted_stage_updates, key=lambda item: item["duration_days"], reverse=True)[:5],
    }


class PMBIHandler(BaseHTTPRequestHandler):
    server_version = "PMBIBackend/1.0"

    def log_message(self, fmt: str, *args) -> None:
        try:
            if sys.stdout:
                sys.stdout.write("%s - - [%s] %s\n" % (self.client_address[0], self.log_date_time_string(), fmt % args))
        except (AttributeError, OSError):
            pass

    def do_GET(self) -> None:
        path = self.clean_path()
        if path in {"/", "/index.html"}:
            self.redirect(DEFAULT_AUTH_PATH)
            return
        if path.startswith("/api/"):
            self.handle_api("GET", path)
            return
        if path == "/login" or path.startswith("/app"):
            self.serve_app(path)
            return
        if path.startswith("/assets/"):
            self.serve_asset(path)
            return
        self.serve_static(path)

    def do_POST(self) -> None:
        path = self.clean_path()
        if path.startswith("/api/"):
            self.handle_api("POST", path)
            return
        self.send_error(HTTPStatus.METHOD_NOT_ALLOWED)

    def clean_path(self) -> str:
        parsed = urllib.parse.urlsplit(self.path)
        path = urllib.parse.unquote(parsed.path)
        if path.endswith("/") and path != "/":
            path = path[:-1]
        return path or "/"

    def read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", "0") or "0")
        if length > 1024 * 1024:
            raise ValueError("Payload too large")
        raw = self.rfile.read(length) if length else b"{}"
        return json.loads(raw.decode("utf-8") or "{}")

    def read_multipart(self) -> cgi.FieldStorage:
        length = int(self.headers.get("Content-Length", "0") or "0")
        if length <= 0:
            raise ValueError("empty_upload")
        if length > MAX_UPLOAD_BYTES:
            raise ValueError("upload_too_large")
        content_type = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type:
            raise ValueError("multipart_required")
        return cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": content_type,
                "CONTENT_LENGTH": str(length),
            },
            keep_blank_values=True,
        )

    def send_json(self, status: int, payload: dict) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_file(self, file_path: Path, content_type: str, download_name: str, inline: bool = False) -> None:
        body = file_path.read_bytes()
        disposition = "inline" if inline else "attachment"
        quoted_name = urllib.parse.quote(download_name)
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type or "application/octet-stream")
        self.send_header(
            "Content-Disposition",
            f"{disposition}; filename*=UTF-8''{quoted_name}",
        )
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.end_headers()
        self.wfile.write(body)

    def redirect(self, location: str) -> None:
        self.send_response(HTTPStatus.FOUND)
        self.send_header("Location", location)
        self.send_header("Cache-Control", "no-store")
        self.end_headers()

    def session_token(self) -> str | None:
        cookie_header = self.headers.get("Cookie")
        if not cookie_header:
            return None
        cookie = SimpleCookie(cookie_header)
        morsel = cookie.get(SESSION_COOKIE)
        return morsel.value if morsel else None

    def bearer_token(self) -> str | None:
        header = self.headers.get("Authorization", "")
        if not header.lower().startswith("bearer "):
            return None
        token = header[7:].strip()
        return token or None

    def clerk_cookie_token(self) -> str | None:
        cookie_header = self.headers.get("Cookie")
        if not cookie_header:
            return None
        cookie = SimpleCookie(cookie_header)
        morsel = cookie.get("__session")
        return morsel.value if morsel else None

    def auth_config(self) -> dict[str, object]:
        return {
            "clerkEnabled": clerk_enabled(),
            "clerkPublishableKey": CLERK_PUBLISHABLE_KEY,
            "clerkSignInFallbackRedirectUrl": CLERK_SIGN_IN_FALLBACK_REDIRECT_URL,
            "clerkSignUpFallbackRedirectUrl": CLERK_SIGN_UP_FALLBACK_REDIRECT_URL,
            "clerkAfterSignOutUrl": LOGIN_PATH,
        }

    def auth_config_script(self) -> str:
        payload = json.dumps(self.auth_config(), ensure_ascii=False)
        return f"<script>window.__PMBI_AUTH__ = {payload};</script>"

    def clerk_sign_in_script(self) -> str:
        if not clerk_enabled():
            return ""
        return (
            '<script async crossorigin="anonymous" '
            'data-clerk-publishable-key="'
            + CLERK_PUBLISHABLE_KEY
            + '" src="https://cdn.jsdelivr.net/npm/@clerk/clerk-js@latest/dist/clerk.browser.js"></script>'
        )

    def current_user_from_clerk(self) -> tuple[dict | None, str | None]:
        token = self.bearer_token() or self.clerk_cookie_token()
        if not token:
            return None, None
        claims = verify_clerk_session_token(token)
        if not claims:
            return None, "bad_clerk_token"
        clerk_user_id = str(claims.get("sub") or "").strip()
        if not clerk_user_id:
            return None, "bad_clerk_token"
        with db() as con:
            row = con.execute(
                "SELECT * FROM users WHERE clerk_user_id = ? AND is_active = 1",
                (clerk_user_id,),
            ).fetchone()
            if row:
                return user_payload(row), None
        try:
            clerk_user = clerk_api_request(f"/users/{urllib.parse.quote(clerk_user_id, safe='')}")
        except Exception:
            return None, "clerk_lookup_failed"
        email = clerk_primary_email(clerk_user)
        phone = clerk_primary_phone(clerk_user)
        full_name = str(clerk_user.get("first_name") or "").strip()
        last_name = str(clerk_user.get("last_name") or "").strip()
        if last_name:
            full_name = (full_name + " " + last_name).strip()
        with db() as con:
            row = None
            if email:
                row = con.execute(
                    "SELECT * FROM users WHERE lower(email) = ? AND is_active = 1",
                    (email.lower(),),
                ).fetchone()
            if not row and email and email.lower() in CLERK_ADMIN_EMAILS:
                row = con.execute(
                    """
                    SELECT *
                    FROM users
                    WHERE is_active = 1 AND lower(login) = 'admin'
                    ORDER BY id
                    LIMIT 1
                    """
                ).fetchone()
            if not row:
                return None, "clerk_user_not_provisioned"
            con.execute(
                """
                UPDATE users
                SET clerk_user_id = ?, email = COALESCE(?, email), phone = COALESCE(?, phone),
                    name = CASE WHEN trim(COALESCE(name, '')) = '' AND ? != '' THEN ? ELSE name END,
                        updated_at = ?
                WHERE id = ?
                """,
                (
                    clerk_user_id,
                    email,
                    phone,
                    full_name,
                    full_name,
                    now_ts(),
                    row["id"],
                ),
            )
            con.commit()
            refreshed = con.execute("SELECT * FROM users WHERE id = ?", (row["id"],)).fetchone()
        return user_payload(refreshed), None

    def current_user(self) -> dict | None:
        if clerk_enabled():
            user, _error = self.current_user_from_clerk()
            if user:
                return user
        token = self.session_token()
        if not token:
            return None
        with db() as con:
            row = con.execute(
                """
                SELECT u.*
                FROM sessions s
                JOIN users u ON u.id = s.user_id
                WHERE s.token_hash = ? AND s.expires_at > ? AND u.is_active = 1
                """,
                (token_hash(token), now_ts()),
            ).fetchone()
            if not row:
                return None
            return user_payload(row)

    def require_user(self) -> dict | None:
        if clerk_enabled():
            user, error = self.current_user_from_clerk()
            if user:
                return user
            if error:
                status = HTTPStatus.FORBIDDEN if error == "clerk_user_not_provisioned" else HTTPStatus.UNAUTHORIZED
                self.send_json(status, {"error": error})
                return None
        user = self.current_user()
        if not user:
            self.send_json(HTTPStatus.UNAUTHORIZED, {"error": "auth_required"})
            return None
        return user

    def require_role(self, roles: set[str]) -> dict | None:
        user = self.require_user()
        if not user:
            return None
        if not user_has_any_role(user, roles):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return None
        return user

    def set_session_cookie(self, token: str) -> None:
        self.send_header(
            "Set-Cookie",
            f"{SESSION_COOKIE}={token}; Path=/; Max-Age={SESSION_TTL_SECONDS}; HttpOnly; SameSite=Lax",
        )

    def clear_session_cookie(self) -> None:
        self.send_header(
            "Set-Cookie",
            f"{SESSION_COOKIE}=; Path=/; Max-Age=0; HttpOnly; SameSite=Lax",
        )

    def handle_api(self, method: str, path: str) -> None:
        try:
            if method == "POST" and path == "/api/auth/login":
                self.api_login()
            elif method == "POST" and path == "/api/auth/logout":
                self.api_logout()
            elif method == "GET" and path == "/api/auth/me":
                self.api_me()
            elif method == "GET" and path == "/api/roles":
                self.api_roles()
            elif method == "POST" and path == "/api/users/manage":
                self.api_users_manage()
            elif method == "POST" and path == "/api/finance/pay-invoice":
                self.api_pay_invoice()
            elif method == "POST" and path == "/api/admin/users":
                self.api_create_user()
            elif method == "GET" and path == "/api/admin/users":
                self.api_users()
            elif method == "GET" and path == "/api/companies":
                self.api_companies()
            elif method == "POST" and path == "/api/companies":
                self.api_create_company()
            elif method == "GET" and path == "/api/projects":
                self.api_projects()
            elif method == "POST" and path == "/api/projects":
                self.api_create_project()
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/update"):
                self.api_update_project(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/delete"):
                self.api_delete_project(path)
            elif method == "GET" and path == "/api/dashboard":
                self.api_dashboard()
            elif method == "GET" and path == "/api/warehouse-items":
                self.api_warehouse_items()
            elif method == "POST" and path == "/api/warehouse-items/receipt":
                self.api_warehouse_receipt()
            elif method == "POST" and path.startswith("/api/warehouse-items/") and path.endswith("/transfer"):
                self.api_warehouse_transfer(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.count("/") == 3:
                self.api_project_detail(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/assignments"):
                self.api_project_assignments(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/assignments"):
                self.api_create_project_assignment(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/materials-summary"):
                self.api_materials_summary(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/warehouse-matches"):
                self.api_project_warehouse_matches(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/material-schedule"):
                self.api_material_schedule(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/material-schedule"):
                self.api_save_material_schedule(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/materials"):
                self.api_create_material(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/bootstrap"):
                self.api_project_bootstrap(path)
            elif method == "POST" and path.startswith("/api/materials/") and path.endswith("/update"):
                self.api_update_material(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/supplier-offers"):
                self.api_project_supplier_offers(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/market-analysis"):
                self.api_project_market_analysis(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/market-counterparty"):
                self.api_create_market_counterparty(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/supplier-offers"):
                self.api_create_supplier_offer(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/supplier-offers/clear-selection"):
                self.api_clear_supplier_selection(path)
            elif method == "POST" and path.startswith("/api/supplier-offers/") and path.endswith("/update"):
                self.api_update_supplier_offer(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/finances"):
                self.api_project_finances(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/finances"):
                self.api_create_finance_entry(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/finances/invoice-upload"):
                self.api_upload_finance_invoice(path)
            elif method == "POST" and path.startswith("/api/finances/") and path.endswith("/update"):
                self.api_update_finance_entry(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/estimate-import"):
                self.api_import_estimate(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/auto-schedule"):
                self.api_project_auto_schedule(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/section-schedule-forecast"):
                self.api_project_section_schedule_forecast(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/schedule-status"):
                self.api_project_schedule_status(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/schedule-status"):
                self.api_update_project_schedule_status(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/analysis"):
                self.api_project_analysis(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/stages"):
                self.api_project_stages(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/stages"):
                self.api_create_project_stage(path)
            elif method == "POST" and path.startswith("/api/stages/") and path.endswith("/update"):
                self.api_update_stage(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/tasks"):
                self.api_project_tasks(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/tasks"):
                self.api_create_task(path)
            elif method == "POST" and path.startswith("/api/tasks/") and path.endswith("/update"):
                self.api_update_task(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/notifications"):
                self.api_project_notifications(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/documents"):
                self.api_project_documents(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/documents"):
                self.api_upload_project_document(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/executive-docs"):
                self.api_project_executive_docs(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/executive-docs"):
                self.api_create_project_executive_doc(path)
            elif method == "GET" and path.startswith("/api/documents/") and path.endswith("/download"):
                self.api_document_file(path, inline=False)
            elif method == "GET" and path.startswith("/api/documents/") and path.endswith("/view"):
                self.api_document_file(path, inline=True)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/daily-logs"):
                self.api_project_daily_logs(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/daily-logs"):
                self.api_create_daily_log(path)
            elif method == "POST" and path.startswith("/api/projects/") and "/daily-logs/" in path and path.endswith("/delete"):
                self.api_delete_daily_log(path)
            elif method == "GET" and path.startswith("/api/projects/") and path.endswith("/chats"):
                self.api_project_chats(path)
            elif method == "GET" and path.startswith("/api/chats/") and path.endswith("/messages"):
                self.api_chat_messages(path)
            elif method == "POST" and path.startswith("/api/chats/") and path.endswith("/messages"):
                self.api_create_chat_message(path)
            elif method == "POST" and path.startswith("/api/projects/") and path.endswith("/stock-moves"):
                self.api_create_stock_move(path)
            else:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "not_found"})
        except json.JSONDecodeError:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "invalid_json"})
        except ValueError as error:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": str(error)})

    def api_login(self) -> None:
        if clerk_enabled():
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "clerk_enabled"})
            return
        payload = self.read_json()
        login = str(payload.get("login", "")).strip()
        password = str(payload.get("password", ""))
        with db() as con:
            row = con.execute(
                "SELECT * FROM users WHERE login = ? AND is_active = 1",
                (login,),
            ).fetchone()
            if not row or not verify_password(password, row["password_hash"]):
                self.send_json(HTTPStatus.UNAUTHORIZED, {"error": "bad_credentials"})
                return

            token = secrets.token_urlsafe(32)
            con.execute(
                """
                INSERT INTO sessions (user_id, token_hash, created_at, expires_at, user_agent, ip)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    row["id"],
                    token_hash(token),
                    now_ts(),
                    now_ts() + SESSION_TTL_SECONDS,
                    self.headers.get("User-Agent", ""),
                    self.client_address[0],
                ),
            )
            con.execute(
                "INSERT INTO audit_log (user_id, action, entity, created_at) VALUES (?, 'login', 'user', ?)",
                (row["id"], now_ts()),
            )
            con.commit()

        body = json.dumps({"user": user_payload(row)}, ensure_ascii=False).encode("utf-8")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.set_session_cookie(token)
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def api_logout(self) -> None:
        token = self.session_token()
        if token:
            with db() as con:
                con.execute("DELETE FROM sessions WHERE token_hash = ?", (token_hash(token),))
                con.commit()
        body = json.dumps({"ok": True}, ensure_ascii=False).encode("utf-8")
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.clear_session_cookie()
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def api_me(self) -> None:
        if clerk_enabled():
            user, error = self.current_user_from_clerk()
            if user:
                self.send_json(HTTPStatus.OK, {"user": user})
                return
            if error:
                status = HTTPStatus.FORBIDDEN if error == "clerk_user_not_provisioned" else HTTPStatus.UNAUTHORIZED
                self.send_json(status, {"error": error})
                return
        user = self.current_user()
        if not user:
            self.send_json(HTTPStatus.UNAUTHORIZED, {"error": "auth_required"})
            return
        self.send_json(HTTPStatus.OK, {"user": user})

    def api_roles(self) -> None:
        user = self.require_user()
        if not user:
            return
        with db() as con:
            rows = con.execute("SELECT code, name, description FROM roles ORDER BY id").fetchall()
        self.send_json(HTTPStatus.OK, {"roles": [dict(row) for row in rows]})

    def api_create_user(self) -> None:
        admin = self.require_role({"admin", "director"})
        if not admin:
            return
        payload = self.read_json()
        login = str(payload.get("login", "")).strip()
        password = str(payload.get("password", ""))
        requested_roles = payload.get("roles")
        if isinstance(requested_roles, list):
            role_codes = [normalize_role(str(item).strip()) for item in requested_roles if str(item).strip()]
        else:
            role_codes = [normalize_role(str(payload.get("role", "")).strip())]
        role_codes = list(dict.fromkeys(role_codes))
        role = role_codes[0] if role_codes else ""
        name = str(payload.get("name", "")).strip() or login
        email = str(payload.get("email", "")).strip().lower() or None
        phone = str(payload.get("phone", "")).strip() or None
        if (
            not login
            or len(password) < 10
            or not role_codes
            or any(role_code not in ROLE_LABELS for role_code in role_codes)
            or (clerk_enabled() and not email)
        ):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "invalid_user_data"})
            return
        clerk_user_id = None
        with db() as con:
            existing_login = con.execute("SELECT id FROM users WHERE login = ?", (login,)).fetchone()
            if existing_login:
                self.send_json(HTTPStatus.CONFLICT, {"error": "login_exists"})
                return
            if clerk_enabled() and email:
                existing_email = con.execute("SELECT id FROM users WHERE lower(email) = ?", (email.lower(),)).fetchone()
                if existing_email:
                    self.send_json(HTTPStatus.CONFLICT, {"error": "email_exists"})
                    return
        if clerk_enabled():
            first_name, last_name = split_person_name(name)
            clerk_payload = {
                "email_address": [email],
                "password": password,
                "username": login,
                "first_name": first_name,
                "last_name": last_name,
                "skip_password_checks": False,
                "skip_password_requirement": False,
            }
            try:
                clerk_user = clerk_api_request("/users", method="POST", payload=clerk_payload)
                clerk_user_id = str(clerk_user.get("id") or "").strip() or None
            except Exception as error:
                self.send_json(HTTPStatus.BAD_GATEWAY, {"error": "clerk_create_user_failed", "detail": str(error)})
                return
        with db() as con:
            try:
                cur = con.execute(
                    """
                    INSERT INTO users (login, email, phone, clerk_user_id, password_hash, role, name, status, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 'active', ?, ?)
                    """,
                    (login, email, phone, clerk_user_id, hash_password(password), role, name, now_ts(), now_ts()),
                )
            except sqlite3.IntegrityError:
                self.send_json(HTTPStatus.CONFLICT, {"error": "login_exists"})
                return
            for role_code in role_codes:
                role_row = con.execute("SELECT id FROM roles WHERE code = ?", (role_code,)).fetchone()
                if role_row:
                    con.execute(
                        """
                        INSERT OR IGNORE INTO user_roles (user_id, role_id, created_at)
                        VALUES (?, ?, ?)
                        """,
                        (cur.lastrowid, role_row["id"], now_ts()),
                    )
            con.execute(
                """
                INSERT INTO audit_log (user_id, action, entity, entity_id, payload, created_at)
                VALUES (?, 'create_user', 'user', ?, ?, ?)
                """,
                (admin["id"], cur.lastrowid, json.dumps({"login": login, "roles": role_codes}, ensure_ascii=False), now_ts()),
            )
            con.commit()
        self.send_json(
            HTTPStatus.CREATED,
            {
                "id": cur.lastrowid,
                "login": login,
                "email": email,
                "phone": phone,
                "clerkUserId": clerk_user_id,
                "role": role,
                "roles": role_codes,
                "name": name,
            },
        )

    def api_users(self) -> None:
        admin = self.require_role({"admin", "director"})
        if not admin:
            return
        with db() as con:
            rows = con.execute(
                """
                SELECT id, login, email, phone, clerk_user_id, role, name, status, is_active, created_at
                FROM users
                ORDER BY id
                """
            ).fetchall()
            role_rows = con.execute(
                """
                SELECT ur.user_id, r.code, r.name
                FROM user_roles ur
                JOIN roles r ON r.id = ur.role_id
                ORDER BY r.id
                """
            ).fetchall()
            roles_by_user: dict[int, list[dict]] = {}
            for role_row in role_rows:
                roles_by_user.setdefault(int(role_row["user_id"]), []).append({
                    "code": normalize_role(role_row["code"]),
                    "name": ROLE_LABELS.get(normalize_role(role_row["code"]), role_row["name"]),
                })
        self.send_json(
            HTTPStatus.OK,
            {
                "users": [
                    {
                        "id": row["id"],
                        "login": row["login"],
                        "email": row["email"],
                        "phone": row["phone"],
                        "clerkUserId": row["clerk_user_id"],
                        "role": normalize_role(row["role"]),
                        "roles": roles_by_user.get(int(row["id"]), [{"code": normalize_role(row["role"]), "name": ROLE_LABELS.get(normalize_role(row["role"]), row["role"])}]),
                        "roleLabel": ROLE_LABELS.get(normalize_role(row["role"]), row["role"]),
                        "name": row["name"],
                        "status": row["status"],
                        "isActive": bool(row["is_active"]),
                        "createdAt": row["created_at"],
                    }
                    for row in rows
                ]
            },
        )

    def set_project_foremen(
        self,
        con: sqlite3.Connection,
        project_id: int,
        foreman_ids: list[int],
        assigned_by: int,
    ) -> list[int]:
        project = con.execute("SELECT id FROM projects WHERE id = ?", (project_id,)).fetchone()
        if not project:
            raise ValueError("project_not_found")
        cleaned_ids = list(dict.fromkeys(int(item) for item in foreman_ids if int(item) > 0))
        if cleaned_ids:
            placeholders = ",".join("?" for _ in cleaned_ids)
            rows = con.execute(
                f"""
                SELECT DISTINCT u.id
                FROM users u
                LEFT JOIN user_roles ur ON ur.user_id = u.id
                LEFT JOIN roles r ON r.id = ur.role_id
                WHERE u.id IN ({placeholders})
                  AND u.is_active = 1
                  AND (u.role = 'foreman' OR r.code = 'foreman')
                ORDER BY u.id
                """,
                cleaned_ids,
            ).fetchall()
            valid_ids = [int(row["id"]) for row in rows]
            if len(valid_ids) != len(cleaned_ids):
                raise ValueError("foreman_not_found")
        else:
            valid_ids = []

        previous_rows = con.execute(
            "SELECT user_id FROM object_assignments WHERE object_id = ? AND role_code = 'foreman'",
            (project_id,),
        ).fetchall()
        previous_ids = [int(row["user_id"]) for row in previous_rows]
        con.execute(
            "DELETE FROM object_assignments WHERE object_id = ? AND role_code = 'foreman'",
            (project_id,),
        )
        for foreman_id in valid_ids:
            con.execute(
                """
                INSERT OR IGNORE INTO object_assignments (object_id, user_id, role_code, responsibility, is_primary, assigned_by, assigned_at)
                VALUES (?, ?, 'foreman', 'РџСЂРѕСЂР°Р± РѕР±СЉРµРєС‚Р°', ?, ?, ?)
                """,
                (project_id, foreman_id, 1 if foreman_id == valid_ids[0] else 0, assigned_by, now_ts()),
            )
            con.execute(
                "INSERT OR IGNORE INTO user_project_access (user_id, project_id) VALUES (?, ?)",
                (foreman_id, project_id),
            )
        for foreman_id in previous_ids:
            if foreman_id in valid_ids:
                continue
            has_other_assignment = con.execute(
                "SELECT 1 FROM object_assignments WHERE object_id = ? AND user_id = ? LIMIT 1",
                (project_id, foreman_id),
            ).fetchone()
            if not has_other_assignment:
                con.execute(
                    "DELETE FROM user_project_access WHERE user_id = ? AND project_id = ?",
                    (foreman_id, project_id),
                )
        con.execute(
            "UPDATE projects SET foreman_id = ?, updated_at = ? WHERE id = ?",
            (valid_ids[0] if valid_ids else None, now_ts(), project_id),
        )
        return valid_ids

    def api_users_manage(self) -> None:
        director = self.require_role({"director"})
        if not director:
            return
        payload = self.read_json()
        action = str(payload.get("action", "create_foreman")).strip() or "create_foreman"

        if action in {"set_project_foremen", "set_access"}:
            try:
                project_id = int(payload.get("project_id", payload.get("projectId")))
                foreman_ids = payload.get("foreman_ids", payload.get("foremanIds", []))
                if not isinstance(foreman_ids, list):
                    raise ValueError("bad_foreman_ids")
                foreman_ids = [int(item) for item in foreman_ids]
            except (TypeError, ValueError):
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_access_payload"})
                return
            with db() as con:
                try:
                    assigned = self.set_project_foremen(con, project_id, foreman_ids, int(director["id"]))
                except ValueError as error:
                    status = HTTPStatus.NOT_FOUND if str(error) == "project_not_found" else HTTPStatus.BAD_REQUEST
                    self.send_json(status, {"error": str(error)})
                    return
                create_audit(
                    con,
                    director["id"],
                    "set_project_foremen",
                    "project",
                    project_id,
                    {"foreman_ids": assigned},
                )
                con.commit()
            self.send_json(HTTPStatus.OK, {"ok": True, "projectId": project_id, "assigned_foremen": assigned})
            return

        login = str(payload.get("login", "")).strip()
        password = str(payload.get("password", ""))
        name = str(payload.get("name", "")).strip() or login
        email = str(payload.get("email", "")).strip().lower() or None
        phone = str(payload.get("phone", "")).strip() or None
        project_ids = payload.get("project_ids", payload.get("projectIds", []))
        if project_ids is None:
            project_ids = []
        if not isinstance(project_ids, list):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_ids"})
            return
        try:
            project_ids = list(dict.fromkeys(int(item) for item in project_ids if int(item) > 0))
        except (TypeError, ValueError):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_ids"})
            return
        if not login or len(password) < 10 or (clerk_enabled() and not email):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "invalid_user_data"})
            return

        clerk_user_id = None
        with db() as con:
            existing_login = con.execute("SELECT id FROM users WHERE login = ?", (login,)).fetchone()
            if existing_login:
                self.send_json(HTTPStatus.CONFLICT, {"error": "login_exists"})
                return
            if clerk_enabled() and email:
                existing_email = con.execute("SELECT id FROM users WHERE lower(email) = ?", (email.lower(),)).fetchone()
                if existing_email:
                    self.send_json(HTTPStatus.CONFLICT, {"error": "email_exists"})
                    return
            if project_ids:
                placeholders = ",".join("?" for _ in project_ids)
                existing_project_count = con.execute(
                    f"SELECT COUNT(*) FROM projects WHERE id IN ({placeholders})",
                    project_ids,
                ).fetchone()[0]
                if int(existing_project_count or 0) != len(project_ids):
                    self.send_json(HTTPStatus.BAD_REQUEST, {"error": "project_not_found"})
                    return

        if clerk_enabled():
            first_name, last_name = split_person_name(name)
            try:
                clerk_user = clerk_api_request(
                    "/users",
                    method="POST",
                    payload={
                        "email_address": [email],
                        "password": password,
                        "username": login,
                        "first_name": first_name,
                        "last_name": last_name,
                        "skip_password_checks": False,
                        "skip_password_requirement": False,
                    },
                )
                clerk_user_id = str(clerk_user.get("id") or "").strip() or None
            except Exception as error:
                self.send_json(HTTPStatus.BAD_GATEWAY, {"error": "clerk_create_user_failed", "detail": str(error)})
                return

        with db() as con:
            try:
                cur = con.execute(
                    """
                    INSERT INTO users (login, email, phone, clerk_user_id, password_hash, role, name, status, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, 'foreman', ?, 'active', ?, ?)
                    """,
                    (login, email, phone, clerk_user_id, hash_password(password), name, now_ts(), now_ts()),
                )
            except sqlite3.IntegrityError:
                self.send_json(HTTPStatus.CONFLICT, {"error": "login_exists"})
                return
            role_row = con.execute("SELECT id FROM roles WHERE code = 'foreman'").fetchone()
            if role_row:
                con.execute(
                    """
                    INSERT OR IGNORE INTO user_roles (user_id, role_id, created_at)
                    VALUES (?, ?, ?)
                    """,
                    (cur.lastrowid, role_row["id"], now_ts()),
                )
            for project_id in project_ids:
                con.execute(
                    """
                    INSERT OR IGNORE INTO object_assignments (object_id, user_id, role_code, responsibility, is_primary, assigned_by, assigned_at)
                    VALUES (?, ?, 'foreman', 'РџСЂРѕСЂР°Р± РѕР±СЉРµРєС‚Р°', 0, ?, ?)
                    """,
                    (project_id, cur.lastrowid, director["id"], now_ts()),
                )
                con.execute(
                    "INSERT OR IGNORE INTO user_project_access (user_id, project_id) VALUES (?, ?)",
                    (cur.lastrowid, project_id),
                )
                con.execute(
                    "UPDATE projects SET foreman_id = COALESCE(foreman_id, ?), updated_at = ? WHERE id = ?",
                    (cur.lastrowid, now_ts(), project_id),
                )
            create_audit(
                con,
                director["id"],
                "create_foreman",
                "user",
                cur.lastrowid,
                {"login": login, "project_ids": project_ids},
            )
            con.commit()
        self.send_json(
            HTTPStatus.CREATED,
            {
                "id": cur.lastrowid,
                "login": login,
                "email": email,
                "phone": phone,
                "clerkUserId": clerk_user_id,
                "role": "foreman",
                "roles": ["foreman"],
                "name": name,
                "projectIds": project_ids,
            },
        )

    def api_companies(self) -> None:
        user = self.require_role({"admin", "director", "purchaser", "foreman"})
        if not user:
            return
        parsed = urllib.parse.urlsplit(self.path)
        query = urllib.parse.parse_qs(parsed.query)
        company_type = str(query.get("type", [""])[0]).strip()
        params: list = []
        where = ""
        if company_type:
            where = "WHERE type = ?"
            params.append(company_type)
        with db() as con:
            rows = con.execute(
                f"""
                SELECT id, type, name, inn, kpp, ogrn, phone, email, address, notes, created_at
                FROM companies
                {where}
                ORDER BY type, name
                """,
                params,
            ).fetchall()
        self.send_json(HTTPStatus.OK, {"companies": [dict(row) for row in rows]})

    def api_create_company(self) -> None:
        user = self.require_role({"admin", "director"})
        if not user:
            return
        payload = self.read_json()
        company_type = str(payload.get("type", "")).strip()
        name = str(payload.get("name", "")).strip()
        if company_type not in {"own_legal_entity", "client", "supplier", "contractor", "other"} or not name:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "invalid_company_data"})
            return
        values = {
            "inn": str(payload.get("inn", "")).strip() or None,
            "kpp": str(payload.get("kpp", "")).strip() or None,
            "ogrn": str(payload.get("ogrn", "")).strip() or None,
            "phone": str(payload.get("phone", "")).strip() or None,
            "email": str(payload.get("email", "")).strip() or None,
            "address": str(payload.get("address", "")).strip() or None,
            "notes": str(payload.get("notes", "")).strip() or None,
        }
        with db() as con:
            cur = con.execute(
                """
                INSERT INTO companies (type, name, inn, kpp, ogrn, phone, email, address, notes, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    company_type,
                    name,
                    values["inn"],
                    values["kpp"],
                    values["ogrn"],
                    values["phone"],
                    values["email"],
                    values["address"],
                    values["notes"],
                    now_ts(),
                ),
            )
            if company_type == "client":
                con.execute(
                    """
                    INSERT INTO clients (company_id, name, contact_person, phone, email, notes)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (cur.lastrowid, name, None, values["phone"], values["email"], values["notes"]),
                )
            create_audit(con, user["id"], "create_company", "company", cur.lastrowid, {"name": name, "type": company_type})
            con.commit()
        self.send_json(HTTPStatus.CREATED, {"id": cur.lastrowid, "type": company_type, "name": name})

    def api_projects(self) -> None:
        user = self.require_user()
        if not user:
            return
        with db() as con:
            if user_has_any_role(user, {"admin", "director"}):
                rows = con.execute("SELECT * FROM projects ORDER BY id DESC").fetchall()
            elif user_has_any_role(user, {"foreman"}):
                rows = con.execute(
                    """
                    SELECT DISTINCT p.*
                    FROM projects p
                    JOIN object_assignments oa ON oa.object_id = p.id
                    WHERE oa.user_id = ? AND oa.role_code = 'foreman'
                    ORDER BY p.id DESC
                    """,
                    (user["id"],),
                ).fetchall()
            else:
                rows = con.execute(
                    """
                    SELECT p.*
                    FROM projects p
                    JOIN user_project_access a ON a.project_id = p.id
                    WHERE a.user_id = ?
                    ORDER BY p.id DESC
                    """,
                    (user["id"],),
                ).fetchall()
            projects = [serialize_project(row, user) for row in rows]
        self.send_json(HTTPStatus.OK, {"projects": projects})

    def api_create_project(self) -> None:
        user = self.require_user()
        if not user:
            return
        if not user_has_any_role(user, {"admin", "director"}):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        payload = self.read_json()
        title = str(payload.get("title", "")).strip()
        address = str(payload.get("address", "")).strip()
        client_name = str(payload.get("client_name", payload.get("clientName", ""))).strip()
        customer_company_id = payload.get("customer_company_id", payload.get("customerCompanyId"))
        own_legal_entity_id = payload.get("own_legal_entity_id", payload.get("ownLegalEntityId"))
        try:
            customer_company_id = int(customer_company_id) if customer_company_id else None
            own_legal_entity_id = int(own_legal_entity_id) if own_legal_entity_id else None
        except (TypeError, ValueError):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_company_id"})
            return
        if not title or not address:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "title_address_client_required"})
            return
        budget = float(payload.get("budget", 0) or 0)
        started_at = str(payload.get("started_at", payload.get("startedAt", ""))).strip() or None
        deadline_at = str(payload.get("deadline_at", payload.get("deadlineAt", ""))).strip() or None
        city = str(payload.get("city", "")).strip() or None
        region = str(payload.get("region", "")).strip() or None
        contract_date = str(payload.get("contract_date", payload.get("contractDate", ""))).strip() or None
        description = str(payload.get("description", "")).strip() or None
        with db() as con:
            customer_company = None
            own_company = None
            if customer_company_id:
                customer_company = con.execute(
                    "SELECT id, name FROM companies WHERE id = ? AND type IN ('client','other')",
                    (customer_company_id,),
                ).fetchone()
                if not customer_company:
                    self.send_json(HTTPStatus.BAD_REQUEST, {"error": "customer_company_not_found"})
                    return
                if not client_name:
                    client_name = customer_company["name"]
            if own_legal_entity_id:
                own_company = con.execute(
                    "SELECT id FROM companies WHERE id = ? AND type = 'own_legal_entity'",
                    (own_legal_entity_id,),
                ).fetchone()
                if not own_company:
                    self.send_json(HTTPStatus.BAD_REQUEST, {"error": "own_legal_entity_not_found"})
                    return
            if not client_name:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "title_address_client_required"})
                return
            client_id = None
            if customer_company_id:
                client_row = con.execute("SELECT id FROM clients WHERE company_id = ?", (customer_company_id,)).fetchone()
                if client_row:
                    client_id = client_row["id"]
                else:
                    client_cur = con.execute(
                        "INSERT INTO clients (company_id, name) VALUES (?, ?)",
                        (customer_company_id, client_name),
                    )
                    client_id = client_cur.lastrowid
            cur = con.execute(
                """
                INSERT INTO projects (
                    title, client_id, customer_company_id, own_legal_entity_id,
                    address, city, region, client_name, contract_no, contract_date,
                    director_id, status, progress, budget, paid, spent,
                    started_at, deadline_at, description, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Подготовка', 0, ?, 0, 0, ?, ?, ?, ?, ?)
                """,
                (
                    title,
                    client_id,
                    customer_company_id,
                    own_legal_entity_id,
                    address,
                    city,
                    region,
                    client_name,
                    str(payload.get("contract_no", payload.get("contractNo", ""))).strip() or None,
                    contract_date,
                    user["id"] if user_has_any_role(user, {"director"}) else None,
                    budget,
                    started_at,
                    deadline_at,
                    description,
                    now_ts(),
                    now_ts(),
                ),
            )
            project_id = cur.lastrowid
            con.execute(
                """
                INSERT OR IGNORE INTO object_assignments (object_id, user_id, role_code, responsibility, is_primary, assigned_by, assigned_at)
                VALUES (?, ?, ?, ?, 1, ?, ?)
                """,
                (project_id, user["id"], user["role"], "Создатель объекта и руководитель запуска", user["id"], now_ts()),
            )
            con.execute(
                "INSERT OR IGNORE INTO user_project_access (user_id, project_id) VALUES (?, ?)",
                (user["id"], project_id),
            )
            default_stages = [
                ("Подготовка", 1, started_at, None, 0, "Директор", 0),
                ("Закупка материалов", 2, started_at, None, 0, "Закупщик", 1),
                ("Основные работы", 3, None, deadline_at, 0, "Прораб", 1),
                ("Исполнительная документация", 4, None, deadline_at, 0, "Прораб", 0),
                ("Сдача объекта", 5, None, deadline_at, 0, "Директор", 0),
            ]
            con.executemany(
                """
                INSERT INTO work_stages (project_id, title, position, planned_start, planned_end, progress, responsible, depends_on_materials, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                [(project_id, *stage, now_ts()) for stage in default_stages],
            )
            con.execute(
                "INSERT INTO chats (project_id, chat_type, title, created_at) VALUES (?, 'team', 'Внутренний чат команды', ?)",
                (project_id, now_ts()),
            )
            con.execute(
                "INSERT INTO chats (project_id, chat_type, title, created_at) VALUES (?, 'client', 'Чат с заказчиком', ?)",
                (project_id, now_ts()),
            )
            con.execute(
                """
                INSERT INTO documents (project_id, title, doc_type, status, is_client_visible, created_at)
                VALUES (?, 'Договор подряда', 'contract', 'draft', 0, ?)
                """,
                (project_id, now_ts()),
            )
            create_audit(con, user["id"], "create_project", "project", project_id, {"title": title})
            con.commit()
            row = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
        self.send_json(HTTPStatus.CREATED, {"project": serialize_project(row, user)})

    def api_update_project(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_role({"admin", "director"})
        if not user:
            return
        payload = self.read_json()
        with db() as con:
            current = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            if not current:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "project_not_found"})
                return

            title = str(payload.get("title", current["title"] or "")).strip()
            address = str(payload.get("address", current["address"] or "")).strip()
            client_name = str(payload.get("client_name", payload.get("clientName", current["client_name"] or ""))).strip()
            status = str(payload.get("status", current["status"] or "")).strip() or "Подготовка"
            contract_no = str(payload.get("contract_no", payload.get("contractNo", current["contract_no"] or ""))).strip() or None
            contract_date = str(payload.get("contract_date", payload.get("contractDate", current["contract_date"] or ""))).strip() or None
            started_at = str(payload.get("started_at", payload.get("startedAt", current["started_at"] or ""))).strip() or None
            deadline_at = str(payload.get("deadline_at", payload.get("deadlineAt", current["deadline_at"] or ""))).strip() or None
            city = str(payload.get("city", current["city"] or "")).strip() or None
            region = str(payload.get("region", current["region"] or "")).strip() or None
            description = str(payload.get("description", current["description"] or "")).strip() or None

            if not title or not address or not client_name:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "title_address_client_required"})
                return

            try:
                budget_value = payload.get("budget", current["budget"])
                budget = float(budget_value or 0)
            except (TypeError, ValueError):
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_budget"})
                return

            con.execute(
                """
                UPDATE projects
                SET
                    title = ?,
                    address = ?,
                    city = ?,
                    region = ?,
                    client_name = ?,
                    contract_no = ?,
                    contract_date = ?,
                    status = ?,
                    budget = ?,
                    started_at = ?,
                    deadline_at = ?,
                    description = ?,
                    updated_at = ?
                WHERE id = ?
                """,
                (
                    title,
                    address,
                    city,
                    region,
                    client_name,
                    contract_no,
                    contract_date,
                    status,
                    budget,
                    started_at,
                    deadline_at,
                    description,
                    now_ts(),
                    project_id,
                ),
            )
            create_audit(con, user["id"], "update_project", "project", project_id, {"title": title, "status": status})
            con.commit()
            row = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
        self.send_json(HTTPStatus.OK, {"project": serialize_project(row, user)})

    def api_delete_project(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_role({"admin", "director"})
        if not user:
            return
        with db() as con:
            current = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            if not current:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "project_not_found"})
                return
            title = current["title"] or ""
            con.execute("DELETE FROM projects WHERE id = ?", (project_id,))
            create_audit(con, user["id"], "delete_project", "project", project_id, {"title": title})
            con.commit()
        self.send_json(HTTPStatus.OK, {"ok": True, "deleted_id": project_id})

    def api_dashboard(self) -> None:
        user = self.require_user()
        if not user:
            return
        with db() as con:
            if user_has_any_role(user, {"admin", "director"}):
                projects = con.execute("SELECT * FROM projects").fetchall()
            elif user_has_any_role(user, {"foreman"}):
                projects = con.execute(
                    """
                    SELECT DISTINCT p.*
                    FROM projects p
                    JOIN object_assignments oa ON oa.object_id = p.id
                    WHERE oa.user_id = ? AND oa.role_code = 'foreman'
                    ORDER BY p.id DESC
                    """,
                    (user["id"],),
                ).fetchall()
            else:
                projects = con.execute(
                    """
                    SELECT p.*
                    FROM projects p
                    LEFT JOIN user_project_access a ON a.project_id = p.id
                    LEFT JOIN object_assignments oa ON oa.object_id = p.id
                    WHERE a.user_id = ? OR oa.user_id = ?
                    GROUP BY p.id
                    """,
                    (user["id"], user["id"]),
                ).fetchall()
            project_ids = [row["id"] for row in projects]
            total_budget = sum(float(row["budget"] or 0) for row in projects)
            total_paid = sum(float(row["paid"] or 0) for row in projects)
            total_spent = sum(float(row["spent"] or 0) for row in projects)
            active = sum(1 for row in projects if "работ" in str(row["status"]).lower())
            shortages = 0
            critical_items = []
            for project in projects:
                for item in material_summary_rows(con, int(project["id"])):
                    if item["missingQty"] > 0:
                        work_date = item.get("stageStartDate") or item.get("needByDate") or item.get("stageEndDate") or ""
                        parsed_work_date = parse_iso_date(str(work_date or ""))
                        shortages += 1
                        critical_items.append({
                            "projectId": project["id"],
                            "projectTitle": project["title"],
                            "title": item["title"],
                            "missingQty": item["missingQty"],
                            "unit": item["unit"],
                            "sectionTitle": item.get("sectionTitle") or "",
                            "stageTitle": item.get("stageTitle") or "",
                            "needByDate": item.get("needByDate") or "",
                            "workDate": work_date,
                            "daysUntilWork": (parsed_work_date - parse_iso_date(TODAY_ISO)).days if parsed_work_date and parse_iso_date(TODAY_ISO) else None,
                        })
            open_tasks = 0
            if project_ids:
                placeholders = ",".join("?" for _ in project_ids)
                open_tasks = con.execute(
                    f"SELECT COUNT(*) FROM tasks WHERE project_id IN ({placeholders}) AND status != 'done'",
                    project_ids,
                ).fetchone()[0]
                task_rows = con.execute(
                    f"""
                    SELECT t.title, t.status, t.priority, t.due_at, t.created_at, p.title AS project_title
                    FROM tasks t
                    JOIN projects p ON p.id = t.project_id
                    WHERE t.project_id IN ({placeholders})
                    ORDER BY CASE WHEN t.status = 'done' THEN 1 ELSE 0 END, t.due_at IS NULL, t.due_at ASC, t.id DESC
                    LIMIT 6
                    """,
                    project_ids,
                ).fetchall()
                document_rows = con.execute(
                    f"""
                    SELECT d.title, d.doc_type, d.status, d.created_at, p.title AS project_title
                    FROM documents d
                    JOIN projects p ON p.id = d.project_id
                    WHERE d.project_id IN ({placeholders})
                    {"" if user["role"] != "customer" else "AND d.is_client_visible = 1"}
                    ORDER BY d.id DESC
                    LIMIT 5
                    """,
                    project_ids,
                ).fetchall()
                daily_log_rows = con.execute(
                    f"""
                    SELECT l.title, l.report_date, l.work_done, l.created_at, p.title AS project_title
                    FROM daily_logs l
                    JOIN projects p ON p.id = l.project_id
                    WHERE l.project_id IN ({placeholders})
                    {"" if user["role"] != "customer" else "AND l.is_client_visible = 1"}
                    ORDER BY l.report_date DESC, l.id DESC
                    LIMIT 6
                    """,
                    project_ids,
                ).fetchall()
                if user["role"] == "customer":
                    message_rows = con.execute(
                        f"""
                        SELECT m.body, m.created_at, c.title AS chat_title, p.title AS project_title
                        FROM chat_messages m
                        JOIN chats c ON c.id = m.chat_id
                        JOIN projects p ON p.id = c.project_id
                        WHERE c.project_id IN ({placeholders}) AND c.chat_type = 'client'
                        ORDER BY m.id DESC
                        LIMIT 5
                        """,
                        project_ids,
                    ).fetchall()
                    stock_rows = []
                else:
                    message_rows = con.execute(
                        f"""
                        SELECT m.body, m.created_at, c.title AS chat_title, p.title AS project_title
                        FROM chat_messages m
                        JOIN chats c ON c.id = m.chat_id
                        JOIN projects p ON p.id = c.project_id
                        WHERE c.project_id IN ({placeholders})
                        ORDER BY m.id DESC
                        LIMIT 5
                        """,
                        project_ids,
                    ).fetchall()
                    stock_rows = con.execute(
                        f"""
                        SELECT sm.move_type, sm.qty, sm.comment, sm.created_at, e.title AS material_title, e.unit, p.title AS project_title
                        FROM stock_moves sm
                        JOIN estimate_items e ON e.id = sm.estimate_item_id
                        JOIN projects p ON p.id = sm.project_id
                        WHERE sm.project_id IN ({placeholders})
                        ORDER BY sm.id DESC
                        LIMIT 5
                        """,
                        project_ids,
                    ).fetchall()
            else:
                task_rows = []
                document_rows = []
                daily_log_rows = []
                message_rows = []
                stock_rows = []
            recent_activity = []
            for row in task_rows:
                recent_activity.append({
                    "kind": "task",
                    "title": row["title"],
                    "text": f"{row['project_title']} · {row['priority']} · до {row['due_at'] or 'без срока'}",
                    "createdAt": row["created_at"],
                })
            for row in document_rows:
                recent_activity.append({
                    "kind": "document",
                    "title": row["title"],
                    "text": f"{row['project_title']} · {row['status']}",
                    "createdAt": row["created_at"],
                })
            for row in daily_log_rows:
                recent_activity.append({
                    "kind": "log",
                    "title": row["title"],
                    "text": f"{row['project_title']} · {row['report_date']} · {row['work_done'][:80]}",
                    "createdAt": row["created_at"],
                })
            for row in message_rows:
                recent_activity.append({
                    "kind": "message",
                    "title": row["chat_title"],
                    "text": f"{row['project_title']} · {row['body'][:90]}",
                    "createdAt": row["created_at"],
                })
            for row in stock_rows:
                move_label = {
                    "purchase": "Закупка",
                    "receipt": "Поступление",
                    "use": "Использовано",
                    "writeoff": "Списание",
                }.get(row["move_type"], row["move_type"])
                recent_activity.append({
                    "kind": "stock",
                    "title": row["material_title"],
                    "text": f"{row['project_title']} · {move_label} {row['qty']} {row['unit']}",
                    "createdAt": row["created_at"],
                })
            recent_activity = sorted(recent_activity, key=lambda item: item["createdAt"] or 0, reverse=True)[:8]
            today_actions = []
            if critical_items and user["role"] != "customer":
                first = critical_items[0]
                today_actions.append(f"Закрыть нехватку: {first['title']} — нужно ещё {first['missingQty']} {first['unit']}.")
            if open_tasks:
                today_actions.append(f"Разобрать открытые задачи: {open_tasks} шт. по активным объектам.")
            if user_has_any_role(user, {"admin", "director"}) and total_paid - total_spent < 0:
                today_actions.append("Проверить кассовый разрыв: расходы сейчас выше оплат.")
            if user["role"] == "purchaser":
                today_actions.append("Обновить поступления и списания на складе, чтобы смета показывала реальные остатки.")
            if user["role"] == "customer":
                today_actions.append("Проверить график работ, видимые документы и сообщения команды по объекту.")
            if not today_actions:
                today_actions.append("Критических блокеров нет — держим график и фиксируем факт каждый день.")
        data = {
            "projectsCount": len(projects),
            "activeProjects": active,
            "avgProgress": round(sum(float(row["progress"] or 0) for row in projects) / len(projects), 1) if projects else 0,
            "shortagesCount": 0 if user["role"] == "customer" else shortages,
            "openTasksCount": open_tasks,
            "criticalItems": [] if user["role"] == "customer" else critical_items[:10],
            "todayActions": today_actions[:6],
            "recentActivity": recent_activity,
            "projects": [serialize_project(row, user) for row in projects[:6]],
        }
        if user_has_any_role(user, {"admin", "director"}):
            data.update({
                "totalBudget": total_budget,
                "totalPaid": total_paid,
                "totalSpent": total_spent,
                "profitNow": total_paid - total_spent,
            })
        self.send_json(HTTPStatus.OK, data)

    def parse_project_id(self, path: str) -> int | None:
        parts = path.strip("/").split("/")
        try:
            return int(parts[2])
        except (IndexError, ValueError):
            return None

    def parse_daily_log_id(self, path: str) -> int | None:
        parts = path.strip("/").split("/")
        try:
            return int(parts[4])
        except (IndexError, ValueError):
            return None

    def parse_chat_id(self, path: str) -> int | None:
        parts = path.strip("/").split("/")
        try:
            return int(parts[2])
        except (IndexError, ValueError):
            return None

    def parse_document_id(self, path: str) -> int | None:
        parts = path.strip("/").split("/")
        try:
            return int(parts[2])
        except (IndexError, ValueError):
            return None

    def parse_stage_id(self, path: str) -> int | None:
        parts = path.strip("/").split("/")
        try:
            return int(parts[2])
        except (IndexError, ValueError):
            return None

    def parse_task_id(self, path: str) -> int | None:
        parts = path.strip("/").split("/")
        try:
            return int(parts[2])
        except (IndexError, ValueError):
            return None

    def parse_material_id(self, path: str) -> int | None:
        parts = path.strip("/").split("/")
        try:
            return int(parts[2])
        except (IndexError, ValueError):
            return None

    def parse_warehouse_item_id(self, path: str) -> int | None:
        parts = path.strip("/").split("/")
        try:
            return int(parts[2])
        except (IndexError, ValueError):
            return None

    def parse_supplier_offer_id(self, path: str) -> int | None:
        parts = path.strip("/").split("/")
        try:
            return int(parts[2])
        except (IndexError, ValueError):
            return None

    def parse_finance_id(self, path: str) -> int | None:
        parts = path.strip("/").split("/")
        try:
            return int(parts[2])
        except (IndexError, ValueError):
            return None

    def can_access_project(self, user: dict, project_id: int) -> bool:
        if user_has_any_role(user, {"admin", "director"}):
            return True
        with db() as con:
            if user_has_any_role(user, {"foreman"}):
                row = con.execute(
                    """
                    SELECT 1
                    FROM object_assignments
                    WHERE user_id = ? AND object_id = ? AND role_code = 'foreman'
                    """,
                    (user["id"], project_id),
                ).fetchone()
            else:
                row = con.execute(
                    """
                    SELECT 1
                    FROM user_project_access
                    WHERE user_id = ? AND project_id = ?
                    UNION
                    SELECT 1
                    FROM object_assignments
                    WHERE user_id = ? AND object_id = ?
                    """,
                    (user["id"], project_id, user["id"], project_id),
                ).fetchone()
        return bool(row)

    def require_project_access(self, project_id: int) -> dict | None:
        user = self.require_user()
        if not user:
            return None
        if not self.can_access_project(user, project_id):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "project_forbidden"})
            return None
        return user

    def api_project_detail(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        with db() as con:
            row = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
        if not row:
            self.send_json(HTTPStatus.NOT_FOUND, {"error": "project_not_found"})
            return
        self.send_json(HTTPStatus.OK, {"project": serialize_project(row, user)})

    def api_project_assignments(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        with db() as con:
            rows = con.execute(
                """
                SELECT oa.*, u.name AS user_name, u.login AS user_login
                FROM object_assignments oa
                JOIN users u ON u.id = oa.user_id
                WHERE oa.object_id = ?
                ORDER BY oa.is_primary DESC, oa.assigned_at DESC, oa.id DESC
                """,
                (project_id,),
            ).fetchall()
        self.send_json(
            HTTPStatus.OK,
            {
                "assignments": [
                    {
                        "id": row["id"],
                        "objectId": row["object_id"],
                        "userId": row["user_id"],
                        "userName": row["user_name"],
                        "userLogin": row["user_login"],
                        "roleCode": normalize_role(row["role_code"]),
                        "roleLabel": ROLE_LABELS.get(normalize_role(row["role_code"]), row["role_code"]),
                        "responsibility": row["responsibility"],
                        "isPrimary": bool(row["is_primary"]),
                        "assignedAt": row["assigned_at"],
                    }
                    for row in rows
                ]
            },
        )

    def api_create_project_assignment(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        admin = self.require_role({"admin", "director"})
        if not admin:
            return
        payload = self.read_json()
        try:
            user_id = int(payload.get("user_id", payload.get("userId")))
        except (TypeError, ValueError):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_user_id"})
            return
        role_code = normalize_role(str(payload.get("role_code", payload.get("roleCode", ""))).strip())
        responsibility = str(payload.get("responsibility", "")).strip() or None
        is_primary = 1 if payload.get("is_primary", payload.get("isPrimary", False)) else 0
        if role_code not in ROLE_LABELS:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_role"})
            return
        with db() as con:
            project = con.execute("SELECT id FROM projects WHERE id = ?", (project_id,)).fetchone()
            assignee = con.execute("SELECT id, role FROM users WHERE id = ? AND is_active = 1", (user_id,)).fetchone()
            if not project:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "project_not_found"})
                return
            if not assignee:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "user_not_found"})
                return
            if is_primary:
                con.execute(
                    "UPDATE object_assignments SET is_primary = 0 WHERE object_id = ? AND role_code = ?",
                    (project_id, role_code),
                )
            cur = con.execute(
                """
                INSERT INTO object_assignments (object_id, user_id, role_code, responsibility, is_primary, assigned_by, assigned_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(object_id, user_id, role_code) DO UPDATE SET
                    responsibility = excluded.responsibility,
                    is_primary = excluded.is_primary,
                    assigned_by = excluded.assigned_by,
                    assigned_at = excluded.assigned_at
                """,
                (project_id, user_id, role_code, responsibility, is_primary, admin["id"], now_ts()),
            )
            con.execute(
                "INSERT OR IGNORE INTO user_project_access (user_id, project_id) VALUES (?, ?)",
                (user_id, project_id),
            )
            if role_code == "director":
                con.execute("UPDATE projects SET director_id = ?, updated_at = ? WHERE id = ?", (user_id, now_ts(), project_id))
            elif role_code == "foreman":
                con.execute("UPDATE projects SET foreman_id = ?, updated_at = ? WHERE id = ?", (user_id, now_ts(), project_id))
            elif role_code == "purchaser":
                con.execute("UPDATE projects SET buyer_id = ?, updated_at = ? WHERE id = ?", (user_id, now_ts(), project_id))
            elif role_code == "customer":
                con.execute("UPDATE projects SET client_user_id = ?, updated_at = ? WHERE id = ?", (user_id, now_ts(), project_id))
            create_audit(
                con,
                admin["id"],
                "assign_user_to_project",
                "object_assignment",
                cur.lastrowid,
                {"project_id": project_id, "user_id": user_id, "role_code": role_code},
            )
            con.commit()
        self.send_json(HTTPStatus.CREATED, {"ok": True})

    def api_materials_summary(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        with db() as con:
            items = material_summary_rows(con, project_id)
        self.send_json(HTTPStatus.OK, {"items": items})

    def api_project_warehouse_matches(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if user["role"] == "customer":
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        with db() as con:
            matches = warehouse_matches_for_project(con, project_id)
        self.send_json(
            HTTPStatus.OK,
            {
                "threshold": FUZZY_MATCH_THRESHOLD,
                "matches": {str(material_id): match for material_id, match in matches.items()},
            },
        )

    def api_warehouse_items(self) -> None:
        user = self.require_role({"admin", "director", "foreman", "purchaser"})
        if not user:
            return
        with db() as con:
            rows = warehouse_item_rows(con)
        self.send_json(HTTPStatus.OK, {"items": [serialize_warehouse_item(row) for row in rows]})

    def api_warehouse_receipt(self) -> None:
        user = self.require_role({"admin", "director", "foreman", "purchaser"})
        if not user:
            return
        payload = self.read_json()
        mode = str(payload.get("mode", "manual")).strip()
        if mode not in {"manual", "return"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_receipt_mode"})
            return
        try:
            qty = float(payload.get("qty", 0) or 0)
        except (TypeError, ValueError):
            qty = 0.0
        if qty <= 0:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_qty"})
            return

        con = db()
        try:
            con.execute("BEGIN IMMEDIATE")
            if mode == "manual":
                item_type = str(payload.get("item_type", payload.get("itemType", "material"))).strip()
                if item_type not in {"material", "tool"}:
                    con.rollback()
                    self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_item_type"})
                    return
                name = str(payload.get("name", "")).strip()
                unit = str(payload.get("unit", "шт")).strip() or "шт"
                if not name:
                    con.rollback()
                    self.send_json(HTTPStatus.BAD_REQUEST, {"error": "name_required"})
                    return
                warehouse_item_id = payload.get("warehouse_item_id", payload.get("warehouseItemId"))
                try:
                    warehouse_item_id = int(warehouse_item_id) if warehouse_item_id not in (None, "", 0, "0") else None
                except (TypeError, ValueError):
                    con.rollback()
                    self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_warehouse_item_id"})
                    return
                condition_status = str(payload.get("condition_status", payload.get("conditionStatus", ""))).strip()
                if item_type == "tool" and condition_status not in {"Новый", "Б/У", "Требует ремонта"}:
                    condition_status = "Б/У"
                item_row = add_qty_to_warehouse(
                    con,
                    item_type=item_type,
                    name=name,
                    qty=qty,
                    unit=unit,
                    sku=str(payload.get("sku", "")).strip() or None,
                    condition_status=condition_status,
                    category=str(payload.get("category", "")).strip() or ("Инструмент" if item_type == "tool" else None),
                    warehouse_item_id=warehouse_item_id,
                )
                create_audit(
                    con,
                    user["id"],
                    "warehouse_manual_receipt",
                    "warehouse_item",
                    item_row["id"],
                    {"name": item_row["name"], "qty": qty, "unit": item_row["unit"]},
                )
                con.commit()
                self.send_json(HTTPStatus.CREATED, {"ok": True, "mode": mode, "warehouseItem": serialize_warehouse_item(item_row)})
                return

            try:
                project_id = int(payload.get("project_id", payload.get("projectId", 0)) or 0)
                estimate_item_id = int(payload.get("estimate_item_id", payload.get("estimateItemId", 0)) or 0)
            except (TypeError, ValueError):
                project_id = 0
                estimate_item_id = 0
            if not project_id or not estimate_item_id:
                con.rollback()
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_or_material_id"})
                return
            if not self.can_access_project(user, project_id):
                con.rollback()
                self.send_json(HTTPStatus.FORBIDDEN, {"error": "project_forbidden"})
                return
            material = con.execute(
                """
                SELECT id, project_id, title, unit, planned_qty, planned_price, item_kind, article,
                       procurement_status, warehouse_source, warehouse_item_id, notes
                FROM estimate_items
                WHERE id = ? AND project_id = ?
                """,
                (estimate_item_id, project_id),
            ).fetchone()
            if not material:
                con.rollback()
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "material_not_found"})
                return
            summary = {int(item["id"]): item for item in material_summary_rows(con, project_id)}.get(estimate_item_id)
            available = float(summary.get("stockQty") or 0) if summary else 0.0
            if qty > available:
                con.rollback()
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "qty_exceeds_object_stock", "available": available})
                return

            raw_kind = str(material["item_kind"] or "").strip().lower()
            notes = str(material["notes"] or "").lower()
            item_type = "tool" if raw_kind == "tool" or "инструмент" in notes or "tool" in notes else "material"
            item_row = add_qty_to_warehouse(
                con,
                item_type=item_type,
                name=str(material["title"] or "").strip(),
                qty=qty,
                unit=str(material["unit"] or "шт").strip() or "шт",
                sku=str(material["article"] or "").strip() or None,
                warehouse_item_id=material["warehouse_item_id"],
                category="Инструмент" if item_type == "tool" else None,
            )
            message = f"Произведен возврат на склад: {material['title']} — {qty:g} {material['unit']}"
            con.execute(
                """
                INSERT INTO stock_moves (project_id, estimate_item_id, move_type, qty, price, comment, created_by, created_at)
                VALUES (?, ?, 'use', ?, 0, ?, ?, ?)
                """,
                (project_id, estimate_item_id, qty, message, user["id"], now_ts()),
            )
            remaining = max(available - qty, 0)
            con.execute(
                """
                UPDATE estimate_items
                SET procurement_status = ?, warehouse_source = 'warehouse', warehouse_item_id = ?, updated_at = ?
                WHERE id = ?
                """,
                ("Возвращено на склад" if remaining <= 0 else "Частично возвращено на склад", item_row["id"], now_ts(), estimate_item_id),
            )
            create_audit(
                con,
                user["id"],
                "warehouse_return_from_project",
                "estimate_item",
                estimate_item_id,
                {
                    "project_id": project_id,
                    "warehouse_item_id": item_row["id"],
                    "qty": qty,
                    "unit": material["unit"],
                    "message": message,
                },
            )
            con.commit()
            items = material_summary_rows(con, project_id)
            self.send_json(
                HTTPStatus.CREATED,
                {
                    "ok": True,
                    "mode": mode,
                    "message": message,
                    "warehouseItem": serialize_warehouse_item(item_row),
                    "items": items,
                },
            )
        except Exception:
            con.rollback()
            raise
        finally:
            con.close()

    def api_warehouse_transfer(self, path: str) -> None:
        warehouse_item_id = self.parse_warehouse_item_id(path)
        if not warehouse_item_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_warehouse_item_id"})
            return
        user = self.require_role({"admin", "director", "foreman", "purchaser"})
        if not user:
            return
        payload = self.read_json()
        try:
            project_id = int(payload.get("project_id", payload.get("projectId", 0)) or 0)
        except (TypeError, ValueError):
            project_id = 0
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        if not self.can_access_project(user, project_id):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "project_forbidden"})
            return
        try:
            qty = float(payload.get("qty", 0) or 0)
        except (TypeError, ValueError):
            qty = 0.0
        if qty <= 0:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_qty"})
            return

        con = db()
        try:
            con.execute("BEGIN IMMEDIATE")
            warehouse_item = con.execute(
                """
                SELECT id, item_type, category, name, sku, unit, qty, condition_status, created_at, updated_at
                FROM warehouse_items
                WHERE id = ?
                """,
                (warehouse_item_id,),
            ).fetchone()
            if not warehouse_item:
                con.rollback()
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "warehouse_item_not_found"})
                return
            available = float(warehouse_item["qty"] or 0)
            if qty > available:
                con.rollback()
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "qty_exceeds_stock", "available": available})
                return
            project = con.execute("SELECT id, title FROM projects WHERE id = ?", (project_id,)).fetchone()
            if not project:
                con.rollback()
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "project_not_found"})
                return

            material_rows = con.execute(
                """
                SELECT id, project_id, title, unit, planned_qty, planned_price, item_kind, section_title,
                       article, procurement_status, warehouse_source, warehouse_item_id, need_by_date, notes, stage_id
                FROM estimate_items
                WHERE project_id = ?
                """,
                (project_id,),
            ).fetchall()
            best_material = None
            best_score = 0.0
            for material in material_rows:
                if normalize_estimate_item_kind(resolved_estimate_item_kind(material)) == "work":
                    continue
                score = warehouse_search_score(material, warehouse_item)
                if score > best_score:
                    best_score = score
                    best_material = material

            summary_by_id = {int(item["id"]): item for item in material_summary_rows(con, project_id)}
            estimate_item_id = None
            transfer_status = "Со склада"
            matched_existing = bool(best_material and best_score >= FUZZY_MATCH_THRESHOLD)
            if matched_existing:
                estimate_item_id = int(best_material["id"])
                summary = summary_by_id.get(estimate_item_id, {})
                old_missing = float(summary.get("missingQty") or 0)
                new_missing = max(old_missing - qty, 0)
                transfer_status = "Передано со склада" if new_missing <= 0 else "Частично передано со склада"
                con.execute(
                    """
                    UPDATE estimate_items
                    SET procurement_status = ?, warehouse_source = 'warehouse', warehouse_item_id = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (transfer_status, warehouse_item_id, now_ts(), estimate_item_id),
                )
            else:
                item_kind = "tool" if warehouse_item["item_type"] == "tool" else "material"
                notes = "Со склада"
                if warehouse_item["item_type"] == "tool":
                    notes = "Со склада; Тип: Инструмент"
                cur = con.execute(
                    """
                    INSERT INTO estimate_items (
                        project_id, title, unit, planned_qty, planned_price, item_kind, article,
                        procurement_status, warehouse_source, warehouse_item_id, notes, updated_at
                    )
                    VALUES (?, ?, ?, ?, 0, ?, ?, 'Со склада', 'warehouse', ?, ?, ?)
                    """,
                    (
                        project_id,
                        warehouse_item["name"],
                        warehouse_item["unit"],
                        qty,
                        item_kind,
                        warehouse_item["sku"] or None,
                        warehouse_item_id,
                        notes,
                        now_ts(),
                    ),
                )
                estimate_item_id = cur.lastrowid

            comment = str(payload.get("comment", "")).strip()
            move_comment = comment or f"Выдано со склада: {warehouse_item['name']}"
            con.execute(
                """
                INSERT INTO stock_moves (project_id, estimate_item_id, move_type, qty, price, comment, created_by, created_at)
                VALUES (?, ?, 'receipt', ?, 0, ?, ?, ?)
                """,
                (project_id, estimate_item_id, qty, move_comment, user["id"], now_ts()),
            )
            transfer_cur = con.execute(
                """
                INSERT INTO warehouse_transfers (warehouse_item_id, project_id, estimate_item_id, qty, unit, comment, created_by, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (warehouse_item_id, project_id, estimate_item_id, qty, warehouse_item["unit"], comment, user["id"], now_ts()),
            )
            con.execute(
                "UPDATE warehouse_items SET qty = qty - ?, updated_at = ? WHERE id = ?",
                (qty, now_ts(), warehouse_item_id),
            )
            create_audit(
                con,
                user["id"],
                "warehouse_transfer",
                "warehouse_transfer",
                transfer_cur.lastrowid,
                {
                    "warehouse_item_id": warehouse_item_id,
                    "project_id": project_id,
                    "estimate_item_id": estimate_item_id,
                    "qty": qty,
                    "matched_existing": matched_existing,
                    "score": round(best_score, 3),
                },
            )
            con.commit()
            item_row = con.execute(
                """
                SELECT id, item_type, category, name, sku, unit, qty, condition_status, created_at, updated_at
                FROM warehouse_items
                WHERE id = ?
                """,
                (warehouse_item_id,),
            ).fetchone()
            items = material_summary_rows(con, project_id)
        except Exception:
            con.rollback()
            raise
        finally:
            con.close()

        self.send_json(
            HTTPStatus.CREATED,
            {
                "ok": True,
                "transferId": transfer_cur.lastrowid,
                "estimateItemId": estimate_item_id,
                "matchedExisting": matched_existing,
                "matchScore": round(best_score, 3),
                "status": transfer_status,
                "warehouseItem": serialize_warehouse_item(item_row),
                "items": items,
            },
        )

    def api_material_schedule(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if user["role"] == "customer":
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        query = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
        try:
            warning_days = int(query.get("warningDays", query.get("warning_days", ["5"]))[0])
        except (TypeError, ValueError):
            warning_days = 5
        try:
            neutral_days = int(query.get("neutralDays", query.get("neutral_days", ["7"]))[0])
        except (TypeError, ValueError):
            neutral_days = 7
        warning_days = max(1, min(warning_days, 30))
        neutral_days = max(warning_days, min(max(neutral_days, 1), 60))
        fresh = str(query.get("fresh", [""])[0]).lower() in {"1", "true", "yes"}
        with db() as con:
            payload = None
            if not fresh:
                saved = con.execute(
                    "SELECT payload, updated_at FROM material_schedule_snapshots WHERE project_id = ?",
                    (project_id,),
                ).fetchone()
                if saved:
                    try:
                        payload = json.loads(saved["payload"])
                        if isinstance(payload, dict):
                            payload["saved"] = True
                            payload["savedAt"] = saved["updated_at"]
                    except (TypeError, ValueError, json.JSONDecodeError):
                        payload = None
            if payload is None:
                payload = build_material_schedule_payload(
                    con,
                    project_id,
                    warning_days=warning_days,
                    neutral_days=neutral_days,
                )
        self.send_json(HTTPStatus.OK, payload)

    def api_save_material_schedule(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if user["role"] == "customer":
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        payload = self.read_json()
        warning_days = 5
        neutral_days = 7
        try:
            warning_days = int(payload.get("warningDays", payload.get("warning_days", warning_days)))
        except (TypeError, ValueError):
            warning_days = 5
        try:
            neutral_days = int(payload.get("neutralDays", payload.get("neutral_days", neutral_days)))
        except (TypeError, ValueError):
            neutral_days = 7
        warning_days = max(1, min(warning_days, 30))
        neutral_days = max(warning_days, min(max(neutral_days, 1), 60))
        schedule = payload.get("schedule")
        with db() as con:
            if not isinstance(schedule, dict) or not isinstance(schedule.get("items"), list):
                schedule = build_material_schedule_payload(
                    con,
                    project_id,
                    warning_days=warning_days,
                    neutral_days=neutral_days,
                )
            schedule["projectId"] = project_id
            schedule["saved"] = True
            schedule["savedAt"] = now_ts()
            now = now_ts()
            con.execute(
                """
                INSERT INTO material_schedule_snapshots (project_id, payload, created_by, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(project_id) DO UPDATE SET
                    payload = excluded.payload,
                    created_by = excluded.created_by,
                    updated_at = excluded.updated_at
                """,
                (project_id, json.dumps(schedule, ensure_ascii=False), user["id"], now, now),
            )
            create_audit(
                con,
                user["id"],
                "save_material_schedule",
                "project",
                project_id,
                {"items": len(schedule.get("items") or [])},
            )
            con.commit()
        self.send_json(HTTPStatus.OK, schedule)

    def api_create_material(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if not user_has_any_role(user, {"admin", "director"}):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        payload = self.read_json()
        title = str(payload.get("title", "")).strip()
        unit = str(payload.get("unit", "")).strip() or "шт"
        planned_qty, planned_price = normalize_estimate_planned_values(
            unit,
            payload.get("planned_qty", payload.get("plannedQty", 0)),
            payload.get("planned_price", payload.get("plannedPrice", 0)),
        )
        if not title or planned_qty <= 0:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "title_and_qty_required"})
            return
        with db() as con:
            existing = con.execute(
                "SELECT id FROM estimate_items WHERE project_id = ? AND lower(title) = lower(?) AND unit = ?",
                (project_id, title, unit),
            ).fetchone()
            if existing:
                self.send_json(HTTPStatus.CONFLICT, {"error": "material_exists"})
                return
            cur = con.execute(
                """
                INSERT INTO estimate_items (project_id, title, unit, planned_qty, planned_price, item_kind)
                VALUES (?, ?, ?, ?, ?, 'material')
                """,
                (project_id, title, unit, planned_qty, planned_price),
            )
            con.execute(
                """
                INSERT INTO audit_log (user_id, action, entity, entity_id, payload, created_at)
                VALUES (?, 'create_material', 'estimate_item', ?, ?, ?)
                """,
                (user["id"], cur.lastrowid, json.dumps({"project_id": project_id, "title": title}, ensure_ascii=False), now_ts()),
            )
            con.commit()
            items = material_summary_rows(con, project_id)
        self.send_json(HTTPStatus.CREATED, {"id": cur.lastrowid, "items": items})

    def api_update_material(self, path: str) -> None:
        material_id = self.parse_material_id(path)
        if not material_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_material_id"})
            return
        payload = self.read_json()
        with db() as con:
            material = con.execute("SELECT * FROM estimate_items WHERE id = ?", (material_id,)).fetchone()
            if not material:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "material_not_found"})
                return
            user = self.require_project_access(int(material["project_id"]))
            if not user:
                return
            if user["role"] == "customer":
                self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
                return
            stage_id = payload.get("stage_id", payload.get("stageId", material["stage_id"]))
            try:
                stage_id = int(stage_id) if stage_id not in (None, "", 0, "0") else None
            except (TypeError, ValueError):
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_stage_id"})
                return
            if stage_id:
                stage = con.execute(
                    "SELECT id FROM work_stages WHERE id = ? AND project_id = ?",
                    (stage_id, material["project_id"]),
                ).fetchone()
                if not stage:
                    self.send_json(HTTPStatus.BAD_REQUEST, {"error": "stage_not_found"})
                    return
            raw_delivery_days = payload.get("delivery_days", payload.get("deliveryDays", material["delivery_days"]))
            try:
                delivery_days = int(raw_delivery_days) if raw_delivery_days not in (None, "") else None
            except (TypeError, ValueError):
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_delivery_days"})
                return
            if delivery_days is not None:
                delivery_days = max(0, min(delivery_days, 90))
            merged_material = dict(material)
            merged_material.update(payload)
            con.execute(
                """
                UPDATE estimate_items
                SET title = ?, unit = ?, planned_qty = ?, planned_price = ?, stage_id = ?, item_kind = ?, section_title = ?, delivery_days = ?, need_by_date = ?, notes = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    str(payload.get("title", material["title"])).strip() or material["title"],
                    str(payload.get("unit", material["unit"])).strip() or material["unit"],
                    max(0.01, normalize_estimate_planned_qty(
                        payload.get("unit", material["unit"]),
                        payload.get("planned_qty", payload.get("plannedQty", material["planned_qty"])),
                    )),
                    max(0.0, float(payload.get("planned_price", payload.get("plannedPrice", material["planned_price"])) or 0)),
                    stage_id,
                    resolved_estimate_item_kind(merged_material),
                    resolved_estimate_section_title(merged_material),
                    delivery_days,
                    str(payload.get("need_by_date", payload.get("needByDate", material["need_by_date"] or ""))).strip() or None,
                    str(payload.get("notes", material["notes"] or "")).strip() or None,
                    now_ts(),
                    material_id,
                ),
            )
            create_audit(
                con,
                user["id"],
                "update_material",
                "estimate_item",
                material_id,
                {"project_id": material["project_id"], "stage_id": stage_id},
            )
            con.commit()
            items = material_summary_rows(con, int(material["project_id"]))
        self.send_json(HTTPStatus.OK, {"id": material_id, "items": items})

    def api_project_supplier_offers(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if user["role"] == "customer":
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        with db() as con:
            rows = con.execute(
                """
                SELECT so.*, e.title AS material_title, e.planned_price, e.planned_qty, e.unit AS material_unit,
                       c.name AS company_name, u.name AS author_name
                FROM supplier_offers so
                LEFT JOIN estimate_items e ON e.id = so.estimate_item_id
                LEFT JOIN companies c ON c.id = so.company_id
                LEFT JOIN users u ON u.id = so.created_by
                WHERE so.project_id = ?
                ORDER BY CASE so.status
                    WHEN 'selected' THEN 0
                    WHEN 'quoted' THEN 1
                    WHEN 'called' THEN 2
                    WHEN 'new' THEN 3
                    ELSE 4 END,
                    so.id DESC
                """,
                (project_id,),
            ).fetchall()
        offers = []
        for row in rows:
            item = dict(row)
            planned_price = float(row["planned_price"] or 0)
            price = float(row["price"] or 0)
            qty = float(row["qty"] or row["planned_qty"] or 0)
            delta_per_unit = round(price - planned_price, 2) if planned_price else None
            delta_total = round((price - planned_price) * qty, 2) if planned_price and qty else None
            item["compareToEstimate"] = {
                "plannedPrice": planned_price or None,
                "deltaPerUnit": delta_per_unit,
                "deltaTotal": delta_total,
                "effectLabel": "Экономия" if delta_total is not None and delta_total < 0 else ("Переплата" if delta_total is not None and delta_total > 0 else "Без сравнения"),
            }
            offers.append(item)
        self.send_json(HTTPStatus.OK, {"offers": offers})

    def api_project_market_analysis(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if user["role"] == "customer":
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        query = urllib.parse.urlparse(self.path).query
        params = urllib.parse.parse_qs(query)
        kind = str((params.get("kind") or ["material"])[0] or "material").strip().lower()
        if kind not in {"material", "work"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_kind"})
            return
        try:
            with db() as con:
                payload = build_project_market_analysis(con, project_id, kind)
        except LookupError as exc:
            self.send_json(HTTPStatus.NOT_FOUND, {"error": str(exc)})
            return
        except urllib.error.URLError:
            self.send_json(HTTPStatus.BAD_GATEWAY, {"error": "autobot_unavailable"})
            return
        except Exception:
            self.send_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"error": "market_analysis_failed"})
            return
        self.send_json(HTTPStatus.OK, payload)

    def api_create_market_counterparty(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if not user_can_manage_suppliers(user):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        payload = self.read_json()
        candidate_type = str(payload.get("candidate_type", payload.get("candidateType", "supplier"))).strip() or "supplier"
        if candidate_type not in {"supplier", "contractor"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_candidate_type"})
            return
        name = str(payload.get("name", payload.get("candidate_name", payload.get("candidateName", "")))).strip()
        if not name:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "name_required"})
            return
        source_type = str(payload.get("source_type", payload.get("sourceType", "manual"))).strip() or "manual"
        if source_type not in {"manual", "avito", "other"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_source_type"})
            return
        try:
            estimate_item_id = int(payload.get("estimate_item_id", payload.get("estimateItemId", 0)) or 0)
        except (TypeError, ValueError):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_estimate_item_id"})
            return
        if not estimate_item_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "estimate_item_required"})
            return

        def payload_float(key: str) -> float:
            try:
                return float(payload.get(key, 0) or 0)
            except (TypeError, ValueError):
                return 0.0

        phone = str(payload.get("phone", "")).strip() or None
        source_url = str(payload.get("source_url", payload.get("sourceUrl", ""))).strip() or None
        notes = str(payload.get("notes", "")).strip()
        company_notes = "\n".join(part for part in [notes, f"Источник: {source_url}" if source_url else ""] if part) or None
        price = max(0.0, payload_float("price"))
        qty = max(0.0, payload_float("qty"))
        unit = str(payload.get("unit", "")).strip() or None
        timestamp = now_ts()
        with db() as con:
            material = con.execute(
                "SELECT id, title, unit, planned_qty FROM estimate_items WHERE id = ? AND project_id = ?",
                (estimate_item_id, project_id),
            ).fetchone()
            if not material:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "material_not_found"})
                return
            if not unit:
                unit = str(material["unit"] or "") or None
            if qty <= 0:
                qty = float(material["planned_qty"] or 0)
            company_cur = con.execute(
                """
                INSERT INTO companies (type, name, inn, kpp, ogrn, phone, email, address, notes, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (candidate_type, name, None, None, None, phone, None, None, company_notes, timestamp),
            )
            company_id = int(company_cur.lastrowid)
            offer_cur = con.execute(
                """
                INSERT INTO supplier_offers (
                    project_id, estimate_item_id, company_id, candidate_type, candidate_name, source_type, source_url,
                    contact_name, phone, price, qty, unit, status, notes, created_by, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'selected', ?, ?, ?, ?)
                """,
                (
                    project_id,
                    estimate_item_id,
                    company_id,
                    candidate_type,
                    name,
                    source_type,
                    source_url,
                    None,
                    phone,
                    price,
                    qty,
                    unit,
                    notes or None,
                    user["id"],
                    timestamp,
                    timestamp,
                ),
            )
            con.execute(
                """
                UPDATE supplier_offers
                SET status = 'quoted', updated_at = ?
                WHERE project_id = ? AND estimate_item_id = ? AND candidate_type = ? AND id <> ? AND status = 'selected'
                """,
                (timestamp, project_id, estimate_item_id, candidate_type, offer_cur.lastrowid),
            )
            create_audit(con, user["id"], "create_market_counterparty", "company", company_id, {"project_id": project_id, "estimate_item_id": estimate_item_id, "type": candidate_type})
            create_audit(con, user["id"], "select_market_counterparty", "supplier_offer", offer_cur.lastrowid, {"project_id": project_id, "company_id": company_id})
            con.commit()
        self.send_json(
            HTTPStatus.CREATED,
            {
                "company": {
                    "id": company_id,
                    "type": candidate_type,
                    "name": name,
                    "inn": None,
                    "kpp": None,
                    "ogrn": None,
                    "phone": phone,
                    "email": None,
                    "address": None,
                    "notes": company_notes,
                    "created_at": timestamp,
                },
                "offer": {
                    "id": int(offer_cur.lastrowid),
                    "project_id": project_id,
                    "estimate_item_id": estimate_item_id,
                    "company_id": company_id,
                    "candidate_type": candidate_type,
                    "candidate_name": name,
                    "status": "selected",
                    "source_url": source_url,
                    "phone": phone,
                },
            },
        )

    def api_create_supplier_offer(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if not user_can_manage_suppliers(user):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        payload = self.read_json()
        candidate_type = str(payload.get("candidate_type", payload.get("candidateType", "supplier"))).strip() or "supplier"
        if candidate_type not in {"supplier", "contractor"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_candidate_type"})
            return
        status = str(payload.get("status", "new")).strip() or "new"
        if status not in {"new", "called", "quoted", "rejected", "selected"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_status"})
            return
        source_type = str(payload.get("source_type", payload.get("sourceType", "manual"))).strip() or "manual"
        if source_type not in {"manual", "avito", "other"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_source_type"})
            return
        company_id = payload.get("company_id", payload.get("companyId"))
        estimate_item_id = payload.get("estimate_item_id", payload.get("estimateItemId"))
        try:
            company_id = int(company_id) if company_id not in (None, "", 0, "0") else None
            estimate_item_id = int(estimate_item_id) if estimate_item_id not in (None, "", 0, "0") else None
        except (TypeError, ValueError):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_relation_id"})
            return
        candidate_name = str(payload.get("candidate_name", payload.get("candidateName", ""))).strip()
        price = float(payload.get("price", 0) or 0)
        qty = float(payload.get("qty", 0) or 0)
        unit = str(payload.get("unit", "")).strip() or None
        if not candidate_name:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "candidate_name_required"})
            return
        with db() as con:
            if estimate_item_id:
                material = con.execute(
                    "SELECT id, unit, planned_qty FROM estimate_items WHERE id = ? AND project_id = ?",
                    (estimate_item_id, project_id),
                ).fetchone()
                if not material:
                    self.send_json(HTTPStatus.BAD_REQUEST, {"error": "material_not_found"})
                    return
                if not unit:
                    unit = str(material["unit"] or "") or None
                if qty <= 0:
                    qty = float(material["planned_qty"] or 0)
            if company_id:
                company = con.execute("SELECT id FROM companies WHERE id = ?", (company_id,)).fetchone()
                if not company:
                    self.send_json(HTTPStatus.BAD_REQUEST, {"error": "company_not_found"})
                    return
            timestamp = now_ts()
            cur = con.execute(
                """
                INSERT INTO supplier_offers (
                    project_id, estimate_item_id, company_id, candidate_type, candidate_name, source_type, source_url,
                    contact_name, phone, price, qty, unit, status, notes, created_by, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    estimate_item_id,
                    company_id,
                    candidate_type,
                    candidate_name,
                    source_type,
                    str(payload.get("source_url", payload.get("sourceUrl", ""))).strip() or None,
                    str(payload.get("contact_name", payload.get("contactName", ""))).strip() or None,
                    str(payload.get("phone", "")).strip() or None,
                    max(0.0, price),
                    max(0.0, qty),
                    unit,
                    status,
                    str(payload.get("notes", "")).strip() or None,
                    user["id"],
                    timestamp,
                    timestamp,
                ),
            )
            if status == "selected" and estimate_item_id:
                con.execute(
                    """
                    UPDATE supplier_offers
                    SET status = 'quoted', updated_at = ?
                    WHERE project_id = ? AND estimate_item_id = ? AND candidate_type = ? AND id <> ? AND status = 'selected'
                    """,
                    (timestamp, project_id, estimate_item_id, candidate_type, cur.lastrowid),
                )
            create_audit(con, user["id"], "create_supplier_offer", "supplier_offer", cur.lastrowid, {"project_id": project_id, "candidate_name": candidate_name, "status": status})
            con.commit()
        self.send_json(HTTPStatus.CREATED, {"id": cur.lastrowid})

    def api_clear_supplier_selection(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if not user_can_manage_suppliers(user):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        payload = self.read_json()
        candidate_type = str(payload.get("candidate_type", payload.get("candidateType", "supplier"))).strip() or "supplier"
        if candidate_type not in {"supplier", "contractor"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_candidate_type"})
            return
        try:
            estimate_item_id = int(payload.get("estimate_item_id", payload.get("estimateItemId", 0)) or 0)
        except (TypeError, ValueError):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_estimate_item_id"})
            return
        if not estimate_item_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "estimate_item_required"})
            return
        timestamp = now_ts()
        with db() as con:
            con.execute(
                """
                UPDATE supplier_offers
                SET status = 'quoted', updated_at = ?
                WHERE project_id = ? AND estimate_item_id = ? AND candidate_type = ? AND status = 'selected'
                """,
                (timestamp, project_id, estimate_item_id, candidate_type),
            )
            create_audit(con, user["id"], "clear_supplier_selection", "estimate_item", estimate_item_id, {"project_id": project_id, "candidate_type": candidate_type})
            con.commit()
        self.send_json(HTTPStatus.OK, {"project_id": project_id, "estimate_item_id": estimate_item_id, "candidate_type": candidate_type})

    def api_update_supplier_offer(self, path: str) -> None:
        offer_id = self.parse_supplier_offer_id(path)
        if not offer_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_offer_id"})
            return
        payload = self.read_json()
        with db() as con:
            offer = con.execute("SELECT * FROM supplier_offers WHERE id = ?", (offer_id,)).fetchone()
            if not offer:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "offer_not_found"})
                return
            user = self.require_project_access(int(offer["project_id"]))
            if not user:
                return
            if not user_can_manage_suppliers(user):
                self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
                return
            status = str(payload.get("status", offer["status"])).strip() or "new"
            if status not in {"new", "called", "quoted", "rejected", "selected"}:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_status"})
                return
            source_type = str(payload.get("source_type", payload.get("sourceType", offer["source_type"]))).strip() or "manual"
            if source_type not in {"manual", "avito", "other"}:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_source_type"})
                return
            company_id = payload.get("company_id", payload.get("companyId", offer["company_id"]))
            try:
                company_id = int(company_id) if company_id not in (None, "", 0, "0") else None
            except (TypeError, ValueError):
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_company_id"})
                return
            if company_id:
                company = con.execute("SELECT id FROM companies WHERE id = ?", (company_id,)).fetchone()
                if not company:
                    self.send_json(HTTPStatus.BAD_REQUEST, {"error": "company_not_found"})
                    return
            timestamp = now_ts()
            if status == "selected" and offer["estimate_item_id"]:
                con.execute(
                    """
                    UPDATE supplier_offers
                    SET status = 'quoted', updated_at = ?
                    WHERE project_id = ? AND estimate_item_id = ? AND candidate_type = ? AND id <> ? AND status = 'selected'
                    """,
                    (timestamp, offer["project_id"], offer["estimate_item_id"], offer["candidate_type"], offer_id),
                )
            con.execute(
                """
                UPDATE supplier_offers
                SET candidate_name = ?, company_id = ?, source_type = ?, source_url = ?, contact_name = ?, phone = ?,
                    price = ?, qty = ?, unit = ?, status = ?, notes = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    str(payload.get("candidate_name", payload.get("candidateName", offer["candidate_name"]))).strip() or offer["candidate_name"],
                    company_id,
                    source_type,
                    str(payload.get("source_url", payload.get("sourceUrl", offer["source_url"] or ""))).strip() or None,
                    str(payload.get("contact_name", payload.get("contactName", offer["contact_name"] or ""))).strip() or None,
                    str(payload.get("phone", offer["phone"] or "")).strip() or None,
                    max(0.0, float(payload.get("price", offer["price"]) or 0)),
                    max(0.0, float(payload.get("qty", offer["qty"]) or 0)),
                    str(payload.get("unit", offer["unit"] or "")).strip() or None,
                    status,
                    str(payload.get("notes", offer["notes"] or "")).strip() or None,
                    timestamp,
                    offer_id,
                ),
            )
            create_audit(con, user["id"], "update_supplier_offer", "supplier_offer", offer_id, {"project_id": offer["project_id"], "status": status})
            con.commit()
        self.send_json(HTTPStatus.OK, {"id": offer_id})

    def api_project_bootstrap(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if not user_has_any_role(user, {"admin", "director"}):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        payload = self.read_json()
        project_payload = payload.get("project") if isinstance(payload.get("project"), dict) else {}
        stages_payload = payload.get("stages") if isinstance(payload.get("stages"), list) else []
        materials_payload = payload.get("materials") if isinstance(payload.get("materials"), list) else []
        tasks_payload = payload.get("tasks") if isinstance(payload.get("tasks"), list) else []
        replace_existing = bool(payload.get("replace_existing", payload.get("replaceExisting", True)))
        if not stages_payload and not materials_payload and not tasks_payload and not project_payload:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bootstrap_payload_empty"})
            return

        with db() as con:
            project = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            if not project:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "project_not_found"})
                return

            # Update core project fields from the parsed tender/estimate package.
            if project_payload:
                updates = (
                    str(project_payload.get("title", project["title"])).strip() or project["title"],
                    str(project_payload.get("address", project["address"])).strip() or project["address"],
                    str(project_payload.get("client_name", project_payload.get("clientName", project["client_name"]))).strip() or project["client_name"],
                    str(project_payload.get("contract_no", project_payload.get("contractNo", project["contract_no"] or ""))).strip() or None,
                    str(project_payload.get("contract_date", project_payload.get("contractDate", project["contract_date"] or ""))).strip() or None,
                    str(project_payload.get("city", project["city"] or "")).strip() or None,
                    str(project_payload.get("region", project["region"] or "")).strip() or None,
                    float(project_payload.get("budget", project["budget"] or 0) or 0),
                    str(project_payload.get("started_at", project_payload.get("startedAt", project["started_at"] or ""))).strip() or None,
                    str(project_payload.get("deadline_at", project_payload.get("deadlineAt", project["deadline_at"] or ""))).strip() or None,
                    str(project_payload.get("description", project["description"] or "")).strip() or None,
                    now_ts(),
                    project_id,
                )
                con.execute(
                    """
                    UPDATE projects
                    SET title = ?, address = ?, client_name = ?, contract_no = ?, contract_date = ?,
                        city = ?, region = ?, budget = ?, started_at = ?, deadline_at = ?, description = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    updates,
                )

            if replace_existing:
                con.execute("DELETE FROM work_stages WHERE project_id = ?", (project_id,))
                con.execute("DELETE FROM estimate_items WHERE project_id = ?", (project_id,))
                con.execute("DELETE FROM tasks WHERE project_id = ?", (project_id,))

            stage_map: dict[str, int] = {}

            def insert_stage_nodes(nodes: list[dict], parent_id: int | None = None, depth: int = 0) -> None:
                for index, item in enumerate(nodes, start=1):
                    if not isinstance(item, dict):
                        continue
                    title = str(item.get("title", "")).strip()
                    if not title:
                        continue
                    stage_kind = str(item.get("stage_kind", item.get("stageKind", ("section" if depth == 0 else "work")))).strip() or "work"
                    status_code = str(item.get("status_code", item.get("statusCode", "not_started"))).strip() or "not_started"
                    key = str(item.get("key", item.get("code", title))).strip() or title
                    cur = con.execute(
                        """
                        INSERT INTO work_stages (
                            project_id, title, position, parent_id, stage_kind, status_code,
                            planned_start, planned_end, customer_start, customer_end, progress,
                            responsible, notes, is_client_visible, depends_on_materials, created_at, updated_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            project_id,
                            title,
                            int(item.get("position", index) or index),
                            parent_id,
                            stage_kind,
                            status_code,
                            str(item.get("planned_start", item.get("plannedStart", ""))).strip() or None,
                            str(item.get("planned_end", item.get("plannedEnd", ""))).strip() or None,
                            str(item.get("customer_start", item.get("customerStart", ""))).strip() or None,
                            str(item.get("customer_end", item.get("customerEnd", ""))).strip() or None,
                            max(0, min(100, int(item.get("progress", 0) or 0))),
                            str(item.get("responsible", "")).strip() or None,
                            str(item.get("notes", "")).strip() or None,
                            1 if item.get("is_client_visible", item.get("isClientVisible", True)) else 0,
                            1 if item.get("depends_on_materials", item.get("dependsOnMaterials", False)) else 0,
                            now_ts(),
                            now_ts(),
                        ),
                    )
                    stage_id = int(cur.lastrowid)
                    stage_map[key] = stage_id
                    stage_map[title] = stage_id
                    children = item.get("children") if isinstance(item.get("children"), list) else []
                    if children:
                        insert_stage_nodes(children, stage_id, depth + 1)

            insert_stage_nodes(stages_payload)

            for item in materials_payload:
                if not isinstance(item, dict):
                    continue
                title = str(item.get("title", "")).strip()
                if not title:
                    continue
                stage_ref = str(item.get("stage_key", item.get("stageKey", item.get("stage_title", item.get("stageTitle", ""))))).strip()
                stage_id = stage_map.get(stage_ref) if stage_ref else None
                item_kind = resolved_estimate_item_kind(item)
                section_title = resolved_estimate_section_title(item)
                unit = str(item.get("unit", "шт")).strip() or "шт"
                planned_qty, planned_price = normalize_estimate_planned_values(
                    unit,
                    item.get("planned_qty", item.get("plannedQty", 0)),
                    item.get("planned_price", item.get("plannedPrice", 0)),
                )
                con.execute(
                    """
                    INSERT INTO estimate_items (project_id, title, unit, planned_qty, planned_price, stage_id, item_kind, section_title, need_by_date, notes, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        project_id,
                        title,
                        unit,
                        max(0.01, planned_qty),
                        max(0.0, planned_price),
                        stage_id,
                        item_kind,
                        section_title,
                        str(item.get("need_by_date", item.get("needByDate", ""))).strip() or None,
                        str(item.get("notes", "")).strip() or None,
                        now_ts(),
                    ),
                )

            for item in tasks_payload:
                if not isinstance(item, dict):
                    continue
                title = str(item.get("title", "")).strip()
                if not title:
                    continue
                assignee_id = item.get("assignee_id", item.get("assigneeId"))
                try:
                    assignee_id = int(assignee_id) if assignee_id not in (None, "", 0, "0") else None
                except (TypeError, ValueError):
                    assignee_id = None
                con.execute(
                    """
                    INSERT INTO tasks (project_id, title, description, status, priority, assignee_id, due_at, created_by, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        project_id,
                        title,
                        str(item.get("description", "")).strip(),
                        str(item.get("status", "open")).strip() or "open",
                        str(item.get("priority", "normal")).strip() or "normal",
                        assignee_id,
                        str(item.get("due_at", item.get("dueAt", ""))).strip() or None,
                        user["id"],
                        now_ts(),
                        now_ts(),
                    ),
                )

            # If parser didn't send tasks, generate minimal startup tasks from imported structure.
            if not tasks_payload:
                generated = []
                if stages_payload:
                    generated.append(("Подтвердить график и структуру этапов", "Проверить импортированные разделы и сроки объекта", "high"))
                if materials_payload:
                    generated.append(("Проверить потребность по материалам", "Сверить импортированные материалы и даты потребности", "high"))
                if project_payload.get("contract_no") or project_payload.get("contract_date"):
                    generated.append(("Проверить договорные условия объекта", "Убедиться, что предстроительные условия и даты старта учтены", "normal"))
                for title, description, priority in generated:
                    con.execute(
                        """
                        INSERT INTO tasks (project_id, title, description, status, priority, created_by, created_at, updated_at)
                        VALUES (?, ?, ?, 'open', ?, ?, ?, ?)
                        """,
                        (project_id, title, description, priority, user["id"], now_ts(), now_ts()),
                    )

            create_audit(
                con,
                user["id"],
                "bootstrap_project",
                "project",
                project_id,
                {
                    "replace_existing": replace_existing,
                    "stages": len(stages_payload),
                    "materials": len(materials_payload),
                    "tasks": len(tasks_payload),
                },
            )
            mark_project_schedule_draft(con, project_id)
            con.commit()
            project_row = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            stages = con.execute("SELECT COUNT(*) FROM work_stages WHERE project_id = ?", (project_id,)).fetchone()[0]
            materials = con.execute("SELECT COUNT(*) FROM estimate_items WHERE project_id = ?", (project_id,)).fetchone()[0]
            tasks = con.execute("SELECT COUNT(*) FROM tasks WHERE project_id = ?", (project_id,)).fetchone()[0]

        self.send_json(
            HTTPStatus.OK,
            {
                "project": serialize_project(project_row, user),
                "summary": {"stages": stages, "materials": materials, "tasks": tasks},
            },
        )

    def api_import_estimate(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if not user_has_any_role(user, {"admin", "director"}):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return

        payload = self.read_json()
        items = payload.get("items")
        replace = bool(payload.get("replace", False))
        if not isinstance(items, list) or not items:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "items_required"})
            return

        normalized = []
        for item in items:
            title = str(item.get("title", "")).strip()
            unit = str(item.get("unit", "")).strip() or "шт"
            planned_qty, planned_price = normalize_estimate_planned_values(
                unit,
                item.get("planned_qty", item.get("plannedQty", 0)),
                item.get("planned_price", item.get("plannedPrice", 0)),
            )
            if not title or planned_qty <= 0:
                continue
            item_kind = resolved_estimate_item_kind(item)
            section_title = resolved_estimate_section_title(item)
            normalized.append((project_id, title, unit, planned_qty, planned_price, item_kind, section_title))

        if not normalized:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "no_valid_items"})
            return

        with db() as con:
            if replace:
                con.execute("DELETE FROM estimate_items WHERE project_id = ?", (project_id,))
            imported = 0
            for normalized_item in normalized:
                _, title, unit, planned_qty, planned_price, item_kind, section_title = normalized_item
                existing = con.execute(
                    "SELECT id FROM estimate_items WHERE project_id = ? AND lower(title) = lower(?) AND unit = ?",
                    (project_id, title, unit),
                ).fetchone()
                if existing:
                    con.execute(
                        """
                        UPDATE estimate_items
                        SET planned_qty = ?, planned_price = ?, item_kind = ?, section_title = ?
                        WHERE id = ?
                        """,
                        (planned_qty, planned_price, item_kind, section_title, existing["id"]),
                    )
                else:
                    con.execute(
                        """
                        INSERT INTO estimate_items (project_id, title, unit, planned_qty, planned_price, item_kind, section_title)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        normalized_item,
                    )
                imported += 1
            con.execute(
                """
                INSERT INTO audit_log (user_id, action, entity, entity_id, payload, created_at)
                VALUES (?, 'import_estimate', 'project', ?, ?, ?)
                """,
                (
                    user["id"],
                    project_id,
                    json.dumps({"count": imported, "replace": replace}, ensure_ascii=False),
                    now_ts(),
                ),
            )
            con.commit()

            summary = material_summary_rows(con, project_id)
        self.send_json(HTTPStatus.CREATED, {"imported": imported, "items": summary})

    def api_project_auto_schedule(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if not user_can_manage_schedule(user):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        payload = self.read_json()
        with db() as con:
            project = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            if not project:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "project_not_found"})
                return
            stages = con.execute(
                "SELECT * FROM work_stages WHERE project_id = ? ORDER BY position, id",
                (project_id,),
            ).fetchall()
            materials = con.execute(
                """
                SELECT id, title, unit, planned_qty, planned_price, stage_id, need_by_date, notes
                FROM estimate_items
                WHERE project_id = ?
                ORDER BY id
                """,
                (project_id,),
            ).fetchall()
            if not stages:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "stages_required"})
                return
            if not materials:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "materials_required"})
                return
            requested_start = parse_iso_date(payload.get("start_date", payload.get("startDate")))
            project_start = parse_iso_date(project["started_at"])
            start_at = requested_start or project_start or date(2026, 7, 25)
            plan = build_auto_schedule_plan(project, stages, materials, start_at)
            for item in plan["stage_updates"]:
                con.execute(
                    """
                    UPDATE work_stages
                    SET planned_start = ?, planned_end = ?, customer_start = ?, customer_end = ?,
                        depends_on_materials = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        item["planned_start"],
                        item["planned_end"],
                        item["customer_start"],
                        item["customer_end"],
                        item["depends_on_materials"],
                        now_ts(),
                        item["id"],
                    ),
                )
            auto_stage_map = {item["id"]: item["stage_id"] for item in plan["auto_linked_materials"]}
            for item in plan["material_updates"]:
                target_stage_id = auto_stage_map.get(item["id"])
                if target_stage_id:
                    con.execute(
                        """
                        UPDATE estimate_items
                        SET stage_id = ?, need_by_date = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        (target_stage_id, item["need_by_date"], now_ts(), item["id"]),
                    )
                else:
                    con.execute(
                        "UPDATE estimate_items SET need_by_date = ?, updated_at = ? WHERE id = ?",
                        (item["need_by_date"], now_ts(), item["id"]),
                    )
            if not project["started_at"]:
                con.execute(
                    "UPDATE projects SET started_at = ?, updated_at = ? WHERE id = ?",
                    (plan["project_start"], now_ts(), project_id),
                )
            mark_project_schedule_draft(con, project_id, generated_at=TODAY_ISO)
            material_schedule = build_material_schedule_payload(con, project_id)
            create_audit(
                con,
                user["id"],
                "auto_schedule_project",
                "project",
                project_id,
                {
                    "start_date": plan["project_start"],
                    "finish_date": plan["project_end"],
                    "auto_linked_materials": len(plan["auto_linked_materials"]),
                    "deadline_overrun_days": plan["deadline_overrun_days"],
                },
            )
            con.commit()
            project_row = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
        self.send_json(
            HTTPStatus.OK,
            {
                "project": serialize_project(project_row, user),
                "summary": {
                    "projectStart": plan["project_start"],
                    "projectEnd": plan["project_end"],
                    "stagesPlanned": len(plan["stage_updates"]),
                    "materialsPlanned": len(plan["material_updates"]),
                    "materialsAutoLinked": len(plan["auto_linked_materials"]),
                    "deadlineOverrunDays": plan["deadline_overrun_days"],
                    "longestStages": plan["longest_stages"],
                    "procurementHotspots": plan["material_updates"][:8],
                },
                "materialSchedule": material_schedule,
            },
        )

    def api_project_section_schedule_forecast(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        payload = self.read_json()
        requested_start = parse_iso_date(payload.get("start_date", payload.get("startDate")))
        with db() as con:
            project = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            if not project:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "project_not_found"})
                return
            rows = con.execute(
                """
                SELECT id, title, unit, planned_qty, item_kind, section_title
                FROM estimate_items
                WHERE project_id = ?
                ORDER BY id
                """,
                (project_id,),
            ).fetchall()
        work_items = [
            row
            for row in rows
            if normalize_estimate_item_kind(row["item_kind"]) == "work"
        ]
        if not work_items:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "works_required"})
            return
        project_start = parse_iso_date(project["started_at"])
        start_at = requested_start or project_start or date(2026, 7, 27)
        forecast = build_section_schedule_forecast(project, work_items, start_at)
        self.send_json(HTTPStatus.OK, forecast)

    def api_project_schedule_status(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        with db() as con:
            project = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
        if not project:
            self.send_json(HTTPStatus.NOT_FOUND, {"error": "project_not_found"})
            return
        self.send_json(HTTPStatus.OK, {"scheduleControl": project_schedule_payload(project)})

    def api_update_project_schedule_status(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if not user_can_manage_schedule(user):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        payload = self.read_json()
        schedule_type = str(payload.get("schedule_type", payload.get("scheduleType", "internal"))).strip() or "internal"
        action = str(payload.get("action", "approve")).strip() or "approve"
        if schedule_type not in {"internal", "customer", "both"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_schedule_type"})
            return
        if action not in {"approve", "reset_to_draft"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_action"})
            return
        with db() as con:
            project = update_project_schedule_status(con, project_id, schedule_type, action)
            if not project:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "project_not_found"})
                return
            create_audit(
                con,
                user["id"],
                "update_project_schedule_status",
                "project",
                project_id,
                {"schedule_type": schedule_type, "action": action},
            )
            con.commit()
        self.send_json(HTTPStatus.OK, {"project": serialize_project(project, user)})

    def api_project_analysis(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        with db() as con:
            project = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            materials = material_summary_rows(con, project_id)
            open_tasks = con.execute(
                "SELECT COUNT(*) FROM tasks WHERE project_id = ? AND status != 'done'",
                (project_id,),
            ).fetchone()[0]
            stages = con.execute(
                "SELECT * FROM work_stages WHERE project_id = ? ORDER BY position",
                (project_id,),
            ).fetchall()
        if not project:
            self.send_json(HTTPStatus.NOT_FOUND, {"error": "project_not_found"})
            return

        shortages = [item for item in materials if item["missingQty"] > 0]
        low_stock = [
            item for item in materials
            if item["stockQty"] <= max(item["plannedQty"] * 0.1, 1) and item["usageProgress"] < 100
        ]
        avg_stage_progress = round(sum(float(stage["progress"]) for stage in stages) / len(stages), 1) if stages else float(project["progress"] or 0)
        material_purchase_progress = round(
            sum(item["purchaseProgress"] for item in materials) / len(materials),
            1,
        ) if materials else 0

        risks = []
        actions = []
        if shortages:
            for item in shortages:
                risks.append({
                    "level": "critical",
                    "title": f"Не хватает: {item['title']}",
                    "text": f"По смете нужно {item['plannedQty']} {item['unit']}, закрыто закупкой/поступлением {max(item['purchasedQty'], item['receivedQty'])}. Не хватает {item['missingQty']} {item['unit']}.",
                })
                actions.append(f"Создать заявку на закупку: {item['title']} — {item['missingQty']} {item['unit']}.")
        if low_stock:
            for item in low_stock:
                risks.append({
                    "level": "warning",
                    "title": f"Малый остаток: {item['title']}",
                    "text": f"Остаток {item['stockQty']} {item['unit']}, использовано {item['usedQty']} из {item['plannedQty']}.",
                })
        if open_tasks:
            actions.append(f"Проверить открытые задачи по объекту: {open_tasks}.")
        if not actions:
            actions.append("Критичных действий сейчас нет. Продолжать контроль списаний и графика.")

        payload = {
            "projectId": project_id,
            "projectProgress": project["progress"],
            "stageProgress": avg_stage_progress,
            "materialPurchaseProgress": material_purchase_progress,
            "shortagesCount": len(shortages),
            "openTasksCount": open_tasks,
            "risks": risks,
            "actions": actions,
        }
        if user["role"] == "customer":
            payload["risks"] = [risk for risk in risks if risk["level"] != "critical"][:2]
            payload["actions"] = ["Команда ведёт контроль материалов и графика. Критичные внутренние действия скрыты."]
        self.send_json(HTTPStatus.OK, payload)

    def api_project_stages(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        with db() as con:
            if user["role"] == "customer":
                rows = con.execute(
                    """
                    SELECT *
                    FROM work_stages
                    WHERE project_id = ? AND is_client_visible = 1
                    ORDER BY position, id
                    """,
                    (project_id,),
                ).fetchall()
            else:
                rows = con.execute(
                    "SELECT * FROM work_stages WHERE project_id = ? ORDER BY position, id",
                    (project_id,),
                ).fetchall()
        self.send_json(HTTPStatus.OK, {"stages": [dict(row) for row in rows]})

    def api_create_project_stage(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if not user_can_manage_documents(user):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        payload = self.read_json()
        title = str(payload.get("title", "")).strip()
        stage_kind = str(payload.get("stage_kind", payload.get("stageKind", "section"))).strip() or "section"
        if not title or stage_kind not in {"section", "subsection", "work"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "invalid_stage_data"})
            return
        parent_id = payload.get("parent_id", payload.get("parentId"))
        try:
            parent_id = int(parent_id) if parent_id not in (None, "", 0, "0") else None
        except (TypeError, ValueError):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_parent_id"})
            return
        with db() as con:
            project = con.execute("SELECT id FROM projects WHERE id = ?", (project_id,)).fetchone()
            if not project:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "project_not_found"})
                return
            if parent_id:
                parent = con.execute(
                    "SELECT id FROM work_stages WHERE id = ? AND project_id = ?",
                    (parent_id, project_id),
                ).fetchone()
                if not parent:
                    self.send_json(HTTPStatus.BAD_REQUEST, {"error": "parent_not_found"})
                    return
            next_position = con.execute(
                "SELECT COALESCE(MAX(position), 0) + 1 FROM work_stages WHERE project_id = ?",
                (project_id,),
            ).fetchone()[0]
            cur = con.execute(
                """
                INSERT INTO work_stages (
                    project_id, title, position, parent_id, stage_kind, status_code,
                    planned_start, planned_end, customer_start, customer_end, progress,
                    responsible, notes, is_client_visible, depends_on_materials, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    title,
                    int(payload.get("position", next_position) or next_position),
                    parent_id,
                    stage_kind,
                    str(payload.get("status_code", payload.get("statusCode", "not_started"))).strip() or "not_started",
                    str(payload.get("planned_start", payload.get("plannedStart", ""))).strip() or None,
                    str(payload.get("planned_end", payload.get("plannedEnd", ""))).strip() or None,
                    str(payload.get("customer_start", payload.get("customerStart", ""))).strip() or None,
                    str(payload.get("customer_end", payload.get("customerEnd", ""))).strip() or None,
                    max(0, min(100, int(payload.get("progress", 0) or 0))),
                    str(payload.get("responsible", "")).strip() or None,
                    str(payload.get("notes", "")).strip() or None,
                    1 if payload.get("is_client_visible", payload.get("isClientVisible", True)) else 0,
                    1 if payload.get("depends_on_materials", payload.get("dependsOnMaterials", False)) else 0,
                    now_ts(),
                    now_ts(),
                ),
            )
            mark_project_schedule_draft(con, project_id)
            create_audit(con, user["id"], "create_stage", "work_stage", cur.lastrowid, {"project_id": project_id, "title": title, "stage_kind": stage_kind})
            project_row = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            row = con.execute("SELECT * FROM work_stages WHERE id = ?", (cur.lastrowid,)).fetchone()
            con.commit()
        self.send_json(HTTPStatus.CREATED, {"stage": dict(row), "project": serialize_project(project_row, user)})

    def api_update_stage(self, path: str) -> None:
        stage_id = self.parse_stage_id(path)
        if not stage_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_stage_id"})
            return
        payload = self.read_json()
        with db() as con:
            stage = con.execute("SELECT * FROM work_stages WHERE id = ?", (stage_id,)).fetchone()
            if not stage:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "stage_not_found"})
                return
            user = self.require_project_access(int(stage["project_id"]))
            if not user:
                return
            if not user_can_manage_documents(user):
                self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
                return
            status_code = str(payload.get("status_code", payload.get("statusCode", stage["status_code"]))).strip() or "not_started"
            if status_code not in {"not_started", "started", "in_progress", "completed", "approved", "blocked", "overdue"}:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_status"})
                return
            progress = max(0, min(100, int(payload.get("progress", stage["progress"]) or 0)))
            con.execute(
                """
                UPDATE work_stages
                SET title = ?, stage_kind = ?, status_code = ?, planned_start = ?, planned_end = ?,
                    customer_start = ?, customer_end = ?, fact_start = ?, fact_end = ?, progress = ?,
                    responsible = ?, notes = ?, is_client_visible = ?, depends_on_materials = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    str(payload.get("title", stage["title"])).strip() or stage["title"],
                    str(payload.get("stage_kind", payload.get("stageKind", stage["stage_kind"]))).strip() or stage["stage_kind"],
                    status_code,
                    str(payload.get("planned_start", payload.get("plannedStart", stage["planned_start"] or ""))).strip() or None,
                    str(payload.get("planned_end", payload.get("plannedEnd", stage["planned_end"] or ""))).strip() or None,
                    str(payload.get("customer_start", payload.get("customerStart", stage["customer_start"] or ""))).strip() or None,
                    str(payload.get("customer_end", payload.get("customerEnd", stage["customer_end"] or ""))).strip() or None,
                    str(payload.get("fact_start", payload.get("factStart", stage["fact_start"] or ""))).strip() or None,
                    str(payload.get("fact_end", payload.get("factEnd", stage["fact_end"] or ""))).strip() or None,
                    progress,
                    str(payload.get("responsible", stage["responsible"] or "")).strip() or None,
                    str(payload.get("notes", stage["notes"] or "")).strip() or None,
                    1 if payload.get("is_client_visible", payload.get("isClientVisible", bool(stage["is_client_visible"]))) else 0,
                    1 if payload.get("depends_on_materials", payload.get("dependsOnMaterials", bool(stage["depends_on_materials"]))) else 0,
                    now_ts(),
                    stage_id,
                ),
            )
            mark_project_schedule_draft(con, int(stage["project_id"]))
            create_audit(con, user["id"], "update_stage", "work_stage", stage_id, {"project_id": stage["project_id"], "status_code": status_code, "progress": progress})
            project_row = con.execute("SELECT * FROM projects WHERE id = ?", (stage["project_id"],)).fetchone()
            row = con.execute("SELECT * FROM work_stages WHERE id = ?", (stage_id,)).fetchone()
            con.commit()
        self.send_json(HTTPStatus.OK, {"stage": dict(row), "project": serialize_project(project_row, user)})

    def api_project_tasks(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if user["role"] == "customer":
            self.send_json(HTTPStatus.OK, {"tasks": []})
            return
        with db() as con:
            rows = con.execute(
                """
                SELECT t.*, u.name AS assignee_name
                FROM tasks t
                LEFT JOIN users u ON u.id = t.assignee_id
                WHERE t.project_id = ?
                ORDER BY CASE t.priority WHEN 'high' THEN 0 WHEN 'normal' THEN 1 ELSE 2 END, t.id DESC
                """,
                (project_id,),
            ).fetchall()
        self.send_json(HTTPStatus.OK, {"tasks": [dict(row) for row in rows]})

    def api_create_task(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if not user_has_any_role(user, {"admin", "director"}):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        payload = self.read_json()
        title = str(payload.get("title", "")).strip()
        if not title:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "title_required"})
            return
        assignee_id = payload.get("assignee_id", payload.get("assigneeId"))
        try:
            assignee_id = int(assignee_id) if assignee_id not in (None, "", 0, "0") else None
        except (TypeError, ValueError):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_assignee_id"})
            return
        with db() as con:
            if assignee_id:
                assignee = con.execute("SELECT id FROM users WHERE id = ?", (assignee_id,)).fetchone()
                if not assignee:
                    self.send_json(HTTPStatus.BAD_REQUEST, {"error": "assignee_not_found"})
                    return
            cur = con.execute(
                """
                INSERT INTO tasks (project_id, title, description, status, priority, assignee_id, due_at, created_by, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    title,
                    str(payload.get("description", "")).strip(),
                    str(payload.get("status", "open")).strip() or "open",
                    str(payload.get("priority", "normal")).strip() or "normal",
                    assignee_id,
                    str(payload.get("due_at", "")).strip() or None,
                    user["id"],
                    now_ts(),
                    now_ts(),
                ),
            )
            con.commit()
        self.send_json(HTTPStatus.CREATED, {"id": cur.lastrowid})

    def api_update_task(self, path: str) -> None:
        task_id = self.parse_task_id(path)
        if not task_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_task_id"})
            return
        payload = self.read_json()
        with db() as con:
            task = con.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
            if not task:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "task_not_found"})
                return
            user = self.require_project_access(int(task["project_id"]))
            if not user:
                return
            if user["role"] == "customer":
                self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
                return
            status = str(payload.get("status", task["status"])).strip() or "open"
            if status not in {"open", "in_progress", "review", "done"}:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_status"})
                return
            priority = str(payload.get("priority", task["priority"])).strip() or "normal"
            if priority not in {"low", "normal", "high"}:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_priority"})
                return
            assignee_id = payload.get("assignee_id", payload.get("assigneeId", task["assignee_id"]))
            try:
                assignee_id = int(assignee_id) if assignee_id not in (None, "", 0, "0") else None
            except (TypeError, ValueError):
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_assignee_id"})
                return
            if assignee_id:
                assignee = con.execute("SELECT id FROM users WHERE id = ?", (assignee_id,)).fetchone()
                if not assignee:
                    self.send_json(HTTPStatus.BAD_REQUEST, {"error": "assignee_not_found"})
                    return
            con.execute(
                """
                UPDATE tasks
                SET title = ?, description = ?, status = ?, priority = ?, assignee_id = ?, due_at = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    str(payload.get("title", task["title"])).strip() or task["title"],
                    str(payload.get("description", task["description"] or "")).strip(),
                    status,
                    priority,
                    assignee_id,
                    str(payload.get("due_at", payload.get("dueAt", task["due_at"] or ""))).strip() or None,
                    now_ts(),
                    task_id,
                ),
            )
            create_audit(
                con,
                user["id"],
                "update_task",
                "task",
                task_id,
                {"project_id": task["project_id"], "status": status, "priority": priority},
            )
            row = con.execute(
                """
                SELECT t.*, u.name AS assignee_name
                FROM tasks t
                LEFT JOIN users u ON u.id = t.assignee_id
                WHERE t.id = ?
                """,
                (task_id,),
            ).fetchone()
            con.commit()
        self.send_json(HTTPStatus.OK, {"task": dict(row)})

    def api_project_notifications(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        today = TODAY_ISO
        today_date = parse_iso_date(today) or date(2026, 7, 26)
        soon_limit = (today_date + timedelta(days=2)).isoformat()
        with db() as con:
            project = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            open_tasks = con.execute(
                """
                SELECT t.*, u.name AS assignee_name
                FROM tasks t
                LEFT JOIN users u ON u.id = t.assignee_id
                WHERE t.project_id = ? AND t.status != 'done'
                ORDER BY t.due_at IS NULL, t.due_at ASC, t.id DESC
                """,
                (project_id,),
            ).fetchall()
            latest_log = con.execute(
                """
                SELECT id, report_date, title, blockers, created_at
                FROM daily_logs
                WHERE project_id = ?
                ORDER BY report_date DESC, id DESC
                LIMIT 1
                """,
                (project_id,),
            ).fetchone()
            blocker_logs = con.execute(
                """
                SELECT id, report_date, title, blockers
                FROM daily_logs
                WHERE project_id = ? AND TRIM(COALESCE(blockers, '')) != ''
                ORDER BY report_date DESC, id DESC
                LIMIT 3
                """,
                (project_id,),
            ).fetchall()
            materials = []
            section_start_dates: dict[str, str] = {}
            if user["role"] == "customer":
                stage_rows = con.execute(
                    """
                    SELECT id, title, parent_id, stage_kind, status_code, planned_start, planned_end, progress, responsible
                    FROM work_stages
                    WHERE project_id = ? AND is_client_visible = 1
                    ORDER BY position, id
                    """,
                    (project_id,),
                ).fetchall()
            else:
                stage_rows = con.execute(
                    """
                    SELECT id, title, parent_id, stage_kind, status_code, planned_start, planned_end, progress, responsible
                    FROM work_stages
                    WHERE project_id = ?
                    ORDER BY position, id
                    """,
                    (project_id,),
                ).fetchall()
                materials = material_summary_rows(con, project_id)
                work_rows = con.execute(
                    """
                    SELECT id, title, unit, planned_qty, item_kind, section_title
                    FROM estimate_items
                    WHERE project_id = ?
                    ORDER BY id
                    """,
                    (project_id,),
                ).fetchall()
                work_items = [row for row in work_rows if normalize_estimate_item_kind(row["item_kind"]) == "work"]
                if project and work_items:
                    forecast_start = parse_iso_date(project["started_at"]) or today_date
                    forecast = build_section_schedule_forecast(project, work_items, forecast_start)
                    for section in forecast.get("sections", []):
                        title = str(section.get("title") or "").strip()
                        start_date = str(section.get("startDate") or "").strip()
                        if title and start_date:
                            section_start_dates[title] = start_date
        stage_items = [dict(row) for row in stage_rows]
        stage_map = {int(item["id"]): item for item in stage_items if item.get("id") is not None}

        def stage_section_title(item: dict) -> str:
            current = item
            root = item
            guard = 0
            while current and current.get("parent_id") and guard < 32:
                parent = stage_map.get(int(current["parent_id"]))
                if not parent:
                    break
                root = parent
                current = parent
                guard += 1
            title = str((root or item).get("title") or item.get("title") or "").strip()
            return title or "График работ"

        stage_matches = []
        for item in stage_items:
            normalized_title = normalize_schedule_text(item.get("title"))
            if normalized_title:
                stage_matches.append((len(normalized_title), normalized_title, item))
        stage_matches.sort(reverse=True, key=lambda entry: entry[0])

        def annotate_task_section(item: dict) -> dict:
            text = normalize_schedule_text(" ".join([str(item.get("title") or ""), str(item.get("description") or "")]))
            for _, normalized_title, stage in stage_matches:
                if normalized_title and normalized_title in text:
                    item["stageTitle"] = str(stage.get("title") or "").strip()
                    item["sectionTitle"] = stage_section_title(stage)
                    return item
            item["sectionTitle"] = "Задачи объекта"
            return item

        overdue_tasks = []
        due_soon_tasks = []
        for row in open_tasks:
            item = annotate_task_section(dict(row))
            due_at = str(item.get("due_at") or "")
            if due_at and due_at < today:
                overdue_tasks.append(item)
            elif due_at and due_at <= soon_limit:
                due_soon_tasks.append(item)
        problem_stages = []
        for item in stage_items:
            item["sectionTitle"] = stage_section_title(item)
            status_code = str(item.get("status_code") or "")
            progress = int(item.get("progress") or 0)
            planned_end = str(item.get("planned_end") or "")
            if status_code == "blocked" or (planned_end and planned_end < today and progress < 100):
                problem_stages.append(item)
        procurement = build_procurement_alerts(materials, stage_rows, today_date, section_start_dates) if user["role"] != "customer" else {"items": [], "summary": {"critical": 0, "soon": 0, "watch": 0}}
        data = {
            "today": today,
            "missingDailyReport": False if user["role"] == "customer" else not bool(latest_log and latest_log["report_date"] == today),
            "latestDailyLog": dict(latest_log) if latest_log else None,
            "overdueTasks": overdue_tasks[:8],
            "dueSoonTasks": due_soon_tasks[:8],
            "blockerLogs": [dict(row) for row in blocker_logs],
            "problemStages": problem_stages[:8],
            "procurementAlerts": procurement["items"],
            "procurementSummary": procurement["summary"],
        }
        self.send_json(HTTPStatus.OK, data)

    def recalc_project_finance_totals(self, con: sqlite3.Connection, project_id: int) -> None:
        spent = con.execute(
            """
            SELECT COALESCE(SUM(amount), 0)
            FROM finance_entries
            WHERE project_id = ? AND direction = 'expense' AND status = 'paid'
            """,
            (project_id,),
        ).fetchone()[0]
        paid = con.execute(
            """
            SELECT COALESCE(SUM(amount), 0)
            FROM finance_entries
            WHERE project_id = ? AND direction = 'income' AND status = 'paid'
            """,
            (project_id,),
        ).fetchone()[0]
        con.execute(
            "UPDATE projects SET spent = ?, paid = ?, updated_at = ? WHERE id = ?",
            (float(spent or 0), float(paid or 0), now_ts(), project_id),
        )

    def api_project_finances(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if not user_can_view_finances(user):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        if user_has_any_role(user, {"foreman"}) and not user_can_manage_finances(user):
            self.send_json(
                HTTPStatus.OK,
                {
                    "items": [],
                    "summary": {},
                    "canUploadInvoice": True,
                    "limited": True,
                },
            )
            return
        with db() as con:
            rows = con.execute(
                """
                SELECT
                    f.*,
                    u.name AS created_by_name,
                    d.original_name AS document_original_name,
                    d.mime_type AS document_mime_type,
                    d.file_ext AS document_file_ext,
                    d.storage_path AS document_storage_path
                FROM finance_entries f
                LEFT JOIN users u ON u.id = f.created_by
                LEFT JOIN documents d ON d.id = f.document_id
                WHERE f.project_id = ?
                ORDER BY COALESCE(f.paid_date, f.planned_date, '') DESC, f.id DESC
                """,
                (project_id,),
            ).fetchall()
            estimate_total = con.execute(
                """
                SELECT COALESCE(SUM(planned_qty * planned_price), 0)
                FROM estimate_items
                WHERE project_id = ?
                """,
                (project_id,),
            ).fetchone()[0]
        items = []
        for row in rows:
            item = dict(row)
            if item.get("document_id"):
                item["document"] = {
                    "id": item["document_id"],
                    "original_name": item.get("document_original_name"),
                    "mime_type": item.get("document_mime_type"),
                    "file_ext": item.get("document_file_ext"),
                    "download_url": f"/api/documents/{item['document_id']}/download",
                    "view_url": f"/api/documents/{item['document_id']}/view",
                    "can_preview": bool(item.get("document_storage_path"))
                    and (
                        str(item.get("document_mime_type") or "").startswith("image/")
                        or str(item.get("document_file_ext") or "") == ".pdf"
                    ),
                }
            items.append(item)
        summary = {
            "estimateTotal": float(estimate_total or 0),
            "plannedExpense": sum(float(item["amount"] or 0) for item in items if item["direction"] == "expense" and item["status"] != "cancelled"),
            "paidExpense": sum(float(item["amount"] or 0) for item in items if item["direction"] == "expense" and item["status"] == "paid"),
            "paidIncome": sum(float(item["amount"] or 0) for item in items if item["direction"] == "income" and item["status"] == "paid"),
        }
        summary["balance"] = summary["paidIncome"] - summary["paidExpense"]
        self.send_json(HTTPStatus.OK, {"items": items, "summary": summary})

    def api_create_finance_entry(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        payload = self.read_json()
        direction = str(payload.get("direction", "expense")).strip() or "expense"
        if direction not in {"income", "expense"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_direction"})
            return
        payment_kind = str(payload.get("payment_kind", payload.get("paymentKind", "cash"))).strip() or "cash"
        if payment_kind not in {"cash", "bank_no_vat", "bank_vat"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_payment_kind"})
            return
        status = str(payload.get("status", "planned")).strip() or "planned"
        if status not in {"planned", "approved", "paid", "cancelled"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_status"})
            return
        if not user_can_manage_finances(user):
            if not user_has_any_role(user, {"foreman"}) or direction != "expense" or status not in {"planned", "approved"}:
                self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
                return
        try:
            amount = float(payload.get("amount", 0) or 0)
            vat_percent = float(payload.get("vat_percent", payload.get("vatPercent", 0)) or 0)
        except (TypeError, ValueError):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_amount"})
            return
        if amount <= 0:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "amount_required"})
            return
        paid_date = str(payload.get("paid_date", payload.get("paidDate", ""))).strip() or None
        if not user_can_manage_finances(user):
            paid_date = None
        with db() as con:
            cur = con.execute(
                """
                INSERT INTO finance_entries (
                    project_id, direction, category, payment_kind, vat_percent, amount,
                    planned_date, paid_date, counterparty_name, status, notes,
                    created_by, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    direction,
                    str(payload.get("category", "")).strip() or None,
                    payment_kind,
                    vat_percent,
                    amount,
                    str(payload.get("planned_date", payload.get("plannedDate", ""))).strip() or None,
                    paid_date,
                    str(payload.get("counterparty_name", payload.get("counterpartyName", ""))).strip() or None,
                    status,
                    str(payload.get("notes", "")).strip() or None,
                    user["id"],
                    now_ts(),
                    now_ts(),
                ),
            )
            self.recalc_project_finance_totals(con, project_id)
            create_audit(
                con,
                user["id"],
                "create_finance_entry",
                "finance_entry",
                cur.lastrowid,
                {"project_id": project_id, "direction": direction, "amount": amount, "status": status},
            )
            con.commit()
        self.send_json(HTTPStatus.CREATED, {"id": cur.lastrowid})

    def api_upload_finance_invoice(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if not user_can_manage_finances(user) and not user_has_any_role(user, {"foreman"}):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return

        form = self.read_multipart()
        upload = form["file"] if "file" in form else None
        if upload is None or not getattr(upload, "file", None) or not getattr(upload, "filename", ""):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "file_required"})
            return

        original_name = sanitize_filename(upload.filename)
        file_ext = document_extension(original_name)
        if file_ext not in FINANCE_INVOICE_EXTENSIONS:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_invoice_format"})
            return

        raw = upload.file.read()
        if not raw:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "empty_file"})
            return
        if len(raw) > MAX_UPLOAD_BYTES:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "upload_too_large"})
            return

        parsed_invoice: dict[str, object] | None = None
        if file_ext in {".xlsx", ".xls"}:
            try:
                parsed_invoice = parse_finance_excel_invoice(raw, file_ext)
            except FinanceExcelParseError:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": FINANCE_EXCEL_PARSE_ERROR})
                return

        category = str((parsed_invoice or {}).get("category") or form.getfirst("category", "")).strip()
        counterparty_name = str((parsed_invoice or {}).get("counterparty_name") or form.getfirst("counterparty_name", "")).strip()
        planned_date = str(form.getfirst("planned_date", "")).strip() or None
        payment_kind = str(form.getfirst("payment_kind", "bank_no_vat")).strip() or "bank_no_vat"
        if payment_kind not in {"cash", "bank_no_vat", "bank_vat"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_payment_kind"})
            return
        status = str(form.getfirst("status", "approved")).strip() or "approved"
        if status not in {"planned", "approved"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_status"})
            return
        try:
            amount = float((parsed_invoice or {}).get("amount") or form.getfirst("amount", "0") or 0)
        except (TypeError, ValueError):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_amount"})
            return
        if amount <= 0:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "amount_required"})
            return
        if not category:
            category = Path(original_name).stem

        notes = str(form.getfirst("notes", "")).strip() or None
        vat_percent = 20 if payment_kind == "bank_vat" else 0
        mime_type = upload.type or mimetypes.guess_type(original_name)[0] or "application/octet-stream"
        storage_name = f"{now_ts()}_{secrets.token_hex(8)}{file_ext}"
        file_path = project_documents_dir(project_id) / storage_name
        file_path.write_bytes(raw)

        with db() as con:
            doc_cur = con.execute(
                """
                INSERT INTO documents (
                    project_id, title, doc_type, status, original_name, storage_name, storage_path,
                    mime_type, file_ext, size_bytes, notes, uploaded_by, is_client_visible, created_at, updated_at
                )
                VALUES (?, ?, 'invoice', 'submitted', ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
                """,
                (
                    project_id,
                    category,
                    original_name,
                    storage_name,
                    str(file_path.relative_to(PROJECT_ROOT)),
                    mime_type,
                    file_ext,
                    len(raw),
                    notes,
                    user["id"],
                    now_ts(),
                    now_ts(),
                ),
            )
            finance_cur = con.execute(
                """
                INSERT INTO finance_entries (
                    project_id, direction, category, payment_kind, vat_percent, amount,
                    planned_date, paid_date, counterparty_name, document_id, status, notes,
                    created_by, created_at, updated_at
                )
                VALUES (?, 'expense', ?, ?, ?, ?, ?, NULL, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    category,
                    payment_kind,
                    vat_percent,
                    amount,
                    planned_date,
                    counterparty_name or None,
                    doc_cur.lastrowid,
                    status,
                    notes,
                    user["id"],
                    now_ts(),
                    now_ts(),
                ),
            )
            self.recalc_project_finance_totals(con, project_id)
            create_audit(
                con,
                user["id"],
                "upload_finance_invoice",
                "finance_entry",
                finance_cur.lastrowid,
                {
                    "project_id": project_id,
                    "document_id": doc_cur.lastrowid,
                    "amount": amount,
                    "status": status,
                    "parsed_excel": bool(parsed_invoice),
                },
            )
            con.commit()
        self.send_json(
            HTTPStatus.CREATED,
            {
                "id": finance_cur.lastrowid,
                "document_id": doc_cur.lastrowid,
                "parsedInvoice": parsed_invoice,
                "status": status,
            },
        )

    def api_update_finance_entry(self, path: str) -> None:
        finance_id = self.parse_finance_id(path)
        if not finance_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_finance_id"})
            return
        payload = self.read_json()
        with db() as con:
            row = con.execute("SELECT * FROM finance_entries WHERE id = ?", (finance_id,)).fetchone()
            if not row:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "finance_not_found"})
                return
            user = self.require_project_access(int(row["project_id"]))
            if not user:
                return
            if not user_can_manage_finances(user):
                self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
                return
            status = str(payload.get("status", row["status"])).strip() or row["status"]
            if status not in {"planned", "approved", "paid", "cancelled"}:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_status"})
                return
            if status == "paid" and str(row["status"] or "") != "paid" and not user_can_pay_invoices(user):
                self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
                return
            con.execute(
                """
                UPDATE finance_entries
                SET planned_date = ?, paid_date = ?, status = ?, notes = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    str(payload.get("planned_date", payload.get("plannedDate", row["planned_date"] or ""))).strip() or None,
                    str(payload.get("paid_date", payload.get("paidDate", row["paid_date"] or ""))).strip() or None,
                    status,
                    str(payload.get("notes", row["notes"] or "")).strip() or None,
                    now_ts(),
                    finance_id,
                ),
            )
            self.recalc_project_finance_totals(con, int(row["project_id"]))
            create_audit(
                con,
                user["id"],
                "update_finance_entry",
                "finance_entry",
                finance_id,
                {"project_id": row["project_id"], "status": status},
            )
            con.commit()
        self.send_json(HTTPStatus.OK, {"ok": True})

    def api_pay_invoice(self) -> None:
        director = self.require_role({"director"})
        if not director:
            return
        payload = self.read_json()
        try:
            finance_id = int(payload.get("finance_id", payload.get("financeId", payload.get("invoice_id", payload.get("invoiceId", payload.get("id"))))))
        except (TypeError, ValueError):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_finance_id"})
            return
        paid_date = str(payload.get("paid_date", payload.get("paidDate", TODAY_ISO))).strip() or TODAY_ISO
        with db() as con:
            row = con.execute("SELECT * FROM finance_entries WHERE id = ?", (finance_id,)).fetchone()
            if not row:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "finance_not_found"})
                return
            if row["direction"] != "expense":
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "not_expense_invoice"})
                return
            user = self.require_project_access(int(row["project_id"]))
            if not user:
                return
            if not user_can_pay_invoices(user):
                self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
                return
            con.execute(
                """
                UPDATE finance_entries
                SET status = 'paid', paid_date = ?, updated_at = ?
                WHERE id = ?
                """,
                (paid_date, now_ts(), finance_id),
            )
            self.recalc_project_finance_totals(con, int(row["project_id"]))
            create_audit(
                con,
                director["id"],
                "pay_invoice",
                "finance_entry",
                finance_id,
                {"project_id": row["project_id"], "paid_date": paid_date},
            )
            con.commit()
        self.send_json(HTTPStatus.OK, {"ok": True, "id": finance_id, "status": "paid", "paidDate": paid_date})

    def api_project_documents(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        with db() as con:
            if user["role"] == "customer":
                rows = con.execute(
                    """
                    SELECT d.*, u.name AS uploaded_by_name, s.title AS stage_title
                    FROM documents d
                    LEFT JOIN users u ON u.id = d.uploaded_by
                    LEFT JOIN work_stages s ON s.id = d.stage_id
                    WHERE d.project_id = ? AND d.is_client_visible = 1
                    ORDER BY d.id DESC
                    """,
                    (project_id,),
                ).fetchall()
            else:
                rows = con.execute(
                    """
                    SELECT d.*, u.name AS uploaded_by_name, s.title AS stage_title
                    FROM documents d
                    LEFT JOIN users u ON u.id = d.uploaded_by
                    LEFT JOIN work_stages s ON s.id = d.stage_id
                    WHERE d.project_id = ?
                    ORDER BY d.id DESC
                    """,
                    (project_id,),
                ).fetchall()
        documents = []
        for row in rows:
            item = dict(row)
            item["download_url"] = f"/api/documents/{row['id']}/download"
            item["view_url"] = f"/api/documents/{row['id']}/view"
            item["can_preview"] = bool(row["storage_path"]) and str(row["mime_type"] or "").startswith(("image/", "text/")) or str(row["file_ext"] or "") == ".pdf"
            documents.append(item)
        self.send_json(HTTPStatus.OK, {"documents": documents})

    def api_project_executive_docs(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return

        with db() as con:
            project = con.execute("SELECT id, title FROM projects WHERE id = ?", (project_id,)).fetchone()
            if not project:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "project_not_found"})
                return
            stages = con.execute(
                """
                SELECT id, title, stage_kind, status_code, progress, planned_end, fact_end, parent_id
                FROM work_stages
                WHERE project_id = ?
                ORDER BY position ASC, id ASC
                """,
                (project_id,),
            ).fetchall()
            docs = con.execute(
                """
                SELECT id, title, doc_type, status, stage_id, template_code, is_client_visible, created_at
                FROM documents
                WHERE project_id = ?
                ORDER BY id DESC
                """,
                (project_id,),
            ).fetchall()

        docs_by_stage: dict[int, list[dict]] = {}
        for row in docs:
            stage_id = int(row["stage_id"] or 0)
            if not stage_id:
                continue
            docs_by_stage.setdefault(stage_id, []).append(dict(row))

        checklist = []
        required_total = 0
        ready_total = 0
        for stage in stages:
            stage_templates = executive_templates_for_stage(stage)
            if not stage_templates:
                continue
            stage_docs = docs_by_stage.get(int(stage["id"]), [])
            items = []
            for template in stage_templates:
                same_docs = [
                    item for item in stage_docs
                    if str(item.get("template_code") or item.get("doc_type") or "") == template["code"]
                ]
                ready_count = sum(1 for item in same_docs if executive_ready_status(item.get("status")))
                required_total += 0 if template["optional"] else 1
                ready_total += 0 if template["optional"] else min(ready_count, 1)
                items.append(
                    {
                        "code": template["code"],
                        "title": template["title"],
                        "docType": template["doc_type"],
                        "optional": template["optional"],
                        "defaultNotes": template["default_notes"],
                        "existingCount": len(same_docs),
                        "readyCount": ready_count,
                        "isReady": ready_count > 0,
                    }
                )
            checklist.append(
                {
                    "stageId": int(stage["id"]),
                    "stageTitle": stage["title"],
                    "stageKind": stage["stage_kind"],
                    "statusCode": stage["status_code"],
                    "progress": int(stage["progress"] or 0),
                    "plannedEnd": stage["planned_end"],
                    "factEnd": stage["fact_end"],
                    "items": items,
                }
            )

        self.send_json(
            HTTPStatus.OK,
            {
                "project": {"id": int(project["id"]), "title": project["title"]},
                "canManage": user_can_manage_documents(user),
                "summary": {
                    "stages": len(checklist),
                    "required": required_total,
                    "ready": ready_total,
                    "missing": max(0, required_total - ready_total),
                },
                "checklist": checklist,
            },
        )

    def api_create_project_executive_doc(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if not user_can_manage_documents(user):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return

        payload = self.read_json()
        template_code = str(payload.get("template_code", payload.get("templateCode", ""))).strip()
        stage_id = int(payload.get("stage_id", payload.get("stageId", 0)) or 0)
        if template_code not in EXECUTIVE_TEMPLATE_RULES:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_template_code"})
            return
        if not stage_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_stage_id"})
            return

        template = EXECUTIVE_TEMPLATE_RULES[template_code]
        with db() as con:
            stage = con.execute(
                """
                SELECT id, project_id, title, stage_kind, status_code, progress
                FROM work_stages
                WHERE id = ? AND project_id = ?
                """,
                (stage_id, project_id),
            ).fetchone()
            if not stage:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "stage_not_found"})
                return
            allowed_templates = {item["code"] for item in executive_templates_for_stage(stage)}
            if template_code not in allowed_templates:
                self.send_json(HTTPStatus.BAD_REQUEST, {"error": "template_not_allowed_for_stage"})
                return

            existing_count = con.execute(
                """
                SELECT COUNT(*) FROM documents
                WHERE project_id = ? AND stage_id = ? AND (template_code = ? OR doc_type = ?)
                """,
                (project_id, stage_id, template_code, template["doc_type"]),
            ).fetchone()[0]
            suffix = f" #{existing_count + 1}" if existing_count else ""
            title = f"{template['title']} — {stage['title']}{suffix}"
            title = f"{template['title']} - {stage['title']}{suffix}"
            notes_parts = [template["default_notes"]]
            custom_notes = str(payload.get("notes", "")).strip()
            if custom_notes:
                notes_parts.append(custom_notes)
            cur = con.execute(
                """
                INSERT INTO documents (
                    project_id, title, doc_type, status, notes, uploaded_by,
                    is_client_visible, created_at, updated_at, stage_id, template_code, generated_by_system
                )
                VALUES (?, ?, ?, 'draft', ?, ?, ?, ?, ?, ?, ?, 1)
                """,
                (
                    project_id,
                    title,
                    template["doc_type"],
                    "\n".join(part for part in notes_parts if part).strip() or None,
                    user["id"],
                    0,
                    now_ts(),
                    now_ts(),
                    stage_id,
                    template_code,
                ),
            )
            create_audit(
                con,
                user["id"],
                "create_executive_document",
                "document",
                cur.lastrowid,
                {"project_id": project_id, "stage_id": stage_id, "template_code": template_code},
            )
            row = con.execute(
                """
                SELECT d.*, u.name AS uploaded_by_name, s.title AS stage_title
                FROM documents d
                LEFT JOIN users u ON u.id = d.uploaded_by
                LEFT JOIN work_stages s ON s.id = d.stage_id
                WHERE d.id = ?
                """,
                (cur.lastrowid,),
            ).fetchone()
            con.commit()

        item = dict(row)
        item["download_url"] = f"/api/documents/{row['id']}/download"
        item["view_url"] = f"/api/documents/{row['id']}/view"
        item["can_preview"] = False
        self.send_json(HTTPStatus.CREATED, {"document": item})

    def api_upload_project_document(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if not user_can_manage_documents(user):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return

        form = self.read_multipart()
        upload = form["file"] if "file" in form else None
        if upload is None or not getattr(upload, "file", None) or not getattr(upload, "filename", ""):
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "file_required"})
            return

        original_name = sanitize_filename(upload.filename)
        raw = upload.file.read()
        if not raw:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "empty_file"})
            return
        if len(raw) > MAX_UPLOAD_BYTES:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "upload_too_large"})
            return

        doc_type = str(form.getfirst("doc_type", "file")).strip() or "file"
        status = str(form.getfirst("status", "draft")).strip() or "draft"
        is_client_visible = 1 if str(form.getfirst("is_client_visible", "0")).strip() in {"1", "true", "yes", "on"} else 0
        title = str(form.getfirst("title", "")).strip() or Path(original_name).stem
        notes = str(form.getfirst("notes", "")).strip() or None
        file_ext = document_extension(original_name)
        mime_type = upload.type or mimetypes.guess_type(original_name)[0] or "application/octet-stream"
        storage_name = f"{now_ts()}_{secrets.token_hex(8)}{file_ext}"
        file_path = project_documents_dir(project_id) / storage_name
        file_path.write_bytes(raw)

        with db() as con:
            cur = con.execute(
                """
                INSERT INTO documents (
                    project_id, title, doc_type, status, original_name, storage_name, storage_path,
                    mime_type, file_ext, size_bytes, notes, uploaded_by, is_client_visible, created_at, updated_at,
                    stage_id, template_code, generated_by_system
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0)
                """,
                (
                    project_id,
                    title,
                    doc_type,
                    status,
                    original_name,
                    storage_name,
                    str(file_path.relative_to(PROJECT_ROOT)),
                    mime_type,
                    file_ext,
                    len(raw),
                    notes,
                    user["id"],
                    is_client_visible,
                    now_ts(),
                    now_ts(),
                    int(str(form.getfirst("stage_id", "0")).strip() or "0") or None,
                    str(form.getfirst("template_code", "")).strip() or None,
                ),
            )
            create_audit(
                con,
                user["id"],
                "upload_document",
                "document",
                cur.lastrowid,
                {"project_id": project_id, "title": title, "doc_type": doc_type},
            )
            row = con.execute(
                """
                SELECT d.*, u.name AS uploaded_by_name, s.title AS stage_title
                FROM documents d
                LEFT JOIN users u ON u.id = d.uploaded_by
                LEFT JOIN work_stages s ON s.id = d.stage_id
                WHERE d.id = ?
                """,
                (cur.lastrowid,),
            ).fetchone()
            con.commit()

        item = dict(row)
        item["download_url"] = f"/api/documents/{row['id']}/download"
        item["view_url"] = f"/api/documents/{row['id']}/view"
        item["can_preview"] = str(item["mime_type"] or "").startswith(("image/", "text/")) or str(item["file_ext"] or "") == ".pdf"
        self.send_json(HTTPStatus.CREATED, {"document": item})

    def api_document_file(self, path: str, inline: bool) -> None:
        document_id = self.parse_document_id(path)
        if not document_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_document_id"})
            return
        user = self.require_user()
        if not user:
            return
        with db() as con:
            row = con.execute("SELECT * FROM documents WHERE id = ?", (document_id,)).fetchone()
        if not row:
            self.send_json(HTTPStatus.NOT_FOUND, {"error": "document_not_found"})
            return
        if not self.can_access_project(user, int(row["project_id"])):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "document_forbidden"})
            return
        if user["role"] == "customer" and int(row["is_client_visible"] or 0) != 1:
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "document_forbidden"})
            return
        storage_path = row["storage_path"]
        if not storage_path:
            self.send_json(HTTPStatus.NOT_FOUND, {"error": "file_not_uploaded"})
            return
        file_path = (PROJECT_ROOT / storage_path).resolve()
        try:
            file_path.relative_to(PROJECT_ROOT.resolve())
        except ValueError:
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "document_forbidden"})
            return
        if not file_path.is_file():
            self.send_json(HTTPStatus.NOT_FOUND, {"error": "file_missing"})
            return
        can_inline = str(row["mime_type"] or "").startswith(("image/", "text/")) or str(row["file_ext"] or "") == ".pdf"
        self.send_file(
            file_path,
            str(row["mime_type"] or "application/octet-stream"),
            str(row["original_name"] or row["title"] or f"document-{document_id}"),
            inline=inline and can_inline,
        )

    def api_project_daily_logs(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        with db() as con:
            if user["role"] == "customer":
                rows = con.execute(
                    """
                    SELECT l.*, u.name AS author_name
                    FROM daily_logs l
                    LEFT JOIN users u ON u.id = l.created_by
                    WHERE l.project_id = ? AND l.is_client_visible = 1
                    ORDER BY l.report_date DESC, l.id DESC
                    """,
                    (project_id,),
                ).fetchall()
            else:
                rows = con.execute(
                    """
                    SELECT l.*, u.name AS author_name
                    FROM daily_logs l
                    LEFT JOIN users u ON u.id = l.created_by
                    WHERE l.project_id = ?
                    ORDER BY l.report_date DESC, l.id DESC
                    """,
                    (project_id,),
                ).fetchall()
        self.send_json(HTTPStatus.OK, {"logs": [dict(row) for row in rows]})

    def api_create_daily_log(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if user["role"] in {"customer"}:
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        payload = self.read_json()
        title = str(payload.get("title", "")).strip()
        work_done = str(payload.get("work_done", "")).strip()
        if not title or not work_done:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "title_and_work_required"})
            return
        report_date = str(payload.get("report_date", "")).strip() or time.strftime("%Y-%m-%d")
        try:
            workers_count = max(0, int(payload.get("workers_count", 0) or 0))
        except (TypeError, ValueError):
            workers_count = 0
        progress_percent_raw = payload.get("progress_percent", payload.get("progressPercent"))
        try:
            progress_percent = max(0.0, min(100.0, float(progress_percent_raw))) if progress_percent_raw not in (None, "", False) else None
        except (TypeError, ValueError):
            progress_percent = None
        with db() as con:
            cur = con.execute(
                """
                INSERT INTO daily_logs (
                    project_id, report_date, title, work_done, workers_count, equipment, blockers, next_steps,
                    progress_percent, raw_input, is_client_visible, created_by, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    report_date,
                    title,
                    work_done,
                    workers_count,
                    str(payload.get("equipment", "")).strip(),
                    str(payload.get("blockers", "")).strip(),
                    str(payload.get("next_steps", "")).strip(),
                    progress_percent,
                    str(payload.get("raw_input", payload.get("rawInput", ""))).strip() or None,
                    1 if payload.get("is_client_visible", True) else 0,
                    user["id"],
                    now_ts(),
                    now_ts(),
                ),
            )
            if progress_percent is not None:
                status = "completed" if progress_percent >= 100 else ("active" if progress_percent > 0 else None)
                if status:
                    con.execute(
                        "UPDATE projects SET progress = ?, status = ?, updated_at = ? WHERE id = ?",
                        (int(round(progress_percent)), status, now_ts(), project_id),
                    )
                else:
                    con.execute(
                        "UPDATE projects SET progress = ?, updated_at = ? WHERE id = ?",
                        (int(round(progress_percent)), now_ts(), project_id),
                    )
            con.execute(
                """
                INSERT INTO audit_log (user_id, action, entity, entity_id, payload, created_at)
                VALUES (?, 'create_daily_log', 'daily_log', ?, ?, ?)
                """,
                (user["id"], cur.lastrowid, json.dumps({"project_id": project_id, "title": title, "progress_percent": progress_percent}, ensure_ascii=False), now_ts()),
            )
            log_row = con.execute(
                """
                SELECT l.*, u.name AS author_name
                FROM daily_logs l
                LEFT JOIN users u ON u.id = l.created_by
                WHERE l.id = ?
                """,
                (cur.lastrowid,),
            ).fetchone()
            project_row = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            con.commit()
        self.send_json(HTTPStatus.CREATED, {"id": cur.lastrowid, "log": dict(log_row), "project": serialize_project(project_row, user)})

    def api_delete_daily_log(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        log_id = self.parse_daily_log_id(path)
        if not project_id or not log_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_log_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if user["role"] in {"customer"}:
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        with db() as con:
            log_row = con.execute(
                "SELECT * FROM daily_logs WHERE id = ? AND project_id = ?",
                (log_id, project_id),
            ).fetchone()
            if not log_row:
                self.send_json(HTTPStatus.NOT_FOUND, {"error": "log_not_found"})
                return
            project_row = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            con.execute("DELETE FROM daily_logs WHERE id = ?", (log_id,))
            latest_progress_row = con.execute(
                """
                SELECT progress_percent
                FROM daily_logs
                WHERE project_id = ? AND progress_percent IS NOT NULL
                ORDER BY report_date DESC, id DESC
                LIMIT 1
                """,
                (project_id,),
            ).fetchone()
            if latest_progress_row and latest_progress_row["progress_percent"] is not None:
                progress_value = int(round(float(latest_progress_row["progress_percent"] or 0)))
                status_value = "completed" if progress_value >= 100 else ("active" if progress_value > 0 else None)
                if status_value:
                    con.execute(
                        "UPDATE projects SET progress = ?, status = ?, updated_at = ? WHERE id = ?",
                        (progress_value, status_value, now_ts(), project_id),
                    )
                else:
                    con.execute(
                        "UPDATE projects SET progress = ?, updated_at = ? WHERE id = ?",
                        (progress_value, now_ts(), project_id),
                    )
            else:
                current_status = str(project_row["status"] or "") if project_row else ""
                fallback_status = "Подготовка" if current_status in {"active", "completed"} else current_status
                con.execute(
                    "UPDATE projects SET progress = 0, status = ?, updated_at = ? WHERE id = ?",
                    (fallback_status or "Подготовка", now_ts(), project_id),
                )
            con.execute(
                """
                INSERT INTO audit_log (user_id, action, entity, entity_id, payload, created_at)
                VALUES (?, 'delete_daily_log', 'daily_log', ?, ?, ?)
                """,
                (
                    user["id"],
                    log_id,
                    json.dumps(
                        {
                            "project_id": project_id,
                            "title": log_row["title"],
                            "report_date": log_row["report_date"],
                        },
                        ensure_ascii=False,
                    ),
                    now_ts(),
                ),
            )
            refreshed_project = con.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            con.commit()
        self.send_json(
            HTTPStatus.OK,
            {
                "ok": True,
                "deletedId": log_id,
                "project": serialize_project(refreshed_project, user) if refreshed_project else None,
            },
        )

    def api_project_chats(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        with db() as con:
            if user["role"] == "customer":
                rows = con.execute(
                    "SELECT * FROM chats WHERE project_id = ? AND chat_type = 'client' ORDER BY id",
                    (project_id,),
                ).fetchall()
            else:
                rows = con.execute(
                    "SELECT * FROM chats WHERE project_id = ? ORDER BY CASE chat_type WHEN 'team' THEN 0 ELSE 1 END, id",
                    (project_id,),
                ).fetchall()
        self.send_json(HTTPStatus.OK, {"chats": [dict(row) for row in rows]})

    def can_access_chat(self, user: dict, chat_id: int) -> bool:
        with db() as con:
            row = con.execute(
                """
                SELECT c.project_id, c.chat_type
                FROM chats c
                WHERE c.id = ?
                """,
                (chat_id,),
            ).fetchone()
        if not row:
            return False
        if not self.can_access_project(user, int(row["project_id"])):
            return False
        return not (user["role"] == "customer" and row["chat_type"] != "client")

    def api_chat_messages(self, path: str) -> None:
        chat_id = self.parse_chat_id(path)
        if not chat_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_chat_id"})
            return
        user = self.require_user()
        if not user:
            return
        if not self.can_access_chat(user, chat_id):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "chat_forbidden"})
            return
        with db() as con:
            rows = con.execute(
                """
                SELECT m.*, COALESCE(u.name, 'Система') AS author_name, COALESCE(u.role, 'system') AS author_role
                FROM chat_messages m
                LEFT JOIN users u ON u.id = m.user_id
                WHERE m.chat_id = ?
                ORDER BY m.id
                """,
                (chat_id,),
            ).fetchall()
        self.send_json(HTTPStatus.OK, {"messages": [dict(row) for row in rows]})

    def api_create_chat_message(self, path: str) -> None:
        chat_id = self.parse_chat_id(path)
        if not chat_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_chat_id"})
            return
        user = self.require_user()
        if not user:
            return
        if not self.can_access_chat(user, chat_id):
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "chat_forbidden"})
            return
        payload = self.read_json()
        body = str(payload.get("body", "")).strip()
        if not body:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "body_required"})
            return
        with db() as con:
            cur = con.execute(
                "INSERT INTO chat_messages (chat_id, user_id, body, created_at) VALUES (?, ?, ?, ?)",
                (chat_id, user["id"], body, now_ts()),
            )
            con.commit()
        self.send_json(HTTPStatus.CREATED, {"id": cur.lastrowid})

    def api_create_stock_move(self, path: str) -> None:
        project_id = self.parse_project_id(path)
        if not project_id:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_project_id"})
            return
        user = self.require_project_access(project_id)
        if not user:
            return
        if user["role"] == "customer":
            self.send_json(HTTPStatus.FORBIDDEN, {"error": "forbidden"})
            return
        payload = self.read_json()
        move_type = str(payload.get("move_type", "")).strip()
        if move_type not in {"purchase", "receipt", "use", "writeoff"}:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_move_type"})
            return
        qty = float(payload.get("qty", 0) or 0)
        if qty <= 0:
            self.send_json(HTTPStatus.BAD_REQUEST, {"error": "bad_qty"})
            return
        with db() as con:
            cur = con.execute(
                """
                INSERT INTO stock_moves (project_id, estimate_item_id, move_type, qty, price, comment, created_by, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    payload.get("estimate_item_id"),
                    move_type,
                    qty,
                    float(payload.get("price", 0) or 0),
                    str(payload.get("comment", "")).strip(),
                    user["id"],
                    now_ts(),
                ),
            )
            con.commit()
        self.send_json(HTTPStatus.CREATED, {"id": cur.lastrowid})

    def render_template(self, template_name: str, variables: dict[str, str]) -> bytes:
        template = (FRONTEND_TEMPLATES / template_name).read_text(encoding="utf-8")
        for key, value in variables.items():
            template = template.replace("{{" + key + "}}", value)
        return template.encode("utf-8")

    def project_cards_fallback_html(self, user: dict) -> str:
        with db() as con:
            if user_has_any_role(user, {"admin", "director", "foreman"}):
                rows = con.execute("SELECT * FROM projects ORDER BY id DESC").fetchall()
            else:
                rows = con.execute(
                    """
                    SELECT p.*
                    FROM projects p
                    JOIN user_project_access a ON a.project_id = p.id
                    WHERE a.user_id = ?
                    ORDER BY p.id DESC
                    """,
                    (user["id"],),
                ).fetchall()

        if not rows:
            return '<div class="muted">Объекты пока не найдены.</div>'

        def esc(value: object) -> str:
            return html.escape("" if value is None else str(value), quote=True)

        def project_money(value: object) -> str:
            if value is None:
                return "Скрыто"
            try:
                amount = float(value or 0)
            except (TypeError, ValueError):
                return esc(value)
            formatted = f"{amount:,.0f}".replace(",", " ")
            return f"{formatted} ₽"

        def project_progress(value: object) -> int:
            try:
                return max(0, min(100, int(round(float(value or 0)))))
            except (TypeError, ValueError):
                return 0

        def is_completed(row: sqlite3.Row) -> bool:
            status = str(row["status"] or "").lower()
            return project_progress(row["progress"]) >= 100 or "сдан" in status or "заверш" in status

        sorted_rows = sorted(rows, key=lambda row: (1 if is_completed(row) else 0, -int(row["id"])))
        cards: list[str] = []
        for row in sorted_rows:
            progress = project_progress(row["progress"])
            completed = is_completed(row)
            status_badge = (
                '<span class="badge success">Завершен</span>'
                if completed
                else f'<span class="badge">{esc(row["status"] or "В работе")}</span>'
            )
            cards.append(
                f'<article class="project-card {"project-completed" if completed else ""}" data-project-id="{esc(row["id"])}">'
                '<div class="project-top">'
                f'<div><h3>{esc(row["title"])}</h3><p>{esc(row["address"] or "Адрес не указан")}</p></div>'
                f'<div class="project-card-tools"><div class="project-badges">{status_badge}</div></div>'
                '</div>'
                '<div class="meta-grid">'
                f'<div><span>Заказчик</span><strong>{esc(row["client_name"] or "Не указан")}</strong></div>'
                f'<div><span>Бюджет</span><strong>{project_money(row["budget"])}</strong></div>'
                f'<div><span>Дедлайн</span><strong>{esc(row["deadline_at"] or "—")}</strong></div>'
                '</div>'
                '<div class="progress-strong">'
                f'<div class="progress-strong-head"><span>Готовность объекта</span><strong>{progress}%</strong></div>'
                f'<div class="progress-strong-track"><i style="width:{progress}%"></i></div>'
                '</div>'
                '</article>'
            )
        return "".join(cards)

    def strip_role_content(self, content: str, user: dict) -> str:
        if not user_has_any_role(user, {"admin", "director"}):
            content = DIRECTOR_ACTION_BLOCK_RE.sub("", content)
        return content

    def send_html(self, body: bytes, status: int = HTTPStatus.OK) -> None:
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "DENY")
        self.send_header("Referrer-Policy", "no-referrer")
        self.end_headers()
        self.wfile.write(body)

    def serve_app(self, path: str) -> None:
        if path == "/login":
            self.send_html(
                self.render_template(
                    "login.html",
                    {
                        "auth_config": self.auth_config_script(),
                        "auth_head": self.clerk_sign_in_script(),
                    },
                )
            )
            return

        user = self.current_user()
        if not user:
            self.redirect(f"{LOGIN_PATH}?next={urllib.parse.quote(path, safe='/')}")
            return
        if not user_can_open(user, path):
            self.redirect(DEFAULT_AUTH_PATH + "?restricted=1")
            return

        page_info = APP_PAGES.get(path)
        if not page_info:
            self.send_error(HTTPStatus.NOT_FOUND, "Not found")
            return
        page_name, title = page_info
        content_path = FRONTEND_PAGES / f"{page_name}.html"
        if not content_path.is_file():
            self.send_error(HTTPStatus.NOT_FOUND, "Page not found")
            return

        content = self.strip_role_content(content_path.read_text(encoding="utf-8"), user)
        if page_name == "autobot":
            content = content.replace(
                "{{autobot_url}}",
                html.escape(PMBI_AUTOBOT_BASE_URL or "http://127.0.0.1:8765", quote=True),
            )
        if page_name == "projects":
            content = content.replace(
                '<div class="projects-grid" data-projects-list></div>',
                f'<div class="projects-grid" data-projects-list>{self.project_cards_fallback_html(user)}</div>',
            )
        body = self.render_template(
            "base.html",
            {
                "title": title,
                "page": page_name,
                "auth_config": self.auth_config_script(),
                "auth_head": self.clerk_sign_in_script(),
                "content": content,
            },
        )
        self.send_html(body)

    def serve_asset(self, path: str) -> None:
        relative = path.removeprefix("/assets/").lstrip("/")
        candidate = (FRONTEND_ASSETS / relative).resolve()
        try:
            candidate.relative_to(FRONTEND_ASSETS.resolve())
        except ValueError:
            self.send_error(HTTPStatus.FORBIDDEN, "Forbidden")
            return
        if not candidate.is_file():
            self.send_error(HTTPStatus.NOT_FOUND, "Not found")
            return
        content_type = mimetypes.guess_type(str(candidate))[0] or "application/octet-stream"
        if candidate.suffix.lower() in {".html", ".js", ".css", ".json", ".txt"}:
            content_type += "; charset=utf-8"
        body = candidate.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.end_headers()
        self.wfile.write(body)

    def serve_static(self, path: str) -> None:
        user = self.current_user()
        is_page_request = path.endswith(".html") or "." not in Path(path).name
        if path not in PUBLIC_STATIC_PATHS and not user:
            if is_page_request:
                next_path = urllib.parse.quote(path, safe="/")
                self.redirect(f"{LOGIN_PATH}?next={next_path}")
            else:
                self.send_error(HTTPStatus.UNAUTHORIZED, "Authentication required")
            return

        if user and is_page_request and not user_can_open(user, path):
            if path.endswith(".html") or "." not in Path(path).name:
                self.redirect(DEFAULT_AUTH_PATH + "?restricted=1")
            else:
                self.send_error(HTTPStatus.FORBIDDEN, "Forbidden")
            return

        file_path = self.resolve_static_path(path)
        if not file_path:
            self.send_error(HTTPStatus.NOT_FOUND, "Not found")
            return

        content_type = mimetypes.guess_type(str(file_path))[0] or "application/octet-stream"
        if file_path.suffix.lower() in {".html", ".js", ".css", ".json", ".txt"}:
            content_type += "; charset=utf-8"
        body = file_path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "DENY")
        self.send_header("Referrer-Policy", "no-referrer")
        self.end_headers()
        self.wfile.write(body)

    def resolve_static_path(self, path: str) -> Path | None:
        rel = path.lstrip("/") or "index.html"
        candidate = (DEPLOY_ROOT / rel).resolve()
        try:
            candidate.relative_to(DEPLOY_ROOT.resolve())
        except ValueError:
            return None
        if candidate.is_dir():
            candidate = candidate / "index.html"
        if not candidate.is_file():
            return None
        return candidate


def main() -> None:
    init_db()
    print(f"PM.bi backend: http://{HOST}:{PORT}/")
    print(f"Database: {DB_PATH}")
    if BOOTSTRAP_PATH.exists():
        print(f"Initial admin credentials: {BOOTSTRAP_PATH}")
    print("Press Ctrl+C to stop the server.")
    ThreadingHTTPServer((HOST, PORT), PMBIHandler).serve_forever()


def resolved_estimate_item_kind(item: dict | sqlite3.Row) -> str:
    def value(*keys: str) -> object:
        if isinstance(item, sqlite3.Row):
            for key in keys:
                if key in item.keys():
                    return item[key]
            return None
        for key in keys:
            if key in item:
                return item.get(key)
        return None

    note_value = (
        extract_labeled_note_value(value("notes"), ("\u0422\u0438\u043f", "Type"))
        or extract_labeled_note_value(value("notes"), ("Type",))
    )
    if note_value:
        return normalize_estimate_item_kind(note_value)
    raw_value = (
        value("item_kind")
        or value("itemKind")
        or value("type")
        or value("type_label")
        or value("typeLabel")
        or value("position_type")
        or value("positionType")
        or ""
    )
    raw_text = str(raw_value or "").strip()
    if raw_text:
        return normalize_estimate_item_kind(raw_text)
    return normalize_estimate_item_kind(note_value)


def resolved_estimate_section_title(item: dict | sqlite3.Row) -> str | None:
    def value(*keys: str) -> object:
        if isinstance(item, sqlite3.Row):
            for key in keys:
                if key in item.keys():
                    return item[key]
            return None
        for key in keys:
            if key in item:
                return item.get(key)
        return None

    def normalize_section_title_text(value: object) -> str | None:
        text = re.sub(r"\s+", " ", str(value or "").strip())
        if not text:
            return None
        duplicate_match = re.fullmatch(r"(.+?)\s+\1", text)
        if duplicate_match:
            text = duplicate_match.group(1).strip()
        return text or None

    direct_value = (
        value("section_title")
        or value("sectionTitle")
        or value("section_name")
        or value("sectionName")
        or value("section")
        or value("chapter")
        or ""
    )
    direct_text = normalize_section_title_text(direct_value)
    if direct_text:
        return direct_text
    note_text = (
        extract_labeled_note_value(value("notes"), ("\u0420\u0430\u0437\u0434\u0435\u043b", "Section", "Chapter"))
        or extract_labeled_note_value(value("notes"), ("Section", "Chapter"))
    )
    return normalize_section_title_text(note_text)


if __name__ == "__main__":
    main()
