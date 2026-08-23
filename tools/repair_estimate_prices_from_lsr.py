from __future__ import annotations

import argparse
import json
import math
import os
import re
import shutil
import sqlite3
import tempfile
import time
from pathlib import Path

from openpyxl import load_workbook


POSITION_TOTAL_LABELS = ("всего по позиции", "итог по позиции", "всего затрат по позиции")


def _text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _position_number(value: object) -> int | None:
    text = _text(value)
    return int(text) if re.fullmatch(r"\d+", text) else None


def _positive_number(value: object) -> float | None:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    number = float(value)
    return number if math.isfinite(number) and number > 0 else None


def read_lsr_positions(path: Path) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows = list(worksheet.iter_rows(min_col=1, max_col=16, values_only=True))
    finally:
        workbook.close()

    positions: list[dict] = []
    section = ""
    for row_index, row in enumerate(rows):
        first = _text(row[0])
        if "раздел" in first.casefold():
            section = first
        item_no = _position_number(row[0])
        if item_no is None:
            continue
        basis_code = _text(row[1])
        title = _text(row[2])
        if len(title) < 4 or not basis_code or re.fullmatch(r"\d+", basis_code):
            continue
        positions.append(
            {
                "item_no": item_no,
                "basis_code": basis_code,
                "title": title,
                "unit": _text(row[7]),
                "qty": _positive_number(row[8]),
                "unit_price": _positive_number(row[13]),
                "total": _positive_number(row[15]),
                "section": section,
                "row_index": row_index,
            }
        )

    for index, position in enumerate(positions):
        end = positions[index + 1]["row_index"] if index + 1 < len(positions) else len(rows)
        for row in rows[position["row_index"] + 1 : end]:
            label = _text(row[2]).casefold()
            if not any(marker in label for marker in POSITION_TOTAL_LABELS):
                continue
            position["unit_price"] = position["unit_price"] or _positive_number(row[13])
            position["total"] = position["total"] or _positive_number(row[15])
            break

    for position in positions:
        qty = position["qty"]
        price = position["unit_price"]
        total = position["total"]
        if not position["basis_code"] or not position["title"] or qty is None or price is None or total is None:
            raise ValueError(f"Incomplete LSR position: {position}")
        calculated = qty * price
        if abs(calculated - total) > max(0.11, abs(total) * 0.001):
            raise ValueError(
                f"LSR position {position['item_no']} does not balance: "
                f"{qty} × {price} != {total}"
            )
    return positions


def _same_position(left: dict, right: dict) -> bool:
    return (
        _text(left.get("basis_code") or left.get("article")).casefold()
        == _text(right.get("basis_code") or right.get("article")).casefold()
        and _text(left.get("title") or left.get("name")).casefold()
        == _text(right.get("title") or right.get("name")).casefold()
    )


def inspect_database(database: Path, project_id: int, positions: list[dict]) -> tuple[list[dict], list[dict]]:
    connection = sqlite3.connect(database, timeout=30)
    connection.row_factory = sqlite3.Row
    try:
        project = connection.execute(
            "SELECT id, title, contract_no FROM projects WHERE id = ?",
            (project_id,),
        ).fetchone()
        if not project:
            raise ValueError(f"Project {project_id} not found")
        rows = [
            dict(row)
            for row in connection.execute(
                """
                SELECT id, title, unit, planned_qty, planned_price, article
                FROM estimate_items
                WHERE project_id = ?
                ORDER BY id
                """,
                (project_id,),
            )
        ]
    finally:
        connection.close()

    if len(rows) != len(positions):
        raise ValueError(f"Position count mismatch: CRM={len(rows)}, XLSX={len(positions)}")
    for crm_row, source_row in zip(rows, positions):
        if not _same_position(crm_row, source_row):
            raise ValueError(
                f"Position order mismatch at CRM id {crm_row['id']}: "
                f"{crm_row['article']} / {crm_row['title']}"
            )
    changes = [
        {
            "id": crm_row["id"],
            "item_no": source_row["item_no"],
            "old_qty": float(crm_row["planned_qty"] or 0),
            "new_qty": source_row["qty"],
            "old_price": float(crm_row["planned_price"] or 0),
            "new_price": source_row["unit_price"],
        }
        for crm_row, source_row in zip(rows, positions)
        if not math.isclose(float(crm_row["planned_qty"] or 0), source_row["qty"], rel_tol=1e-9, abs_tol=1e-9)
        or not math.isclose(float(crm_row["planned_price"] or 0), source_row["unit_price"], rel_tol=1e-9, abs_tol=0.005)
    ]
    return rows, changes


def _backup_database(database: Path, suffix: str) -> Path:
    backup = database.with_name(f"{database.name}.before-estimate-price-repair-{suffix}.bak")
    source = sqlite3.connect(database, timeout=30)
    target = sqlite3.connect(backup)
    try:
        source.backup(target)
    finally:
        target.close()
        source.close()
    return backup


def repair_database(database: Path, project_id: int, positions: list[dict], suffix: str) -> Path:
    backup = _backup_database(database, suffix)
    connection = sqlite3.connect(database, timeout=30)
    try:
        connection.execute("BEGIN IMMEDIATE")
        updated_at = int(time.time())
        for position, row in zip(
            positions,
            connection.execute(
                "SELECT id FROM estimate_items WHERE project_id = ? ORDER BY id",
                (project_id,),
            ).fetchall(),
        ):
            connection.execute(
                """
                UPDATE estimate_items
                SET planned_qty = ?, planned_price = ?, updated_at = ?
                WHERE id = ? AND project_id = ?
                """,
                (position["qty"], position["unit_price"], updated_at, int(row[0]), project_id),
            )
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()
    return backup


def inspect_autobot_rows(path: Path, positions: list[dict]) -> tuple[list[dict], int]:
    rows = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(rows, list) or len(rows) != len(positions):
        raise ValueError(f"AutoBot position count mismatch: rows={len(rows)}, XLSX={len(positions)}")
    changed = 0
    for row, position in zip(rows, positions):
        if not isinstance(row, dict) or not _same_position(row, position):
            raise ValueError(f"AutoBot position order mismatch at item {position['item_no']}")
        desired = {
            "qty": position["qty"],
            "unit_price": position["unit_price"],
            "total": position["total"],
            "unit": position["unit"],
            "section": position["section"],
        }
        if any(row.get(key) != value for key, value in desired.items()):
            changed += 1
        row.update(desired)
    return rows, changed


def repair_autobot_rows(path: Path, rows: list[dict], suffix: str) -> Path:
    backup = path.with_name(f"rows.before-estimate-price-repair-{suffix}.json")
    shutil.copy2(path, backup)
    fd, temp_name = tempfile.mkstemp(prefix="rows.", suffix=".json", dir=path.parent)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="\n") as handle:
            json.dump(rows, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
        os.replace(temp_name, path)
    except Exception:
        try:
            os.unlink(temp_name)
        except OSError:
            pass
        raise
    return backup


def main() -> None:
    parser = argparse.ArgumentParser(description="Repair CRM/AutoBot prices from explicit LSR columns I/N/P")
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--database", type=Path, required=True)
    parser.add_argument("--project-id", type=int, required=True)
    parser.add_argument("--autobot-rows", type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    positions = read_lsr_positions(args.xlsx)
    _, database_changes = inspect_database(args.database, args.project_id, positions)
    autobot_rows = None
    autobot_changes = 0
    if args.autobot_rows:
        autobot_rows, autobot_changes = inspect_autobot_rows(args.autobot_rows, positions)

    result = {
        "mode": "apply" if args.apply else "dry-run",
        "positions": len(positions),
        "database_changes": len(database_changes),
        "database_zero_prices_before": sum(1 for row in database_changes if row["old_price"] <= 0),
        "autobot_changes": autobot_changes,
        "database_backup": None,
        "autobot_backup": None,
    }
    if args.apply:
        suffix = time.strftime("%Y%m%d-%H%M%S")
        result["database_backup"] = str(repair_database(args.database, args.project_id, positions, suffix))
        if args.autobot_rows and autobot_rows is not None:
            result["autobot_backup"] = str(repair_autobot_rows(args.autobot_rows, autobot_rows, suffix))

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
